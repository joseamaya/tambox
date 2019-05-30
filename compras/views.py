# -*- coding: utf-8 -*- 
from django.views.generic.base import View, TemplateView
from django.views.generic.list import ListView
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Color
from openpyxl.styles import Font
from openpyxl.styles import Side

from compras.models import Proveedor, OrdenCompra, FormaPago, DetalleOrdenCompra, DetalleRequerimiento, OrdenServicios,\
    DetalleOrdenServicios, ConformidadServicio,\
    DetalleConformidadServicio, DetalleCotizacion, Cotizacion
from django.views.generic.edit import FormView, UpdateView, CreateView
from compras.forms import ProveedorForm, DetalleCotizacionForm, CotizacionForm, OrdenCompraForm,\
    OrdenServiciosForm, ConformidadServicioForm, DetalleOrdenCompraFormSet,\
    DetalleOrdenServiciosFormSet, DetalleConformidadServicioFormSet, DetalleCotizacionFormSet,\
    FormularioReporteOrdenesFecha
from django.core.urlresolvers import reverse_lazy, reverse
from django.http.response import HttpResponseRedirect
import json
from django.http import HttpResponse
import datetime
import simplejson
from openpyxl import Workbook
from django.views.generic.detail import DetailView
from django.conf import settings
from io import BytesIO
from reportlab.platypus import Paragraph, TableStyle
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import Table
from reportlab.platypus.flowables import ListFlowable
from reportlab.pdfgen import canvas
#from reportlab.lib.pagesizes import cm
from reportlab.lib.enums import TA_JUSTIFY
from administracion.models import Puesto
import csv
import locale
from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
import os
from django.db import transaction, IntegrityError
from contabilidad.forms import UploadForm
from django.db.models import Q
from django.contrib import messages
from almacen.forms import DetalleIngresoFormSet
from django.shortcuts import render, get_object_or_404

from contabilidad.models import TipoCambio
from productos.models import Producto, UnidadMedida, GrupoProductos
from django.utils.encoding import smart_str
from datetime import date
from compras.reports import ReporteOrdenCompra
from compras.settings import EMPRESA, CONFIGURACION, IMPUESTO_COMPRA
from decimal import Decimal

locale.setlocale(locale.LC_ALL,"")

class Tablero(View):
    
    def get(self, request, *args, **kwargs):
        lista_notificaciones = []
        cant_proveedores = Proveedor.objects.count()        
        cant_productos = Producto.objects.filter(es_servicio=False).count()
        cant_tipos_unidad_medida = UnidadMedida.objects.count()
        cant_grupos_suministros = GrupoProductos.objects.count()
        cant_servicios = Producto.objects.filter(es_servicio=True).count()
        unidad_medida, creado = UnidadMedida.objects.get_or_create(codigo = 'SERV',
                                                                   defaults = {'descripcion':'SERVICIO'})
        if cant_proveedores == 0:
            lista_notificaciones.append("No se ha creado ningún proveedor")
        if creado:
            lista_notificaciones.append("Se ha creado la unidad de medida SERVICIO")
        if cant_productos == 0:
            lista_notificaciones.append("No se ha creado ningún producto")
        if cant_tipos_unidad_medida == 0:
            lista_notificaciones.append("No se ha creado ningún tipo de unidad de medida")
        if cant_grupos_suministros == 0:
            lista_notificaciones.append("No se ha creado ningún grupo de productos")
        if cant_servicios == 0:
            lista_notificaciones.append("No se ha creado ningún servicio")
        context = {'notificaciones':lista_notificaciones}
        return render(request, 'compras/tablero_compras.html', context)
    
class BusquedaCotizacion(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            cotizacion = Cotizacion.objects.get(codigo=codigo)            
            cotizacion_json = {}
            cotizacion_json['ruc'] = cotizacion.proveedor.ruc
            cotizacion_json['razon_social'] = cotizacion.proveedor.razon_social
            cotizacion_json['direccion'] = cotizacion.proveedor.direccion
            data = simplejson.dumps(cotizacion_json)
            return HttpResponse(data, 'application/json')
    
class BusquedaProveedoresRazonSocial(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            razon_social = request.GET['razon_social']            
            proveedores = Proveedor.objects.filter(razon_social__icontains=razon_social)[:20]
            lista_proveedores = []
            for proveedor in proveedores:
                proveedor_json = {}
                proveedor_json['label'] = proveedor.razon_social
                proveedor_json['ruc'] = proveedor.ruc
                proveedor_json['direccion'] = proveedor.direccion
                proveedor_json['orden'] = str(OrdenServicios.objects.ultimo())
                lista_proveedores.append(proveedor_json)
            data = json.dumps(lista_proveedores)
            return HttpResponse(data, 'application/json')
        
class BusquedaProveedoresRUC(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            ruc = request.GET['ruc']
            proveedor = Proveedor.objects.get(ruc=ruc)
            proveedor_json = {}
            proveedor_json['razon_social'] = proveedor.razon_social
            proveedor_json['direccion'] = proveedor.direccion
            proveedor_json['estado'] = proveedor.estado_sunat
            proveedor_json['es_locador'] = proveedor.es_locador
            proveedor_json['orden'] = str(OrdenServicios.objects.ultimo())
            data = simplejson.dumps(proveedor_json)
            return HttpResponse(data, 'application/json')            
        
class CargarProveedores(FormView):
    template_name = 'compras/cargar_proveedores.html'
    form_class = UploadForm
    
    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']            
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT,'archivos',str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        for fila in dataReader:
            Proveedor.objects.get_or_create(ruc=fila[0],
                                            defaults={'razon_social': fila[1],
                                                      'direccion': fila[2],
                                                      'fecha_alta' : datetime.datetime.now(),
                                                      'estado_sunat' : 'ACTIVO',
                                                      'condicion' : 'HABIDO',
                                                      'ciiu' : 'CUALQUIERA'})            
        return HttpResponseRedirect(reverse('compras:proveedores'))
    
class CrearProveedor(CreateView):
    model = Proveedor
    template_name = 'compras/proveedor.html'
    form_class = ProveedorForm
    
    @method_decorator(permission_required('compras.add_proveedor',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearProveedor, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('compras:detalle_proveedor', args=[self.object.pk])
    
    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

class CrearDetalleCotizacion(FormView):
    template_name = 'compras/crear_detalle_cotizacion.html'
    form_class = DetalleCotizacionForm
    
class CrearDetalleOrdenCompra(TemplateView):
        
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            lista_detalles = []            
            det = {}
            det['cotizacion'] = '0'
            det['codigo'] = ''               
            det['nombre'] = ''                    
            det['unidad'] = ''
            det['cantidad'] = '0'
            det['precio'] = '0'
            det['impuesto'] = '0'            
            det['valor'] = '0'
            lista_detalles.append(det)
            formset = DetalleOrdenCompraFormSet(initial=lista_detalles)
            lista_json = []
            for form in formset:
                detalle_json = {}    
                detalle_json['cotizacion'] = str(form['cotizacion'])
                detalle_json['codigo'] = str(form['codigo'])
                detalle_json['nombre'] = str(form['nombre'])
                detalle_json['unidad'] = str(form['unidad'])
                detalle_json['cantidad'] = str(form['cantidad'])
                detalle_json['precio'] = str(form['precio'])
                detalle_json['impuesto'] = str(form['impuesto'])                
                detalle_json['valor'] = str(form['valor'])
                lista_json.append(detalle_json)                                
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')
    
class CrearDetalleOrdenServicios(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            lista_detalles = []            
            det = {}
            det['cotizacion'] = '0'
            det['codigo'] = ''               
            det['nombre'] = ''                    
            det['unidad'] = ''
            det['cantidad'] = '0'
            det['precio'] = '0'
            det['valor'] = '0'
            lista_detalles.append(det)
            formset = DetalleOrdenServiciosFormSet(initial = lista_detalles)
            lista_json = []
            for form in formset:
                detalle_json = {}    
                detalle_json['cotizacion'] = str(form['cotizacion'])
                detalle_json['codigo'] = str(form['codigo'])
                detalle_json['nombre'] = str(form['nombre'])
                detalle_json['unidad'] = str(form['unidad'])
                detalle_json['cantidad'] = str(form['cantidad'])
                detalle_json['precio'] = str(form['precio'])
                detalle_json['valor'] = str(form['valor'])
                lista_json.append(detalle_json)                                
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')
    
class CrearCotizacion(CreateView):
    form_class = CotizacionForm
    template_name = "compras/cotizacion.html"
    model = Cotizacion
    
    def get_initial(self):
        initial = super(CrearCotizacion, self).get_initial()
        initial['fecha'] = date.today().strftime('%d/%m/%Y')
        return initial
    
    def get(self, request, *args, **kwargs):
        self.object = None
        proveedores = Proveedor.objects.all()        
        if not proveedores:
            return HttpResponseRedirect(reverse('compras:crear_proveedor'))
        else:
            form_class = self.get_form_class()
            form = self.get_form(form_class)
            detalle_cotizacion_formset = DetalleCotizacionFormSet()
            return self.render_to_response(self.get_context_data(form = form,
                                                                 detalle_cotizacion_formset = detalle_cotizacion_formset))
        
    def post(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_cotizacion_formset = DetalleCotizacionFormSet(request.POST)
        if form.is_valid() and detalle_cotizacion_formset.is_valid():
            return self.form_valid(form, detalle_cotizacion_formset)
        else:
            return self.form_invalid(form, detalle_cotizacion_formset)
    
    def form_valid(self, form, detalle_cotizacion_formset):
        #try:
        with transaction.atomic():
            self.object = form.save()
            cod_orden = form.cleaned_data.get('orden')
            referencia = self.object.requerimiento
            detalles = []
            detalles_servicios = []
            cont = 1                
            for detalle_cotizacion_form in detalle_cotizacion_formset:
                requerimiento = detalle_cotizacion_form.cleaned_data.get('requerimiento')
                codigo = detalle_cotizacion_form.cleaned_data.get('codigo')
                nombre = detalle_cotizacion_form.cleaned_data.get('nombre')
                unidad = detalle_cotizacion_form.cleaned_data.get('unidad')
                cantidad = detalle_cotizacion_form.cleaned_data.get('cantidad')                    
                detalle_requerimiento = DetalleRequerimiento.objects.get(pk=requerimiento)
                if cantidad:
                    detalle_cotizacion = DetalleCotizacion(detalle_requerimiento = detalle_requerimiento,
                                                           nro_detalle = cont,
                                                           cotizacion = self.object,
                                                           cantidad = cantidad) 
                    detalles.append(detalle_cotizacion)
                    
                        
                    cont = cont + 1
            DetalleCotizacion.objects.bulk_create(detalles, referencia, None)
            return HttpResponseRedirect(reverse('compras:detalle_cotizacion', args=[self.object.codigo]))
        #except IntegrityError:
            #messages.error(self.request, 'Error guardando la cotizacion.')
        
    def form_invalid(self, form, detalle_cotizacion_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_cotizacion_formset = detalle_cotizacion_formset))

class CrearOrdenCompra(CreateView):
    form_class = OrdenCompraForm
    template_name = "compras/orden_compra.html"
    model = OrdenCompra
    
    def get_initial(self):
        initial = super(CrearOrdenCompra, self).get_initial()
        try:
            monto_impuesto = IMPUESTO_COMPRA.monto
        except:
            return HttpResponseRedirect(reverse('contabilidad:configuracion'))
        initial['fecha'] = date.today().strftime('%d/%m/%Y')
        initial['codigo'] = OrdenCompra.objects.ultimo()
        initial['impuesto_actual'] = monto_impuesto
        initial['total'] = 0
        initial['subtotal'] = 0
        initial['impuesto'] = 0
        initial['total_letras'] = ''
        return initial
    
    def get(self, request, *args, **kwargs):
        self.object = None
        formas_pago = FormaPago.objects.all().order_by('descripcion')        
        if not formas_pago:
            return HttpResponseRedirect(reverse('contabilidad:crear_forma_pago'))
        else:
            try:
                CONFIGURACION
                form_class = self.get_form_class()
                form = self.get_form(form_class)
                detalle_orden_compra_formset=DetalleOrdenCompraFormSet()
                return self.render_to_response(self.get_context_data(form=form,
                                                                     detalle_orden_compra_formset=detalle_orden_compra_formset))
            except:
                return HttpResponseRedirect(reverse('contabilidad:configuracion'))            
        
    def post(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_orden_compra_formset = DetalleOrdenCompraFormSet(request.POST)
        if form.is_valid() and detalle_orden_compra_formset.is_valid():
            return self.form_valid(form, detalle_orden_compra_formset)
        else:
            return self.form_invalid(form, detalle_orden_compra_formset)
    
    def form_valid(self, form, detalle_orden_compra_formset):
        try:
            with transaction.atomic():
                self.object = form.save()
                con_igv = form.cleaned_data.get('con_impuesto')
                referencia = self.object.cotizacion
                detalles = []
                cont = 1                
                for detalle_orden_compra_form in detalle_orden_compra_formset:
                    cotizacion = detalle_orden_compra_form.cleaned_data.get('cotizacion')
                    codigo = detalle_orden_compra_form.cleaned_data.get('codigo')
                    cantidad = detalle_orden_compra_form.cleaned_data.get('cantidad')
                    precio = detalle_orden_compra_form.cleaned_data.get('precio')
                    valor = detalle_orden_compra_form.cleaned_data.get('valor')
                    impuesto = detalle_orden_compra_form.cleaned_data.get('impuesto')
                    if cantidad and precio and valor and impuesto:
                        try:
                            detalle_cotizacion = DetalleCotizacion.objects.get(pk=cotizacion)
                            detalle_orden_compra = DetalleOrdenCompra(detalle_cotizacion = detalle_cotizacion,
                                                                      nro_detalle = cont,
                                                                      orden = self.object,
                                                                      cantidad = cantidad,
                                                                      precio = precio) 
                        except DetalleCotizacion.DoesNotExist:
                            producto = Producto.objects.get(pk = codigo)
                            detalle_orden_compra = DetalleOrdenCompra(producto = producto,
                                                                      nro_detalle = cont,
                                                                      orden = self.object,
                                                                      cantidad = cantidad,
                                                                      precio = precio)
                        detalles.append(detalle_orden_compra)                        
                        cont = cont + 1
                if cont>1:
                    DetalleOrdenCompra.objects.bulk_create(detalles,referencia)                
                return HttpResponseRedirect(reverse('compras:detalle_orden_compra', args=[self.object.pk]))
        except IntegrityError:
            messages.error(self.request, 'Error guardando la orden de compra.')
        
    def form_invalid(self, form, detalle_orden_compra_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_orden_compra_formset = detalle_orden_compra_formset))

class CrearOrdenServicios(CreateView):
    form_class = OrdenServiciosForm
    template_name = "compras/orden_servicio.html"
    model = OrdenServicios
    
    def get_initial(self):
        initial = super(CrearOrdenServicios, self).get_initial()
        initial['codigo'] = OrdenServicios.objects.ultimo()
        initial['total'] = 0
        initial['subtotal'] = 0
        initial['impuesto'] = 0
        initial['total_letras'] = ''
        return initial
    
    def get(self, request, *args, **kwargs):
        self.object = None
        formas_pago = FormaPago.objects.all().order_by('descripcion')        
        if not formas_pago:
            return HttpResponseRedirect(reverse('contabilidad:crear_forma_pago'))
        else:
            form_class = self.get_form_class()
            form = self.get_form(form_class)
            detalle_orden_servicios_formset=DetalleOrdenServiciosFormSet()
            return self.render_to_response(self.get_context_data(form=form,
                                                                 detalle_orden_servicios_formset=detalle_orden_servicios_formset))
    
    def post(self, request, *args, **kwargs):                      
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_orden_servicios_formset = DetalleOrdenServiciosFormSet(request.POST)
        if form.is_valid() and detalle_orden_servicios_formset.is_valid():
            return self.form_valid(form, detalle_orden_servicios_formset)
        else:
            return self.form_invalid(form, detalle_orden_servicios_formset)
    
    def form_valid(self, form, detalle_orden_servicios_formset):
        try:
            with transaction.atomic():
                self.object = form.save()
                referencia = self.object.cotizacion
                detalles = []
                cont = 1
                for detalle_orden_servicios_form in detalle_orden_servicios_formset:
                    cotizacion = detalle_orden_servicios_form.cleaned_data.get('cotizacion')
                    codigo = detalle_orden_servicios_form.cleaned_data.get('codigo')
                    cantidad = detalle_orden_servicios_form.cleaned_data.get('cantidad')
                    precio = detalle_orden_servicios_form.cleaned_data.get('precio')
                    valor = detalle_orden_servicios_form.cleaned_data.get('valor')
                    if cantidad and precio and valor:
                        try:
                            detalle_cotizacion = DetalleCotizacion.objects.get(pk=cotizacion)
                            detalle_orden_servicios = DetalleOrdenServicios(detalle_cotizacion = detalle_cotizacion,
                                                                            nro_detalle = cont,
                                                                            orden = self.object,
                                                                            cantidad = cantidad,
                                                                            precio = precio,
                                                                            valor = valor)
                        except DetalleCotizacion.DoesNotExist:
                            producto = Producto.objects.get(pk = codigo)
                            detalle_orden_servicios = DetalleOrdenServicios(producto = producto,
                                                                            nro_detalle = cont,
                                                                            orden = self.object,
                                                                            cantidad = cantidad,
                                                                            precio = precio,
                                                                            valor = valor)
                             
                        detalles.append(detalle_orden_servicios)                        
                        cont = cont + 1
                DetalleOrdenServicios.objects.bulk_create(detalles, referencia)
                return HttpResponseRedirect(reverse('compras:detalle_orden_servicios', args=[self.object.codigo]))                
        except IntegrityError:
                messages.error(self.request, 'Error guardando la cotizacion.')
        
    def form_invalid(self, form, detalle_orden_servicios_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_orden_servicios_formset = detalle_orden_servicios_formset))        
    
class CrearConformidadServicio(CreateView):
    form_class = ConformidadServicioForm
    template_name = "compras/conformidad_servicio.html"
    model = ConformidadServicio
    
    def get_initial(self):
        initial = super(CrearConformidadServicio, self).get_initial()
        initial['total'] = 0
        initial['subtotal'] = 0
        initial['total_letras'] = ''
        return initial
    
    def get(self, request, *args, **kwargs):
        self.object = None
        formas_pago = FormaPago.objects.all().order_by('descripcion')        
        if not formas_pago:
            return HttpResponseRedirect(reverse('compras:crear_forma_pago'))
        else:
            form_class = self.get_form_class()
            form = self.get_form(form_class)
            detalle_conformidad_servicio_formset=DetalleConformidadServicioFormSet()
            return self.render_to_response(self.get_context_data(form=form,
                                                                 detalle_conformidad_servicio_formset=detalle_conformidad_servicio_formset))
            
    def post(self, request, *args, **kwargs):                      
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_conformidad_servicio_formset = DetalleConformidadServicioFormSet(request.POST)
        if form.is_valid() and detalle_conformidad_servicio_formset.is_valid():
            return self.form_valid(form, detalle_conformidad_servicio_formset)
        else:
            return self.form_invalid(form, detalle_conformidad_servicio_formset)
    
    def form_valid(self, form, detalle_conformidad_servicio_formset):
        try:
            with transaction.atomic():
                self.object = form.save()
                referencia = self.object.orden_servicios
                detalles = []
                cont = 1
                for detalle_orden_servicios_form in detalle_conformidad_servicio_formset:
                    orden_servicios = detalle_orden_servicios_form.cleaned_data.get('orden_servicios')
                    cantidad = detalle_orden_servicios_form.cleaned_data.get('cantidad')
                    precio = detalle_orden_servicios_form.cleaned_data.get('precio')
                    valor = detalle_orden_servicios_form.cleaned_data.get('valor')
                    detalle_orden_servicios = DetalleOrdenServicios.objects.get(pk=orden_servicios)
                    if cantidad and precio and valor:
                        detalle_conformidad_servicio = DetalleConformidadServicio(detalle_orden_servicios = detalle_orden_servicios,
                                                                                  nro_detalle = cont,
                                                                                  conformidad = self.object,
                                                                                  cantidad=cantidad) 
                        detalles.append(detalle_conformidad_servicio)                        
                        cont = cont + 1
                DetalleConformidadServicio.objects.bulk_create(detalles, referencia)
                return HttpResponseRedirect(reverse('compras:detalle_conformidad_servicios', args=[self.object.codigo]))
        except IntegrityError:
                messages.error(self.request, 'Error guardando la cotizacion.')
        
    def form_invalid(self, form, detalle_conformidad_servicio_formset):
        return self.render_to_response(self.get_context_data(form=form, 
                                                             detalle_conformidad_servicio_form = detalle_conformidad_servicio_formset))
    
class DetalleProveedor(DetailView):
    model = Proveedor
    template_name = 'compras/detalle_proveedor.html'
    
class DetalleOperacionCotizacion(DetailView):
    model = Cotizacion
    template_name = 'compras/detalle_cotizacion.html'
    
class DetalleOperacionOrdenCompra(DetailView):
    model = OrdenCompra
    template_name = 'compras/detalle_orden_compra.html'

class DetalleOperacionOrdenServicios(DetailView):
    model = OrdenServicios
    template_name = 'compras/detalle_orden_servicios.html'
    
class DetalleOperacionConformidadServicios(DetailView):
    model = ConformidadServicio
    template_name = 'compras/detalle_conformidad_servicios.html'
    
class EliminarCotizacion(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            cotizacion = Cotizacion.objects.get(codigo = codigo)
            cotizacion_json = {}
            cotizacion_json['codigo'] = codigo
            ordenes_servicios = cotizacion.ordenservicios_set.all()
            if len(ordenes_servicios)>0:
                cotizacion_json['ordenes'] = 'SI'
            else:
                cotizacion_json['ordenes'] = 'NO'                
                ordenes_compras = cotizacion.ordencompra_set.all()
                if len(ordenes_compras)>0:
                    cotizacion_json['ordenes'] = 'SI'
                else:
                    cotizacion_json['ordenes'] = 'NO'
                
                with transaction.atomic():
                    cotizacion.eliminar_referencia()
                    cotizacion.eliminar_cotizacion()
                    DetalleCotizacion.objects.filter(cotizacion = cotizacion).delete()
            data = simplejson.dumps(cotizacion_json)
            return HttpResponse(data, 'application/json')
    
class EliminarOrdenCompra(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            orden = OrdenCompra.objects.get(codigo=codigo)
            movimiento_json = {}
            movimiento_json['codigo'] = codigo
            if len(orden.movimiento_set.all())>0:
                movimiento_json['movimientos'] = 'SI'
            else:
                movimiento_json['movimientos'] = 'NO'
                with transaction.atomic():
                    if orden.cotizacion is not None:
                        orden.eliminar_referencia()
                    OrdenCompra.objects.filter(codigo=codigo).update(estado = OrdenCompra.STATUS.CANC, cotizacion=None)
                    DetalleOrdenCompra.objects.filter(orden=orden).delete()                
            data = simplejson.dumps(movimiento_json)
            return HttpResponse(data, 'application/json')
        
class EliminarOrdenServicios(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            orden = OrdenServicios.objects.get(codigo = codigo)
            orden_json = {}
            orden_json['codigo'] = codigo
            if len(orden.conformidadservicio_set.all())>0:                
                orden_json['conformidades'] = 'SI'
            else:
                orden_json['conformidades'] = 'NO'
                with transaction.atomic():
                    if orden.cotizacion is not None:
                        orden.eliminar_referencia()
                    OrdenServicios.objects.filter(codigo = codigo).update(estado = OrdenServicios.STATUS.CANC, cotizacion=None)
                    DetalleOrdenServicios.objects.filter(orden=orden).delete()                
            data = simplejson.dumps(orden_json)
            return HttpResponse(data, 'application/json')
        
class EliminarConformidadServicio(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            conformidad = ConformidadServicio.objects.get(codigo = codigo)
            conformidad_json = {}
            conformidad_json['codigo'] = codigo            
            with transaction.atomic():
                if conformidad.orden_servicios is not None:
                    conformidad.eliminar_referencia()
                ConformidadServicio.objects.filter(codigo = codigo).update(estado = False)
                DetalleConformidadServicio.objects.filter(conformidad = conformidad).delete()                
            data = simplejson.dumps(conformidad_json)
            return HttpResponse(data, 'application/json')
        
class EliminarProveedor(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            ruc = request.GET['ruc']
            proveedor = Proveedor.objects.get(pk=ruc)
            proveedor_json = {}
            proveedor_json['ruc'] = ruc
            Proveedor.objects.filter(pk=ruc).update(estado = False)
            data = simplejson.dumps(proveedor_json)
            return HttpResponse(data, 'application/json')
        
class ListadoProveedores(ListView):
    model = Proveedor
    template_name = 'compras/proveedores.html'
    context_object_name = 'proveedores'
    queryset = Proveedor.objects.filter(estado=True).order_by('razon_social')
    
    @method_decorator(permission_required('compras.ver_tabla_proveedores',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoProveedores, self).dispatch(*args, **kwargs)
    
class ListadoCotizaciones(ListView):
    model = Cotizacion
    template_name = 'compras/cotizaciones.html'
    context_object_name = 'cotizaciones'
    queryset = Cotizacion.objects.exclude(estado=Cotizacion.STATUS.CANC).order_by('codigo')
    
    @method_decorator(permission_required('compras.ver_tabla_cotizaciones',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoCotizaciones, self).dispatch(*args, **kwargs)
    
class ListadoOrdenesCompra(ListView):
    model = OrdenCompra
    template_name = 'compras/ordenes_compra.html'
    context_object_name = 'ordenes_compra'
    queryset = OrdenCompra.objects.exclude(estado=OrdenCompra.STATUS.CANC).order_by('codigo')
    
    @method_decorator(permission_required('compras.ver_tabla_ordenes_compra',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoOrdenesCompra, self).dispatch(*args, **kwargs)
    
class ListadoOrdenesServicios(ListView):
    model = OrdenServicios
    template_name = 'compras/ordenes_servicios.html'
    context_object_name = 'ordenes_servicios'
    queryset = OrdenServicios.objects.filter().order_by('codigo')
    
    @method_decorator(permission_required('compras.ver_tabla_ordenes_servicios',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoOrdenesServicios, self).dispatch(*args, **kwargs)
    
class ListadoOrdenesCompraPorCotizacion(ListView):
    model = OrdenCompra
    template_name = 'compras/ordenes_compra.html'
    context_object_name = 'ordenes_compra'    
    
    @method_decorator(permission_required('compras.ver_tabla_ordenes_compra',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoOrdenesCompraPorCotizacion, self).dispatch(*args, **kwargs)
    
    def get_queryset(self):
        cotizacion = Cotizacion.objects.get(pk=self.kwargs['cotizacion'])
        queryset = cotizacion.ordencompra_set.all()
        return queryset
    
class ListadoOrdenesServiciosPorCotizacion(ListView):
    model = OrdenCompra
    template_name = 'compras/ordenes_servicios.html'
    context_object_name = 'ordenes_servicios'    
    
    @method_decorator(permission_required('compras.ver_tabla_ordenes_servicios',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoOrdenesServiciosPorCotizacion, self).dispatch(*args, **kwargs)
    
    def get_queryset(self):
        cotizacion = Cotizacion.objects.get(pk=self.kwargs['cotizacion'])
        queryset = cotizacion.ordenservicios_set.all()
        return queryset
    
class ListadoConformidadesServicio(ListView):
    model = ConformidadServicio
    template_name = 'compras/conformidades_servicio.html'
    context_object_name = 'conformidades'
    queryset = ConformidadServicio.objects.filter(estado=True).order_by('codigo')
    
    @method_decorator(permission_required('compras.ver_tabla_conformidades_servicio',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoConformidadesServicio, self).dispatch(*args, **kwargs)
    
class ListadoMovimientosPorOrdenCompra(ListView):
    template_name = 'almacen/movimientos.html'
    context_object_name = 'movimientos'    
    
    @method_decorator(permission_required('compras.ver_tabla_movimientos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoMovimientosPorOrdenCompra, self).dispatch(*args, **kwargs)
    
    def get_queryset(self):
        orden_compra = OrdenCompra.objects.get(pk=self.kwargs['orden'])
        queryset = orden_compra.movimiento_set.all()
        return queryset
    
class ListadoConformidadesPorOrdenServicios(ListView):
    template_name = 'compras/conformidades_servicio.html'
    context_object_name = 'conformidades'    
    
    @method_decorator(permission_required('compras.ver_tabla_conformidades_servicio',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoConformidadesPorOrdenServicios, self).dispatch(*args, **kwargs)
    
    def get_queryset(self):
        orden_servicios = OrdenServicios.objects.get(pk=self.kwargs['orden'])
        queryset = orden_servicios.conformidadservicio_set.all()
        return queryset
        
class ModificarProveedor(UpdateView):
    model = Proveedor
    template_name = 'compras/proveedor.html'
    form_class = ProveedorForm
        
    @method_decorator(permission_required('compras.change_proveedor',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarProveedor, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('compras:detalle_proveedor', args=[self.object.pk])
    
    def get_initial(self):
        initial = super(ModificarProveedor, self).get_initial()
        initial['fecha_alta'] = self.object.fecha_alta.strftime('%d/%m/%Y')
        return initial

class ModificarCotizacion(UpdateView):
    form_class = CotizacionForm
    template_name = "compras/cotizacion.html"
    model = Cotizacion
    
    @method_decorator(permission_required('compras.change_cotizacion',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        cotizacion = self.get_object()
        if cotizacion.estado == Cotizacion.STATUS.PEND:
            return super(ModificarCotizacion, self).dispatch(*args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))
    
    def get_initial(self):
        initial = super(ModificarCotizacion, self).get_initial()
        cotizacion = self.object
        initial['codigo'] = cotizacion.codigo
        initial['ruc'] = cotizacion.proveedor.ruc
        initial['razon_social'] = cotizacion.proveedor.razon_social
        initial['direccion'] = cotizacion.proveedor.direccion
        initial['fecha'] = cotizacion.fecha.strftime('%d/%m/%Y')
        initial['referencia'] = cotizacion.requerimiento
        initial['observaciones'] = cotizacion.observaciones        
        return initial
    
    def get_context_data(self, **kwargs):
        cotizacion = self.object
        detalles = DetalleCotizacion.objects.filter(cotizacion=cotizacion).order_by('nro_detalle')
        cant_detalles = detalles.count()
        context = super(ModificarCotizacion, self).get_context_data(**kwargs)
        context['cotizacion'] = cotizacion
        context['detalles'] = detalles
        context['cant_detalles'] = cant_detalles
        return context
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalles = DetalleCotizacion.objects.filter(cotizacion=self.object).order_by('nro_detalle')
        detalles_data = []
        for detalle in detalles:
            d = {'requerimiento': detalle.detalle_requerimiento.pk,
                 'codigo': detalle.detalle_requerimiento.producto.codigo,
                 'nombre': detalle.detalle_requerimiento.producto.descripcion,
                 'unidad': detalle.detalle_requerimiento.producto.unidad_medida.codigo,
                 'cantidad': detalle.cantidad}
            detalles_data.append(d)
        detalle_cotizacion_formset=DetalleCotizacionFormSet(initial=detalles_data)
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_cotizacion_formset = detalle_cotizacion_formset))
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)        
        detalle_cotizacion_formset = DetalleCotizacionFormSet(request.POST)
        if form.is_valid() and detalle_cotizacion_formset.is_valid():
            return self.form_valid(form, detalle_cotizacion_formset)
        else:
            return self.form_invalid(form, detalle_cotizacion_formset)
    
    def form_valid(self, form, detalle_cotizacion_formset):
        try:
            with transaction.atomic():
                self.object.eliminar_referencia()
                form.save()
                detalles = []
                cont = 1                
                for detalle_cotizacion_form in detalle_cotizacion_formset:
                    detalle_requerimiento = detalle_cotizacion_form.cleaned_data.get('requerimiento')
                    codigo = detalle_cotizacion_form.cleaned_data.get('codigo')
                    nombre = detalle_cotizacion_form.cleaned_data.get('nombre')
                    unidad = detalle_cotizacion_form.cleaned_data.get('unidad')
                    cantidad = detalle_cotizacion_form.cleaned_data.get('cantidad')                    
                    detalle_requerimiento = DetalleRequerimiento.objects.get(pk = detalle_requerimiento)
                    if cantidad:
                        detalle_cotizacion = DetalleCotizacion(detalle_requerimiento = detalle_requerimiento,
                                                               nro_detalle = cont,
                                                               cotizacion = self.object,
                                                               cantidad = cantidad) 
                        detalles.append(detalle_cotizacion)                        
                        cont = cont + 1
                DetalleCotizacion.objects.bulk_create(detalles, self.object.requerimiento)
                return HttpResponseRedirect(reverse('compras:detalle_cotizacion', args=[self.object.codigo]))
        except IntegrityError:
                messages.error(self.request, 'Error guardando el requerimiento.')
        
    def form_invalid(self, form, detalle_cotizacion_formset):
        return self.render_to_response(self.get_context_data(form=form))
    
class ModificarConformidadServicio(UpdateView):
    template_name = 'compras/conformidad_servicio.html'
    form_class = ConformidadServicioForm
    model = ConformidadServicio
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        return self.render_to_response(self.get_context_data(form=form))
    
    def get_initial(self):
        initial = super(ModificarConformidadServicio, self).get_initial()
        conformidad = self.object
        initial['cod_conformidad_servicio'] = conformidad.codigo
        initial['orden_servicios'] = conformidad.orden_servicios
        initial['doc_sustento'] = conformidad.doc_sustento
        initial['fecha'] = conformidad.fecha.strftime('%d/%m/%Y')                
        return initial 
        
    def get_context_data(self, **kwargs):
        conformidad = self.object
        detalles = DetalleConformidadServicio.objects.filter(conformidad=conformidad)
        cant_detalles = detalles.count()
        context = super(ModificarConformidadServicio, self).get_context_data(**kwargs)
        context['conformidad'] = conformidad
        context['detalles'] = detalles
        context['cant_detalles'] = cant_detalles
        return context
            
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)        
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)
    
    def form_valid(self, form):
        try:
            with transaction.atomic():
                self.eliminar_referencia()
                form.save()
                return HttpResponseRedirect(reverse('compras:detalle_cotizacion', args=[self.object.codigo]))
        except IntegrityError:
            messages.error(self.request, 'Error guardando el requerimiento.')


class ModificarOrdenCompra(UpdateView):
    template_name = 'compras/orden_compra.html'
    form_class = OrdenCompraForm
    model = OrdenCompra

    @method_decorator(permission_required('compras.change_ordencompra', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        orden_compra = self.get_object()
        if orden_compra.estado == OrdenCompra.STATUS.PEND:
            return super(ModificarOrdenCompra, self).dispatch(*args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.estado == OrdenCompra.STATUS.PEND:
            form_class = self.get_form_class()
            form = self.get_form(form_class)
            detalles = DetalleOrdenCompra.objects.filter(orden=self.object).order_by('nro_detalle')
            detalles_data = []
            for detalle in detalles:
                try:
                    d = {'cotizacion': detalle.detalle_cotizacion.pk,
                         'codigo': detalle.detalle_cotizacion.detalle_requerimiento.producto.codigo,
                         'nombre': detalle.detalle_cotizacion.detalle_requerimiento.producto.descripcion,
                         'unidad': detalle.detalle_cotizacion.detalle_requerimiento.producto.unidad_medida.codigo,
                         'cantidad': detalle.cantidad,
                         'precio': detalle.precio,
                         'impuesto': detalle.impuesto,
                         'valor': detalle.valor}
                except:
                    d = {'cotizacion': '0',
                         'codigo': detalle.producto.codigo,
                         'nombre': detalle.producto.descripcion,
                         'unidad': detalle.producto.unidad_medida.codigo,
                         'cantidad': detalle.cantidad,
                         'precio': detalle.precio,
                         'impuesto': detalle.impuesto,
                         'valor': detalle.valor_sin_igv}
                detalles_data.append(d)
            detalle_orden_compra_formset = DetalleOrdenCompraFormSet(initial=detalles_data)
            return self.render_to_response(self.get_context_data(form=form,
                                                                 detalle_orden_compra_formset=detalle_orden_compra_formset))
        else:
            return HttpResponseRedirect(reverse('compras:ordenes_compra'))

    def get_initial(self):
        initial = super(ModificarOrdenCompra, self).get_initial()
        orden = self.object
        cotizacion = orden.cotizacion
        if cotizacion is None:
            proveedor = orden.proveedor
        else:
            proveedor = orden.cotizacion.proveedor
        initial['codigo'] = orden.codigo
        initial['ruc'] = proveedor.ruc
        initial['razon_social'] = proveedor.razon_social
        initial['direccion'] = proveedor.direccion
        initial['fecha'] = orden.fecha.strftime('%d/%m/%Y')
        initial['formas_pago'] = orden.forma_pago
        initial['referencia'] = orden.cotizacion
        try:
            monto_impuesto = IMPUESTO_COMPRA.monto
        except:
            return HttpResponseRedirect(reverse('contabilidad:configuracion'))
        initial['impuesto_actual'] = monto_impuesto
        initial['total'] = orden.total
        initial['subtotal'] = orden.subtotal
        initial['impuesto'] = orden.impuesto
        initial['total_letras'] = orden.total_letras
        initial['observaciones'] = orden.observaciones
        return initial

    def get_context_data(self, **kwargs):
        orden = self.object
        context = super(ModificarOrdenCompra, self).get_context_data(**kwargs)
        context['orden'] = orden
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_orden_compra_formset = DetalleOrdenCompraFormSet(request.POST)
        if form.is_valid() and detalle_orden_compra_formset.is_valid():
            return self.form_valid(form, detalle_orden_compra_formset)
        else:
            return self.form_invalid(form, detalle_orden_compra_formset)

    def form_valid(self, form, detalle_orden_compra_formset):
        try:
            with transaction.atomic():
                if self.object.cotizacion is not None:
                    self.object.eliminar_referencia()
                DetalleOrdenCompra.objects.filter(orden=self.object).delete()
                self.object = form.save()
                referencia = self.object.cotizacion
                detalles = []
                cont = 1
                for detalle_orden_compra_form in detalle_orden_compra_formset:
                    cotizacion = detalle_orden_compra_form.cleaned_data.get('cotizacion')
                    codigo = detalle_orden_compra_form.cleaned_data.get('codigo')
                    cantidad = detalle_orden_compra_form.cleaned_data.get('cantidad')
                    precio = detalle_orden_compra_form.cleaned_data.get('precio')
                    valor = detalle_orden_compra_form.cleaned_data.get('valor')
                    impuesto = detalle_orden_compra_form.cleaned_data.get('impuesto')
                    if cantidad and precio and valor and impuesto:
                        try:
                            detalle_cotizacion = DetalleCotizacion.objects.get(pk=cotizacion)
                            detalle_orden_compra = DetalleOrdenCompra(detalle_cotizacion=detalle_cotizacion,
                                                                      nro_detalle=cont,
                                                                      orden=self.object,
                                                                      cantidad=cantidad,
                                                                      precio=precio)
                        except DetalleCotizacion.DoesNotExist:
                            producto = Producto.objects.get(pk=codigo)
                            detalle_orden_compra = DetalleOrdenCompra(producto=producto,
                                                                      nro_detalle=cont,
                                                                      orden=self.object,
                                                                      cantidad=cantidad,
                                                                      precio=precio)
                        detalles.append(detalle_orden_compra)
                        cont = cont + 1
                        if cont > 1:
                            DetalleOrdenCompra.objects.bulk_create(detalles, referencia)
                return HttpResponseRedirect(reverse('compras:detalle_orden_compra', args=[self.object.pk]))
        except IntegrityError:
            messages.error(self.request, 'Error guardando la cotizacion.')

    def form_invalid(self, form, detalle_orden_compra_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_orden_compra_formset=detalle_orden_compra_formset))
    
class ModificarOrdenServicios(UpdateView):
    template_name = 'compras/orden_servicio.html'
    form_class = OrdenServiciosForm
    model = OrdenServicios
    
    @method_decorator(permission_required('compras.change_ordenservicios',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        orden_servicios = self.get_object()
        if orden_servicios.estado == OrdenServicios.STATUS.PEND:
            return super(ModificarOrdenServicios, self).dispatch(*args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))
    
    def get_initial(self):
        initial = super(ModificarOrdenServicios, self).get_initial()
        orden = self.object
        cotizacion = orden.cotizacion
        if cotizacion is None:
            proveedor = orden.proveedor
        else:
            proveedor = orden.cotizacion.proveedor
        initial['codigo'] = orden.codigo
        initial['ruc'] = proveedor.ruc
        initial['razon_social'] = proveedor.razon_social
        initial['direccion'] = proveedor.direccion
        initial['fecha'] = orden.fecha.strftime('%d/%m/%Y')
        initial['formas_pago'] = orden.forma_pago
        initial['referencia'] = orden.cotizacion
        initial['proceso'] = orden.proceso
        initial['total'] = orden.total
        initial['subtotal'] = orden.subtotal
        initial['impuesto'] = orden.impuesto
        initial['total_letras'] = orden.total_letras
        initial['observaciones'] = orden.observaciones
        return initial 
        
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.estado == OrdenCompra.STATUS.PEND:            
            form_class = self.get_form_class()
            form = self.get_form(form_class)
            detalles = DetalleOrdenServicios.objects.filter(orden=self.object).order_by('nro_detalle')
            detalles_data = []
            for detalle in detalles:
                try:
                    d = {'cotizacion': detalle.detalle_cotizacion.pk,
                         'codigo': detalle.detalle_cotizacion.detalle_requerimiento.producto.codigo,
                         'nombre': detalle.detalle_cotizacion.detalle_requerimiento.producto.descripcion,
                         'unidad': detalle.detalle_cotizacion.detalle_requerimiento.producto.unidad_medida.codigo,
                         'cantidad': detalle.cantidad,
                         'precio': detalle.precio,
                         'valor': detalle.valor }
                except:
                    d = {'cotizacion': '0',
                         'codigo': detalle.producto.codigo,
                         'nombre': detalle.producto.descripcion,
                         'unidad': detalle.producto.unidad_medida.codigo,
                         'cantidad': detalle.cantidad,
                         'precio': detalle.precio,
                         'valor': detalle.valor }
                detalles_data.append(d)
            detalle_orden_servicios_formset = DetalleOrdenServiciosFormSet(initial=detalles_data)
            return self.render_to_response(self.get_context_data(form=form,
                                                                 detalle_orden_servicios_formset=detalle_orden_servicios_formset))
        else:
            return HttpResponseRedirect(reverse('compras:ordenes_compra'))
    
    def get_context_data(self, **kwargs):
        orden = self.object
        context = super(ModificarOrdenServicios, self).get_context_data(**kwargs)
        context['orden'] = orden                
        return context      
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_orden_servicios_formset = DetalleOrdenServiciosFormSet(request.POST)
        if form.is_valid() and detalle_orden_servicios_formset.is_valid():
            return self.form_valid(form, detalle_orden_servicios_formset)
        else:
            return self.form_invalid(form, detalle_orden_servicios_formset)
    
    def form_valid(self, form, detalle_orden_servicios_formset):
        #try:
        with transaction.atomic():
            if self.object.cotizacion is not None:
                self.object.eliminar_referencia()
            DetalleOrdenServicios.objects.filter(orden = self.object).delete()
            self.object = form.save()
            referencia = self.object.cotizacion
            detalles = []
            cont = 1
            for detalle_orden_servicios_form in detalle_orden_servicios_formset:
                cotizacion = detalle_orden_servicios_form.cleaned_data.get('cotizacion')
                codigo = detalle_orden_servicios_form.cleaned_data.get('codigo')
                cantidad = detalle_orden_servicios_form.cleaned_data.get('cantidad')
                precio = detalle_orden_servicios_form.cleaned_data.get('precio')
                valor = detalle_orden_servicios_form.cleaned_data.get('valor')                    
                if cantidad and precio and valor:
                    try:
                        detalle_cotizacion = DetalleCotizacion.objects.get(pk=cotizacion)
                        detalle_orden_servicios = DetalleOrdenServicios(detalle_cotizacion = detalle_cotizacion,
                                                                        nro_detalle = cont,
                                                                        orden = self.object,
                                                                        cantidad = cantidad,
                                                                        precio = precio,
                                                                        valor = valor) 
                    except:
                        producto = Producto.objects.get(pk = codigo)
                        detalle_orden_servicios = DetalleOrdenServicios(producto = producto,
                                                                        nro_detalle = cont,
                                                                        orden = self.object,
                                                                        cantidad = cantidad,
                                                                        precio = precio,
                                                                        valor = valor)
                    detalles.append(detalle_orden_servicios)                        
                    cont = cont + 1
            DetalleOrdenServicios.objects.bulk_create(detalles,referencia)
            return HttpResponseRedirect(reverse('compras:detalle_orden_servicios', args=[self.object.codigo]))                
        #except IntegrityError:
                #messages.error(self.request, 'Error guardando la Orden de Servicios.')
        
    def form_invalid(self, form, detalle_orden_servicios_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_orden_servicios_formset=detalle_orden_servicios_formset))

class ObtenerDetalleCotizacion(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():            
            cotizacion = request.GET['cotizacion']
            tipo_busqueda =  request.GET['tipo_busqueda']
            if tipo_busqueda == 'PRODUCTOS':
                detalles = DetalleCotizacion.objects.filter(Q(estado = DetalleCotizacion.STATUS.PEND) | Q(estado = DetalleCotizacion.STATUS.ELEG_PARC),
                                                            cotizacion__codigo = cotizacion,
                                                            detalle_requerimiento__producto__es_servicio = False).order_by('nro_detalle')
                try:
                    monto_impuesto = IMPUESTO_COMPRA.monto
                except:
                    monto_impuesto = 0
            elif tipo_busqueda == 'SERVICIOS':
                monto_impuesto = 1
                detalles = DetalleCotizacion.objects.filter(cotizacion__codigo=cotizacion,
                                                            detalle_requerimiento__producto__es_servicio=True).order_by('nro_detalle')                
            
            lista_detalles = []            
            for detalle in detalles:                 
                det = {}       
                det['cotizacion'] = detalle.id
                try:
                    det['codigo'] = detalle.detalle_requerimiento.producto.codigo                
                    det['nombre'] = detalle.detalle_requerimiento.producto.descripcion                    
                    det['precio'] = str(detalle.detalle_requerimiento.producto.precio)
                    cantidad = detalle.cantidad - detalle.detalle_requerimiento.cantidad_comprada
                    det['cantidad'] = str(cantidad)
                    valor = detalle.detalle_requerimiento.producto.precio * cantidad
                    if tipo_busqueda == 'PRODUCTOS':
                        det['unidad'] = detalle.detalle_requerimiento.producto.unidad_medida.codigo
                        base = valor / (monto_impuesto + 1)
                        det['impuesto'] = str(round(valor - base,5))
                        det['valor'] = str(round(valor,5))
                    elif tipo_busqueda == 'SERVICIOS':
                        det['unidad'] = detalle.detalle_requerimiento.producto.unidad_medida.codigo
                        det['valor'] = str(round(valor))
                    lista_detalles.append(det)
                except:
                    pass
            if tipo_busqueda == 'PRODUCTOS':
                formset = DetalleOrdenCompraFormSet(initial=lista_detalles)
            elif tipo_busqueda == 'SERVICIOS':
                formset = DetalleOrdenServiciosFormSet(initial=lista_detalles)
            lista_json = []
            if tipo_busqueda == 'PRODUCTOS':
                for form in formset:
                    detalle_json = {}    
                    detalle_json['cotizacion'] = str(form['cotizacion'])
                    detalle_json['codigo'] = str(form['codigo'])
                    detalle_json['nombre'] = str(form['nombre'])
                    detalle_json['precio'] = str(form['precio'])
                    detalle_json['unidad'] = str(form['unidad'])
                    detalle_json['cantidad'] = str(form['cantidad'])                
                    detalle_json['impuesto'] = str(form['impuesto'])
                    detalle_json['valor'] = str(form['valor'])
                    lista_json.append(detalle_json)
            elif tipo_busqueda == 'SERVICIOS':
                for form in formset:
                    detalle_json = {}    
                    detalle_json['cotizacion'] = str(form['cotizacion'])
                    detalle_json['codigo'] = str(form['codigo'])
                    detalle_json['nombre'] = str(form['nombre'])
                    detalle_json['precio'] = str(form['precio'])
                    detalle_json['unidad'] = str(form['unidad'])
                    detalle_json['cantidad'] = str(form['cantidad'])                
                    detalle_json['valor'] = str(form['valor'])
                    lista_json.append(detalle_json)
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')

class ObtenerDetalleOrdenCompra(TemplateView):

    def obtener_fecha(self,r_fecha):
        anio = int(r_fecha[6:])
        mes = int(r_fecha[3:5])
        dia = int(r_fecha[0:2])
        fecha = datetime.datetime(anio,mes,dia)
        return fecha
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            orden_compra = OrdenCompra.objects.get(codigo=request.GET['orden_compra'])
            fecha = self.obtener_fecha(request.GET['fecha'])
            tipo_cambio = 1
            if orden_compra.dolares:
                try:
                    tipo_cambio = TipoCambio.objects.get(fecha=fecha).monto
                except:
                    tipo_cambio = 0
            lista_detalles = []
            lista_json = []
            if tipo_cambio > 0:
                detalles = DetalleOrdenCompra.objects.filter(orden=orden_compra,
                                                             estado=DetalleOrdenCompra.STATUS.PEND).order_by('nro_detalle')
                for detalle in detalles:
                    det = {}
                    det['orden_compra'] = detalle.id
                    try:
                        det['codigo'] = detalle.detalle_cotizacion.detalle_requerimiento.producto.codigo
                        det['nombre'] = detalle.detalle_cotizacion.detalle_requerimiento.producto.descripcion
                        det['cantidad'] = str(detalle.cantidad-detalle.cantidad_ingresada)
                        det['precio'] = str(round(Decimal(detalle.precio_sin_igv) * tipo_cambio,5))
                        det['unidad'] = detalle.detalle_cotizacion.detalle_requerimiento.producto.unidad_medida.codigo
                        det['valor'] = str(round(Decimal(detalle.valor_sin_igv) * tipo_cambio,5))
                    except:
                        det['codigo'] = detalle.producto.codigo
                        det['nombre'] = detalle.producto.descripcion
                        det['cantidad'] = str(detalle.cantidad-detalle.cantidad_ingresada)
                        det['precio'] = str(round(Decimal(detalle.precio_sin_igv) * tipo_cambio,5))
                        det['unidad'] = detalle.producto.unidad_medida.codigo
                        det['valor'] = str(round(Decimal(detalle.valor_sin_igv) * tipo_cambio,5))
                    lista_detalles.append(det)
                formset = DetalleIngresoFormSet(initial=lista_detalles)
                for form in formset:
                    detalle_json = {}
                    detalle_json['orden_compra'] = str(form['orden_compra'])
                    detalle_json['codigo'] = str(form['codigo'])
                    detalle_json['nombre'] = str(form['nombre'])
                    detalle_json['cantidad'] = str(form['cantidad'])
                    detalle_json['precio'] = str(form['precio'])
                    detalle_json['unidad'] = str(form['unidad'])
                    detalle_json['valor'] = str(form['valor'])
                    lista_json.append(detalle_json)
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')
        
class ObtenerDetalleOrdenServicios(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            orden_servicios = request.GET['orden_servicios']
            detalles = DetalleOrdenServicios.objects.filter(orden__codigo=orden_servicios,
                                                            estado=DetalleOrdenServicios.STATUS.PEND).order_by('nro_detalle')
            lista_detalles = []
            for detalle in detalles:
                try:
                    det = {}       
                    det['orden_servicios'] = detalle.id
                    det['codigo'] = detalle.detalle_cotizacion.detalle_requerimiento.producto.codigo
                    det['servicio'] = detalle.detalle_cotizacion.detalle_requerimiento.producto.descripcion                
                    det['uso'] = detalle.detalle_cotizacion.detalle_requerimiento.uso                    
                    det['precio'] = str(detalle.precio)
                    det['cantidad'] = str(detalle.cantidad)
                    det['valor'] = str(detalle.valor)
                except:
                    det = {}       
                    det['orden_servicios'] = detalle.id
                    det['codigo'] = detalle.producto.codigo
                    det['servicio'] = detalle.producto.descripcion                
                    det['uso'] = detalle.producto.unidad_medida.descripcion                    
                    det['precio'] = str(detalle.precio)
                    det['cantidad'] = str(detalle.cantidad)
                    det['valor'] = str(detalle.valor)
                lista_detalles.append(det)                
            formset = DetalleConformidadServicioFormSet(initial=lista_detalles)
            lista_json = []
            for form in formset:
                detalle_json = {}    
                detalle_json['orden_servicios'] = str(form['orden_servicios'])
                detalle_json['servicio'] = str(form['servicio'])
                detalle_json['uso'] = str(form['uso'])
                detalle_json['precio'] = str(form['precio'])
                detalle_json['cantidad'] = str(form['cantidad'])                
                detalle_json['valor'] = str(form['valor'])
                lista_json.append(detalle_json)                
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')
        
"""class ReportePDFOrdenCompra(View):
    
    def get(self, request, *args, **kwargs): 
        codigo = kwargs['pk']
        orden = OrdenCompra.objects.get(codigo=codigo)        
        response = HttpResponse(content_type='application/pdf')                
        reporte = ReporteOrdenCompra('A4',orden)
        pdf = reporte.imprimir()        
        response.write(pdf)
        return response"""
        
class ReportePDFOrdenCompra(View):  
    
    def cabecera(self,pdf,orden):
        try:
            archivo_imagen = os.path.join(settings.MEDIA_ROOT,str(EMPRESA.logo))
            pdf.drawImage(archivo_imagen, 40, 750, 100, 90, mask='auto',preserveAspectRatio=True)
        except:
            pdf.drawString(40,800,str(archivo_imagen))
        pdf.setFont("Times-Roman", 14)
        pdf.drawString(230, 800, u"ORDEN DE COMPRA")
        pdf.setFont("Times-Roman", 11)
        pdf.drawString(455, 800, u"R.U.C. " + EMPRESA.ruc)
        pdf.setFont("Times-Roman", 13)
        pdf.drawString(250, 780, u"N° "+orden.codigo)
        pdf.setFont("Times-Roman", 10)
        pdf.drawString(430, 780, EMPRESA.distrito + " " + orden.fecha.strftime('%d de %b de %Y'))#orden.fecha.strftime('%d de %B de %Y')
        pdf.setFont("Times-Roman", 10)
        cotizacion = orden.cotizacion
        if cotizacion is None:
            proveedor = orden.proveedor
        else:
            proveedor = orden.cotizacion.proveedor
        pdf.drawString(40, 750, u"SEÑOR(ES): "+ proveedor.razon_social)
        pdf.drawString(440, 750, u"R.U.C.: "+ proveedor.ruc)
        direccion = proveedor.direccion
        if len(direccion)>60:
            pdf.drawString(40,730,u"DIRECCIÓN: "+direccion[0:60])
            pdf.drawString(105, 720,direccion[60:])
        else:            
            pdf.drawString(40,730,u"DIRECCIÓN: "+direccion)
        try:
            pdf.drawString(440, 730, u"TELÉFONO: "+ proveedor.telefono)
        except:
            pdf.drawString(440, 730, u"TELÉFONO: -")
        try:
            pdf.drawString(40, 710, u"REFERENCIA: "+orden.cotizacion.requerimiento.codigo+" - "+orden.cotizacion.requerimiento.oficina.nombre)
        except:
            pdf.drawString(40, 710, u"REFERENCIA: -")
        pdf.drawString(40, 690, u"PROCESO: -")
        pdf.setFont("Times-Roman", 8)
        pdf.drawString(40, 670, u"Sírvase remitirnos según especificaciones que detallamos lo siguiente: ")
        
    def detalle(self,pdf,y,orden):
        encabezados = ('Item', 'Cantidad', 'Unidad', u'Descripción','Precio','Total')
        try:
            detalles = [(detalle.nro_detalle, detalle.cantidad, detalle.detalle_cotizacion.detalle_requerimiento.producto.unidad_medida.descripcion, detalle.detalle_cotizacion.detalle_requerimiento.producto.descripcion, detalle.precio,round(detalle.valor,5)) for detalle in DetalleOrdenCompra.objects.filter(orden=orden)]
        except:
            detalles = [(detalle.nro_detalle, detalle.cantidad, detalle.producto.unidad_medida.descripcion, detalle.producto.descripcion, detalle.precio, round(detalle.precio,5)) for detalle in DetalleOrdenCompra.objects.filter(orden=orden)]
        adicionales = [('','','','','','')]*(15-len(detalles))
        detalle_orden = Table([encabezados] + detalles + adicionales,colWidths=[0.8 * cm, 1.9 * cm, 2 * cm,9.3* cm, 2 * cm, 2.5 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN',(0,0),(5,0),'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black), 
                #('LINEBELOW', (0,1), (5,-1), 0, colors.black),                  
                ('FONTSIZE', (0, 0), (-1, -1), 8),  
                ('ALIGN',(4,1),(-1,-1),'RIGHT'),           
            ]
        ))
        detalle_orden.wrapOn(pdf, 800, 600)
        detalle_orden.drawOn(pdf, 40,y+75)
        #Letras
        total_letras = [("SON: "+orden.total_letras,'')]
        tabla_total_letras = Table(total_letras,colWidths=[16 * cm, 2.5 * cm])
        tabla_total_letras.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (1, 0), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]
        ))
        tabla_total_letras.wrapOn(pdf, 800, 600)
        tabla_total_letras.drawOn(pdf, 40,y+55)
        
    def otros(self,pdf,y,orden):
        encabezados_otros = ('LUGAR DE ENTREGA', 'PLAZO DE ENTREGA', 'FORMA DE PAGO')
        otros = [(EMPRESA.direccion(),u"INMEDIATA",orden.forma_pago.descripcion)]
        tabla_otros = Table([encabezados_otros] + otros,colWidths=[6 * cm, 3.5 * cm, 4.5 * cm], rowHeights=[0.6 * cm, 1 * cm])
        tabla_otros.setStyle(TableStyle(
            [
                ('ALIGN',(0,0),(2,0),'CENTER'),  
                ('GRID', (0, 0), (2, 1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),                
            ]
        ))
        tabla_otros.wrapOn(pdf, 800, 600)
        tabla_otros.drawOn(pdf, 40,y+5)
        
    def cuadro_total(self,pdf,y,orden):
        pdf.drawString(445, y+40, u"SUB-TOTAL: ")
        pdf.drawString(445, y+20, u"IGV: ")
        pdf.drawString(445, y, u"TOTAL: S/")
        total = [[round(orden.subtotal,2)],[round(orden.impuesto,2)],[round(orden.total,2)]]
        tabla_total = Table(total,colWidths=[2.5 * cm])
        tabla_total.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (0, 2), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN',(0,0),(-1,-1),'RIGHT'),
            ]
        ))
        tabla_total.wrapOn(pdf, 800, 600)
        tabla_total.drawOn(pdf, 495,y-2)
        
    def cuadro_observaciones(self,pdf,y,orden):
        p = ParagraphStyle('parrafos')
        p.alignment = TA_JUSTIFY 
        p.fontSize = 10
        p.fontName="Times-Roman"
        obs=Paragraph("Observaciones: "+orden.observaciones,p)
        observaciones = [[obs]]
        tabla_observaciones = Table(observaciones,colWidths=[18.50 * cm], rowHeights=1.8 * cm)
        tabla_observaciones.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (0, 2), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN',(0,0),(-1,-1),'LEFT'),
                ('VALIGN',(0,0),(-1,-1),'TOP'),
            ]
        ))
        tabla_observaciones.wrapOn(pdf, 800, 600)
        tabla_observaciones.drawOn(pdf, 40,y-58)
        
    def afectacion_presupuesta(self,pdf):
        y=320
        pdf.drawString(40, y-90, u"HOJA DE AFECTACIÓN PRESUPUESTAL:")
        p = ParagraphStyle('parrafos')
        p.alignment = TA_JUSTIFY 
        p.fontSize = 8
        p.fontName="Times-Roman"
        lista = ListFlowable([
                          Paragraph("""Consignar el número de la presente Orden de Compra en su Guía de Remisión y Factura. 
                          Facturar a nombre de """ + smart_str(EMPRESA.razon_social),p),
                          Paragraph("El " + smart_str(EMPRESA.razon_social) + """, se reserva el derecho de devolver 
                          la mercaderia, sino se ajusta a las especificaciones requeridas, asimismo de anular la presente 
                          Orden de Compra.""",p),
                          Paragraph("""El pago de toda factura se hará de acuerdo a las condiciones establecidas.""",p)
                          ],bulletType='1'
                         )
        p1=Paragraph("RECIBIDO POR: ",p)        
        pdf.drawString(330, y-150,"FIRMA: ")
        pdf.line(370, y-150, 560, y-150)
        pdf.drawString(330, y-170,"NOMBRE: ")
        pdf.line(370, y-170, 560, y-170)
        pdf.drawString(330, y-190,"DNI: ")
        pdf.line(370, y-190, 560, y-190)
        pdf.setFont("Times-Roman", 6)
        pdf.drawString(525, y-127,"FECHA")
        afectacion = [[(Paragraph("IMPORTANTE:",p), lista),"RECIBIDO POR:"]]
        tabla_afectacion = Table(afectacion,colWidths=[10 * cm, 8.50 * cm])
        tabla_afectacion.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (1, 0), 1, colors.black),
                ('VALIGN',(1,0),(1,0),'TOP'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]
        ))
        tabla_afectacion.wrapOn(pdf, 800, 600)
        tabla_afectacion.drawOn(pdf, 40,y-200)
        fecha = [[' ',' ',' ']]
        tabla_fecha = Table(fecha,colWidths=[0.6 * cm, 0.6 * cm, 0.6 * cm], rowHeights=0.6 * cm)
        tabla_fecha.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 5),                
            ]
        ))
        tabla_fecha.wrapOn(pdf, 800, 600)
        tabla_fecha.drawOn(pdf, 510,y-120)        
        
    def get(self, request, *args, **kwargs):         
        codigo = kwargs['pk']
        orden = OrdenCompra.objects.get(pk=codigo)
        response = HttpResponse(content_type='application/pdf')
        #response['Content-Disposition'] = 'attachment; filename="orden_compra.pdf"'
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)  
        self.cabecera(pdf, orden)
        y=300
        self.detalle(pdf, y, orden)        
        self.otros(pdf, y, orden)
        self.cuadro_total(pdf, y, orden)
        self.cuadro_observaciones(pdf, y, orden)
        self.afectacion_presupuesta(pdf)
        pdf.setFont("Times-Roman", 8)
        pdf.drawString(115, y-250,"Elaborado por")
        pdf.drawString(430, y-250,"Autorizado por")
        pdf.line(70, y-240, 200, y-240)
        pdf.line(390, y-240, 520, y-240)
        pdf.drawCentredString(300, y-280, EMPRESA.direccion())
        pdf.showPage()    
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

class ReporteXLSOrdenCompra(TemplateView):

    def get(self, request, *args, **kwargs):
        orden=get_object_or_404(OrdenCompra, pk=kwargs['pk'])
        detalle_index=0
        cotizacion = orden.cotizacion
        if cotizacion is None:
            proveedor = orden.proveedor
        else:
            proveedor = orden.cotizacion.proveedor
        wb = Workbook()
        ws = wb.active

        ws.column_dimensions['B'].width=18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['I'].width = 15

        ws.merge_cells('B2:I2')
        ws['B2'].font=Font(name='Brush Script MT', size=20, color='0000ff', bold=True)
        ws['B2'].alignment=Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height=25
        ws['B2'] = 'Asociación de Bananeros Orgánicos Solidarios  Salitral'
        ws.merge_cells('B3:I3')
        ws['B3'].font = Font(name='Lucida Bright', size=9, color='148918', bold=True, italic=True)
        ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B3'] = 'Inscrita en Partida Nº 11003959 de la Superintendencia Nacional de'
        ws.merge_cells('B4:I4')
        ws['B4'].font = Font(name='Lucida Bright', size=9, color='148918', bold=True, italic=True)
        ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B4'] = 'los Registros Públicos R.U.C. Nº 20484149748'
        ws.merge_cells('B5:I5')
        ws['B5'].font = Font(name='Lucida Bright', size=9, color='ff0000', bold=True, italic=True)
        ws['B5'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B5'] = 'FLO Certified Banana Producer – FLO ID 2460'
        ws.merge_cells('B6:I6')
        ws['B6'].font = Font(name='Lucida Bright', size=9, color='ff0000', bold=True, italic=True)
        ws['B6'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B6'] = 'Certificate Organic USDA-NOP, EU2092/91, JAS. Registration Nº CU 805878'
        ws.merge_cells('B10:I10')
        ws['B10'].font = Font(name='Calibri', size=12)
        ws['B10'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B10'] = '"AÑO DEL BUEN SERVICIO AL CIUDADANO"'
        ws.merge_cells('D11:F11')
        ws['D11'].font = Font(name='Calibri', size=18, bold=True, underline="single")
        ws['D11'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[11].height = 25
        ws['D11'] = 'ORDEN DE COMPRA'
        ws['G11'].font = Font(name='Calibri', size=11, bold=True, underline="single")
        ws['G11'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G11'] = 'N°'


        ws['B13'].font = Font(size=11, bold=True)
        ws['B13'] = 'DATOS DEL CLIENTE Y DE LA FACTURA'
        ws['B14'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B14'] = 'RAZÓN SOCIAL'
        ws.merge_cells('C14:G14')
        ws['C14'].alignment = Alignment(horizontal="center")
        ws['C14'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C14'] = orden.proveedor.razon_social
        ws['H14'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H14'] = 'FECHA'
        ws['I14'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I14'] = orden.fecha
        ws['B15'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B15'] = 'CONTACTO'
        ws.merge_cells('C15:G15')
        ws['C15'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C15'].alignment = Alignment(horizontal="center")
        ws['C15'] = ''
        ws['H15'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H15'] = 'RUC/NIT'
        ws['I15'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I15'] = proveedor.ruc
        ws['B16'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B16'] = 'DIRECCIÓN'
        ws.merge_cells('C16:G16')
        ws['C16'].alignment = Alignment(horizontal="center")
        ws['C16'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C16'] = proveedor.direccion
        ws['H16'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H16'] = 'TELÉFONO'
        ws['I16'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        try:
            ws['I16'] = proveedor.telefono
        except:
            ws['I16'] = '-'
        ws['B17'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B17'] = 'E-MAIL'
        ws.merge_cells('C17:G17')
        ws['C14'].alignment = Alignment(horizontal="center")
        ws['C17'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C17'] = proveedor.correo
        ws['H17'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H17'] = 'RPM/RPC'
        ws['I17'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I17'] = ''


        ws['B19'].font = Font(size=11, bold=True)
        ws['B19'] = 'DATOS DE ENTREGA'
        ws.merge_cells('B20:C20')
        ws['B20'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B20'] = 'DIRECCIÓN DE ENTREGA'
        ws.merge_cells('D20:F20')
        ws['D20'].alignment = Alignment(horizontal="center")
        ws['D20'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D20'] = str(EMPRESA.direccion())
        ws.merge_cells('G20:H20')
        ws['G20'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G20'] = 'DEPARTAMENTO'
        ws['I20'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I20'] = str(EMPRESA.departamento)
        ws.merge_cells('B21:C22')
        ws['B21'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B21'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B21'] = 'PERSONAS AUTORIZADAS'
        ws.merge_cells('D21:F21')
        ws['D21'].alignment = Alignment(horizontal="center")
        ws['D21'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D21'] = ''
        ws.merge_cells('D22:F22')
        ws['D22'].alignment = Alignment(horizontal="center")
        ws['D22'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D22'] = ''
        ws.merge_cells('G21:H21')
        ws['G21'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G21'] = 'PROVINCIA'
        ws['I21'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I21'] = str(EMPRESA.provincia)
        ws.merge_cells('G22:H22')
        ws['G22'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G22'] = 'DISTRITO'
        ws['I22'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I22'] = str(EMPRESA.distrito)
        ws.merge_cells('B23:C24')
        ws['B23'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B23'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B23'] = 'TELÉFONOS'
        ws.merge_cells('D23:F23')
        ws['D23'].alignment = Alignment(horizontal="center")
        ws['D23'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D23'] = ''
        ws.merge_cells('G23:H23')
        ws['G23'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G23'] = 'COMENTARIO:'
        ws['I23'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I23'] = ''
        ws['B24'].border = Border(bottom=Side(border_style="thin"))
        ws.merge_cells('D24:F24')
        ws['D24'].alignment = Alignment(horizontal="center")
        ws['D24'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['D24'] = ''
        ws.merge_cells('G24:I24')
        ws['D24'].alignment = Alignment(horizontal="center")
        ws['G24'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G24'] = ''
        ws['J24'].border = Border(left=Side(border_style="thin"))
        ws.merge_cells('G25:I25')
        ws['G25'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G25'] = ''
        ws['J25'].border = Border(left=Side(border_style="thin"))
        ws.merge_cells('G26:I26')
        ws['G26'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G26'] = ''
        ws['J26'].border = Border(left=Side(border_style="thin"))


        ws['B28'].font = Font(size=11, bold=True)
        ws['B28'] = 'DATOS DEL PRODUCTO A ADQUIRIR'
        ws.merge_cells('B29:B30')
        ws['B29'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B29'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['B29'] = 'UNIDAD MEDIDA'
        ws.merge_cells('C29:F30')
        ws['C29'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C29'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['C29'] = 'DESCRIPCIÓN DEL PRODUCTO'
        ws.merge_cells('G29:G30')
        ws['G29'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G29'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G29'] = 'CANT.'
        ws['H29'].alignment = Alignment(horizontal="center")
        ws['H29'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H29'] = 'PRECIO'
        ws['I29'].alignment = Alignment(horizontal="center")
        ws['I29'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I29'] = 'PRECIO'
        ws['H30'].alignment = Alignment(horizontal="center")
        ws['H30'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['H30'] = 'UNITARIO'
        ws['I30'].alignment = Alignment(horizontal="center")
        ws['I30'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I30'] = 'TOTAL'

        for item in DetalleOrdenCompra.objects.filter(orden=orden):
            fila=31+detalle_index
            ws['B'+str(fila)].alignment = Alignment(horizontal="center")
            ws['B'+str(fila)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B'+str(fila)] = item.producto.unidad_medida.descripcion
            ws.merge_cells('C' + str(fila) + ':F' + str(fila))
            ws['C'+str(fila)].alignment = Alignment(horizontal="center")
            ws['C'+str(fila)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                      top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['C'+str(fila)] = item.producto.descripcion
            ws['G'+str(fila)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['G'+str(fila)] = item.cantidad
            ws['H'+str(fila)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['H'+str(fila)] = item.precio
            ws['I'+str(fila)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['I'+str(fila)] = item.cantidad*item.precio
            detalle_index += 1

        fila_total=31+detalle_index
        ws.merge_cells('G' + str(fila_total) + ':H' + str(fila_total))
        ws['G'+str(fila_total)].alignment = Alignment(horizontal="center")
        ws['G'+str(fila_total)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                            top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G'+str(fila_total)] = 'SUBTOTAL'
        ws['I' + str(fila_total)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I'+str(fila_total)] = orden.subtotal
        ws.merge_cells('G' + str(fila_total+1) + ':H' + str(fila_total+1))
        ws['G' + str(fila_total+1)].alignment = Alignment(horizontal="center")
        ws['G' + str(fila_total+1)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G'+str(fila_total+1)] = 'IMPUESTO 18% IGV'
        ws['I' + str(fila_total+1)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I'+str(fila_total+1)] = orden.impuesto
        ws.merge_cells('G' + str(fila_total+2) + ':H' + str(fila_total+2))
        ws['G' + str(fila_total+2)].alignment = Alignment(horizontal="center")
        ws['G' + str(fila_total+2)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['G'+str(fila_total+2)] = 'TOTAL'
        ws['I' + str(fila_total+2)].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                                  top=Side(border_style="thin"), bottom=Side(border_style="thin"))
        ws['I'+str(fila_total+2)] = orden.total

        ws['B' + str(fila_total + 7)] = 'SON:'
        ws.merge_cells('C' + str(fila_total + 7) + ':I' + str(fila_total + 7))
        ws['C' + str(fila_total + 7)].font = Font(underline="single")
        ws['C' + str(fila_total + 7)].alignment = Alignment(horizontal="center")
        ws['C' + str(fila_total + 7)] = orden.total_letras

        fila_pago=fila_total+9
        ws.merge_cells('E' + str(fila_pago) + ':F' + str(fila_pago))
        ws['E'+str(fila_pago)] = 'FORMA DE PAGO'

        ws.merge_cells('G' + str(fila_pago) + ':I' + str(fila_pago))
        ws['G' + str(fila_pago)].alignment = Alignment(horizontal="center")
        ws['G' + str(fila_pago)].border = Border(bottom=Side(border_style="thin"))
        ws['G'+str(fila_pago)] = orden.forma_pago.descripcion
        ws.merge_cells('E' + str(fila_pago+1) + ':F' + str(fila_pago+1))
        ws['E'+str(fila_pago+1)] = 'BANCO'
        ws.merge_cells('G' + str(fila_pago+1) + ':I' + str(fila_pago+1))
        ws['G' + str(fila_pago+1)].alignment = Alignment(horizontal="center")
        ws['G' + str(fila_pago+1)].border = Border(bottom=Side(border_style="thin"))
        ws['G'+str(fila_pago+1)] = ''

        ws.merge_cells('E' + str(fila_pago+2) + ':F' + str(fila_pago+2))
        ws['E'+str(fila_pago+2)] = 'NÚMERO DE CTA:'
        ws.merge_cells('G' + str(fila_pago + 2) + ':I' + str(fila_pago + 2))
        ws['G' + str(fila_pago+2)].alignment = Alignment(horizontal="center")
        ws['G' + str(fila_pago+2)].border = Border(bottom=Side(border_style="thin"))
        ws['G'+str(fila_pago+2)] = ''

        ws.merge_cells('E' + str(fila_pago+3) + ':F' + str(fila_pago+3))
        ws['E'+str(fila_pago+3)] = 'TIPO DE CUENTA'
        ws.merge_cells('G' + str(fila_pago + 3) + ':I' + str(fila_pago + 3))
        ws['G' + str(fila_pago+3)].alignment = Alignment(horizontal="center")
        ws['G' + str(fila_pago+3)].border = Border(bottom=Side(border_style="thin"))
        ws['G'+str(fila_pago+3)] = ''

        ws.merge_cells('E' + str(fila_pago+4) + ':F' + str(fila_pago+4))
        ws['E'+str(fila_pago+4)] = 'MONEDA'
        ws.merge_cells('G' + str(fila_pago + 4) + ':I' + str(fila_pago + 4))
        ws['G' + str(fila_pago+4)].alignment = Alignment(horizontal="center")
        ws['G' + str(fila_pago+4)].border = Border(bottom=Side(border_style="thin"))
        ws['G'+str(fila_pago+4)] = ''

        ws.merge_cells('E' + str(fila_pago+5) + ':F' + str(fila_pago+5))
        ws['E'+str(fila_pago+5)] = 'CCI'
        ws.merge_cells('G' + str(fila_pago + 5) + ':I' + str(fila_pago + 5))
        ws['G' + str(fila_pago+5)].alignment = Alignment(horizontal="center")
        ws['G' + str(fila_pago+5)].border = Border(bottom=Side(border_style="thin"))
        ws['G'+str(fila_pago+5)] = ''

        fila_firmas=fila_pago+11
        ws.merge_cells('C' + str(fila_firmas) + ':E' + str(fila_firmas))
        ws['C' + str(fila_firmas)].alignment = Alignment(horizontal="center")
        ws['C' + str(fila_firmas)].border = Border(top=Side(border_style="thin"))
        ws['C' + str(fila_firmas)] = 'RESPONSABLE'
        ws.merge_cells('G' + str(fila_firmas) + ':I' + str(fila_firmas))
        ws['G' + str(fila_firmas)].alignment = Alignment(horizontal="center")
        ws['G' + str(fila_firmas)].border = Border(top=Side(border_style="thin"))
        ws['G' + str(fila_firmas)] = 'APROBADO POR'

        nombre_archivo ="ORDEN_DE_COMPRA_N°.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class ReportePDFOrdenServicios(View):  
    
    def cabecera(self,pdf,orden):
        try:
            archivo_imagen = os.path.join(settings.MEDIA_ROOT,str(EMPRESA.logo))
            pdf.drawImage(archivo_imagen, 40, 750, 120, 90,preserveAspectRatio=True)  
        except:
            pdf.drawString(40,800,str(archivo_imagen))
        pdf.setFont("Times-Roman", 14)
        pdf.drawString(230, 800, u"ORDEN DE SERVICIOS")
        pdf.setFont("Times-Roman", 11)
        pdf.drawString(455, 800, u"R.U.C. " + EMPRESA.ruc)
        pdf.setFont("Times-Roman", 13)
        pdf.drawString(250, 780, u"N°"+orden.codigo)
        pdf.setFont("Times-Roman", 10)
        pdf.drawString(430, 780,EMPRESA.distrito + " " + orden.fecha.strftime('%d de %b de %Y'))
        pdf.setFont("Times-Roman", 10)
        cotizacion = orden.cotizacion
        if cotizacion is None:
            proveedor = orden.proveedor
        else:
            proveedor = orden.cotizacion.proveedor
        pdf.drawString(40, 750, u"SEÑOR(ES): "+ proveedor.razon_social)
        pdf.drawString(440, 750, u"R.U.C.: "+ proveedor.ruc)
        direccion = proveedor.direccion
        if len(direccion)>60:
            pdf.drawString(40,730,u"DIRECCIÓN: "+direccion[0:60])
            pdf.drawString(105, 720,direccion[60:])
        else:            
            pdf.drawString(40,730,u"DIRECCIÓN: "+direccion)
        try:
            pdf.drawString(440, 730, u"TELÉFONO: "+ proveedor.telefono)
        except:
            pdf.drawString(440, 730, u"TELÉFONO: -")
        try:
            pdf.drawString(40, 710, u"REFERENCIA: "+orden.cotizacion.requerimiento.codigo+" - "+orden.cotizacion.requerimiento.oficina.nombre)
        except:
            pdf.drawString(40, 710, u"REFERENCIA: " + orden.nombre_informe)
    
        pdf.drawString(40, 690, u"PROCESO: "+orden.proceso)
        pdf.setFont("Times-Roman", 8)
        pdf.drawString(40, 670, u"Sírvase remitirnos según especificaciones que detallamos lo siguiente: ")
        
    def detalle(self,pdf,y,orden):
        encabezados = ('Item', 'Cantidad', u'Descripción','Precio','Total')
        p = ParagraphStyle('parrafos')
        p.alignment = TA_JUSTIFY 
        p.fontSize = 9
        p.fontName="Times-Roman"
        detalles = []
        cont = 0
        
        for detalle in DetalleOrdenServicios.objects.filter(orden=orden):
            try:
                descripcion = detalle.detalle_cotizacion.detalle_requerimiento.producto.descripcion
                if len(descripcion)>58:
                    cont = cont + 1
                detalles.append((detalle.nro_detalle, detalle.cantidad, Paragraph(descripcion,p), detalle.precio,detalle.valor))
            except:
                descripcion = detalle.producto.descripcion
                if len(descripcion)>58:
                    cont = cont + 1
                detalles.append((detalle.nro_detalle, detalle.cantidad, Paragraph(descripcion,p), detalle.precio,detalle.valor))
                
        #detalles = [(detalle.nro_detalle, detalle.cantidad, Paragraph(detalle.servicio.descripcion+'-'+detalle.descripcion,p), detalle.precio,detalle.valor) for detalle in DetalleOrdenServicios.objects.filter(orden=orden)]
        adicionales = [('','','','','')]*(15-cont-len(detalles))
        detalle_orden = Table([encabezados] + detalles + adicionales,colWidths=[0.8 * cm, 1.9 * cm, 11.3* cm, 2 * cm, 2.5 * cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN',(0,0),(4,0),'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black), 
                #('LINEBELOW', (0,1), (5,-1), 0, colors.black),                  
                ('FONTSIZE', (0, 0), (-1, -1), 8),  
                ('ALIGN',(4,1),(-1,-1),'RIGHT'),           
            ]
        ))
        detalle_orden.wrapOn(pdf, 800, 600)
        detalle_orden.drawOn(pdf, 40,y+75)
        #Letras
        total_letras = [("SON: "+ orden.total_letras,'')]
        tabla_total_letras = Table(total_letras,colWidths=[16 * cm, 2.5 * cm])
        tabla_total_letras.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (1, 0), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]
        ))
        tabla_total_letras.wrapOn(pdf, 800, 600)
        tabla_total_letras.drawOn(pdf, 40,y+55)
        
    def otros(self,pdf,y,orden):
        encabezados_otros = ('LUGAR DE ENTREGA', 'PLAZO DE ENTREGA', 'FORMA DE PAGO')
        otros = [('',u" DÍAS","")]
        tabla_otros = Table([encabezados_otros] + otros,colWidths=[6 * cm, 3.5 * cm, 4.5 * cm], rowHeights=[0.6 * cm, 1 * cm])
        tabla_otros.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (2, 1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),                
            ]
        ))
        tabla_otros.wrapOn(pdf, 800, 600)
        tabla_otros.drawOn(pdf, 40,y+5)
        
    def cuadro_total(self,pdf,y,orden):
        pdf.drawString(445, y+40, u"SUB-TOTAL: ")
        pdf.drawString(445, y+20, u"IGV: ")
        pdf.drawString(445, y, u"TOTAL: ")
        total = [[orden.subtotal],[str(orden.impuesto)],[str(orden.total)]]
        tabla_total = Table(total,colWidths=[2.5 * cm])
        tabla_total.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (0, 2), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN',(0,0),(-1,-1),'RIGHT'),
            ]
        ))
        tabla_total.wrapOn(pdf, 800, 600)
        tabla_total.drawOn(pdf, 495,y-2)
        
    def cuadro_observaciones(self,pdf,y,orden):
        p = ParagraphStyle('parrafos')
        p.alignment = TA_JUSTIFY 
        p.fontSize = 10
        p.fontName="Times-Roman"
        obs=Paragraph("Observaciones: "+orden.observaciones,p)
        observaciones = [[obs]]
        tabla_observaciones = Table(observaciones,colWidths=[18.50 * cm], rowHeights=1.8 * cm)
        tabla_observaciones.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (0, 2), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN',(0,0),(-1,-1),'LEFT'),
                ('VALIGN',(0,0),(-1,-1),'TOP'),
            ]
        ))
        tabla_observaciones.wrapOn(pdf, 800, 600)
        tabla_observaciones.drawOn(pdf, 40,y-58)
        
    def afectacion_presupuesta(self,pdf):
        y=320
        pdf.drawString(40, y-90, u"HOJA DE AFECTACIÓN PRESUPUESTAL:")
        p = ParagraphStyle('parrafos')
        p.alignment = TA_JUSTIFY 
        p.fontSize = 8
        p.fontName="Times-Roman"
        lista = ListFlowable([
                          Paragraph("""Consignar el número de la presente Orden de Compra en su Guía de Remisión y Factura. 
                          Facturar a nombre de """ + smart_str(EMPRESA.razon_social),p),
                          Paragraph("El " + smart_str(EMPRESA.razon_social) + """, se reserva el derecho de devolver 
                          la mercaderia, sino se ajusta a las especificaciones requeridas, asimismo de anular la presente 
                          Orden de Compra.""",p),
                          Paragraph("""El pago de toda factura se hará de acuerdo a las condiciones establecidas.""",p)
                          ],bulletType='1'
                         )
        p1=Paragraph("RECIBIDO POR: ",p)        
        pdf.drawString(330, y-150,"FIRMA: ")
        pdf.line(370, y-150, 560, y-150)
        pdf.drawString(330, y-170,"NOMBRE: ")
        pdf.line(370, y-170, 560, y-170)
        pdf.drawString(330, y-190,"DNI: ")
        pdf.line(370, y-190, 560, y-190)
        pdf.setFont("Times-Roman", 6)
        pdf.drawString(525, y-127,"FECHA")
        afectacion = [[(Paragraph("IMPORTANTE:",p), lista),"RECIBIDO POR:"]]
        tabla_afectacion = Table(afectacion,colWidths=[10 * cm, 8.50 * cm])
        tabla_afectacion.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (1, 0), 1, colors.black),
                ('VALIGN',(1,0),(1,0),'TOP'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]
        ))
        tabla_afectacion.wrapOn(pdf, 800, 600)
        tabla_afectacion.drawOn(pdf, 40,y-200)
        fecha = [[' ',' ',' ']]
        tabla_fecha = Table(fecha,colWidths=[0.6 * cm, 0.6 * cm, 0.6 * cm], rowHeights=0.6 * cm)
        tabla_fecha.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 5),                
            ]
        ))
        tabla_fecha.wrapOn(pdf, 800, 600)
        tabla_fecha.drawOn(pdf, 510,y-120)        
        
    def get(self, request, *args, **kwargs):         
        codigo = kwargs['codigo']
        orden = OrdenServicios.objects.get(codigo=codigo)        
        response = HttpResponse(content_type='application/pdf')
        #response['Content-Disposition'] = 'attachment; filename="orden_compra.pdf"'
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)  
        self.cabecera(pdf, orden)
        y=300
        self.detalle(pdf, y, orden)        
        self.otros(pdf, y, orden)
        self.cuadro_total(pdf, y, orden)
        self.cuadro_observaciones(pdf, y, orden)
        self.afectacion_presupuesta(pdf)
        pdf.setFont("Times-Roman", 8)
        pdf.drawString(115, y-250,"Elaborado por")
        pdf.drawString(430, y-250,"Autorizado por")
        pdf.line(70, y-240, 200, y-240)
        pdf.line(390, y-240, 520, y-240)
        pdf.drawCentredString(300, y-280, EMPRESA.direccion())
        pdf.showPage()    
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response
    
class ReportePDFMemorandoConformidadServicio(View):  
    
    def obtener_puesto(self, oficina, conformidad):
        try:
            puesto = Puesto.objects.get(oficina=oficina,
                                        es_jefatura=True,
                                        fecha_inicio__lte=conformidad.fecha,
                                        fecha_fin = None)
        except Puesto.DoesNotExist:
            puesto = Puesto.objects.get(oficina=oficina,
                                        es_jefatura=True,
                                        fecha_inicio__lte=conformidad.fecha,
                                        fecha_fin__gte=conformidad.fecha)
        return puesto
    
    def puesto_superior(self, oficina, conformidad):
        try:
            puesto_superior = Puesto.objects.get(oficina=oficina,
                                                 es_jefatura=True,
                                                 fecha_inicio__lte=conformidad.fecha,
                                                 fecha_fin = None)
        except Puesto.DoesNotExist:
            puesto_superior = Puesto.objects.get(oficina=oficina,
                                                 es_jefatura=True,
                                                 fecha_inicio__lte=conformidad.fecha,
                                                 fecha_fin__gte=conformidad.fecha)            
        return puesto_superior
    
    def cabecera(self,pdf,conformidad):
        try:
            archivo_imagen = os.path.join(settings.MEDIA_ROOT,str(EMPRESA.logo))
            pdf.drawImage(archivo_imagen, 40, 750, 100, 70,preserveAspectRatio=True)   
        except:
            pdf.drawString(40,750,str(archivo_imagen))         
        pdf.setFont("Times-Roman", 14)
        pdf.drawString(130, 750, u"MEMORANDO DE CONFORMIDAD DEL SERVICIO")
        pdf.setFont("Times-Roman", 13)
        pdf.drawString(250, 730, u"N°"+conformidad.codigo)
        pdf.setFont("Times-Roman", 10)
        pdf.drawString(430, 780, EMPRESA.distrito + " " + conformidad.fecha.strftime('%d de %b de %Y'))
        pdf.drawString(475, 710, conformidad.orden_servicios.codigo)
        requerimiento = conformidad.orden_servicios.cotizacion.requerimiento
        gerencia_inmediata = requerimiento.oficina.gerencia
        solicitante = requerimiento.solicitante
        puesto_solicitante = solicitante.puesto
        if puesto_solicitante is None:            
            puesto_solicitante = self.obtener_puesto(requerimiento.oficina, conformidad)
        puesto_jefe_inmediato = self.puesto_superior(requerimiento.oficina, conformidad)
        jefe_inmediato = puesto_jefe_inmediato.trabajador
        y = 690
        if puesto_solicitante.oficina.codigo == 'GGEN':
            puesto_gerente = self.obtener_puesto(CONFIGURACION.administracion, conformidad)            
        elif puesto_solicitante.oficina.codigo == 'GOPE' and not puesto_solicitante.es_jefatura:
            puesto_gerente = self.obtener_puesto(requerimiento.oficina, conformidad)
        else:
            puesto_gerente = self.obtener_puesto(gerencia_inmediata, conformidad)        
        gerente = puesto_gerente.trabajador
        if puesto_gerente.pk == puesto_jefe_inmediato.pk or puesto_jefe_inmediato.pk == puesto_solicitante.pk:
            pdf.drawString(50, y, u"A           :    "+gerente.nombre_completo())
            y = y-20
            pdf.drawString(50, y, u"                   "+puesto_gerente.nombre)
            y = y-20
            pdf.drawString(50, y, u"DE        :     "+puesto_solicitante.trabajador.nombre_completo())
            y = y-20
            pdf.drawString(50, y, u"                   "+puesto_solicitante.nombre)
            y = y-50
        else:
            pdf.drawString(50, y, u"A           :    "+gerente.nombre_completo())
            y = y-20
            pdf.drawString(50, y, u"                   "+puesto_gerente.nombre)
            y = y-20
            pdf.drawString(50, y, u"                   "+jefe_inmediato.nombre_completo())
            y = y-20
            pdf.drawString(50, y, u"                   "+puesto_jefe_inmediato.nombre)
            y = y-20
            pdf.drawString(50, y, u"DE        :     "+puesto_solicitante.trabajador.nombre_completo())
            y = y-20
            pdf.drawString(50, y, u"                   " + puesto_solicitante.nombre)
            y = y-30
        estilo_parrafo = ParagraphStyle('parrafos')
        estilo_parrafo.alignment = TA_JUSTIFY 
        estilo_parrafo.fontSize = 10
        estilo_parrafo.fontName="Times-Roman"
        cadena_parrafo = u"""Mediante el presente comunico a Ud. que el servicio requerido con REQ DE BIENES Y SERV. N° %s, 
        ha sido concluido a satisfacción, según %s, lo que comunicamos para que proceda al pago del servicio correspondiente que 
        se detalla como sigue: """ % (requerimiento.codigo,conformidad.doc_sustento)
        p1=Paragraph(cadena_parrafo,estilo_parrafo)
        p1.wrapOn(pdf, 500, y - 20 )
        p1.drawOn(pdf, 40,  y - 20)
        
    def detalle(self,pdf,y,conformidad):
        encabezados = ('ITEM', 'DETALLE DEL SERVICIO')
        p = ParagraphStyle('parrafos')
        p.alignment = TA_JUSTIFY 
        p.fontSize = 9
        p.fontName="Times-Roman"
        detalles = []
        cont = 0
        for detalle in DetalleConformidadServicio.objects.filter(conformidad=conformidad):
            descripcion = detalle.detalle_orden_servicios.detalle_cotizacion.detalle_requerimiento.producto.descripcion+'-'+detalle.detalle_orden_servicios.detalle_cotizacion.detalle_requerimiento.uso
            if len(descripcion)>58:
                cont = cont + 1
            detalles.append((detalle.nro_detalle, Paragraph(descripcion,p)))
        adicionales = [('','')]*(8-cont-len(detalles))
        detalle_orden = Table([encabezados] + detalles + adicionales,colWidths=[0.8 * cm, 17* cm])
        detalle_orden.setStyle(TableStyle(
            [
                ('ALIGN',(0,0),(1,0),'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black), 
                #('LINEBELOW', (0,1), (5,-1), 0, colors.black),                  
                ('FONTSIZE', (0, 0), (-1, -1), 8),  
                ('ALIGN',(1,1),(-1,-1),'RIGHT'),           
            ]
        ))
        detalle_orden.wrapOn(pdf, 800, 600)
        detalle_orden.drawOn(pdf, 40,y+75)
        
    def firma(self,pdf,x_texto,y_texto,texto,x_ini_linea,x_fin_linea,y_linea):
        pdf.drawString(x_texto, y_texto,texto)
        pdf.line(x_ini_linea, y_linea, x_fin_linea, y_linea)
        
    def get(self, request, *args, **kwargs):         
        codigo = kwargs['codigo']
        conformidad = ConformidadServicio.objects.get(codigo=codigo)        
        response = HttpResponse(content_type='application/pdf')
        #response['Content-Disposition'] = 'attachment; filename="orden_compra.pdf"'
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)  
        self.cabecera(pdf, conformidad)
        y=300
        self.detalle(pdf, y, conformidad)
        pdf.setFont("Times-Roman", 8)
        self.firma(pdf, 170, y-50, "GERENCIA", 120, 265, y-40)
        self.firma(pdf, 330, y-50, "CONFORMIDAD DEL SOLICITANTE", 320, 470, y-40)
        self.firma(pdf, 130, y-150, "CONFORMIDAD JEFE INMEDIATO", 120, 265, y-140)
        self.firma(pdf, 350, y-150, "UNIDAD DE LOGÍSTICA", 320, 470, y-140)
        pdf.drawCentredString(300, y-280, EMPRESA.direccion())
        pdf.showPage()    
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response
    
class ReportePDFSolicitudCotizacion(View):
    
    def cabecera(self,pdf,cotizacion):
        archivo_imagen = os.path.join(settings.MEDIA_ROOT,str(EMPRESA.logo))
        pdf.drawImage(archivo_imagen, 20, 750, 120, 90,preserveAspectRatio=True)  
        pdf.setFont("Times-Roman", 14)
        encabezado = [[u"SOLICITUD DE COTIZACIÓN"]]
        tabla_encabezado = Table(encabezado,colWidths=[8 * cm])
        tabla_encabezado.setStyle(TableStyle(
            [
                ('ALIGN',(0,0),(0,0),'CENTER'),
                ('GRID', (0, 0), (1, 0), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]
        ))
        tabla_encabezado.wrapOn(pdf, 800, 600)
        tabla_encabezado.drawOn(pdf, 200,800)
        pdf.drawString(270, 780, u"N°"+cotizacion.codigo)
        pdf.setFont("Times-Roman", 10)        
        pdf.drawString(40, 750, u"SEÑOR(ES): "+cotizacion.proveedor.razon_social)
        pdf.drawString(440, 750, u"R.U.C.: "+cotizacion.proveedor.ruc)
        direccion = cotizacion.proveedor.direccion
        if len(direccion)>60:
            pdf.drawString(40,730,u"DIRECCIÓN: "+direccion[0:60])
            pdf.drawString(105, 720,direccion[60:])
        else:            
            pdf.drawString(40,730,u"DIRECCIÓN: "+direccion)
        try:
            pdf.drawString(440, 730, u"TELÉFONO: "+cotizacion.proveedor.telefono)
        except:
            pdf.drawString(440, 730, u"TELÉFONO: -")
        pdf.drawString(40, 710, u"FECHA: "+cotizacion.fecha.strftime('%d/%m/%Y'))
        
    def detalle(self,pdf,y,cotizacion):
        encabezados = ('Nro', 'Descripción', 'Unidad','Cantidad')
        detalles = cotizacion.detallecotizacion_set.all()
        lista_detalles = []
        for detalle in detalles:
            tupla_producto = (detalle.nro_detalle, detalle.detalle_requerimiento.producto.descripcion, detalle.detalle_requerimiento.producto.unidad_medida.descripcion, detalle.cantidad)
            lista_detalles.append(tupla_producto)            
        adicionales = [('','','','')]*(15-len(detalles))
        tabla_detalle = Table([encabezados] + lista_detalles + adicionales,colWidths=[1 * cm, 13.5 * cm, 1.5 * cm, 2* cm])
        tabla_detalle.setStyle(TableStyle(
            [
                ('ALIGN',(0,0),(3,0),'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 7),  
                ('ALIGN',(3,1),(-1,-1),'LEFT'),           
            ]
        ))
        tabla_detalle.wrapOn(pdf, 800, 600)
        tabla_detalle.drawOn(pdf, 40,y+80)
        
    def cuadro_observaciones(self,pdf,y,cotizacion):
        p = ParagraphStyle('parrafos')
        p.alignment = TA_JUSTIFY 
        p.fontSize = 8
        p.fontName="Times-Roman"
        obs=Paragraph("OBSERVACIONES: "+cotizacion.observaciones,p)
        observaciones = [[obs]]
        tabla_observaciones = Table(observaciones,colWidths=[18 * cm], rowHeights=1.8 * cm)
        tabla_observaciones.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (0, 2), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN',(0,0),(-1,-1),'LEFT'),
                ('VALIGN',(0,0),(-1,-1),'TOP'),
            ]
        ))
        tabla_observaciones.wrapOn(pdf, 800, 600)
        tabla_observaciones.drawOn(pdf, 40,y+20)
        
    def get(self, request, *args, **kwargs): 
        codigo = kwargs['codigo']
        cotizacion = Cotizacion.objects.get(codigo=codigo)        
        response = HttpResponse(content_type='application/pdf')
        #response['Content-Disposition'] = 'attachment; filename="resume.pdf"'
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer)  
        self.cabecera(pdf, cotizacion)
        y=300
        self.detalle(pdf, y, cotizacion)
        self.cuadro_observaciones(pdf, y, cotizacion)
        ''''
        self.firmas(pdf, y, cotizacion)'''
        pdf.showPage()    
        pdf.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

class ReporteExcelProveedores(TemplateView):
    
    def get(self, request, *args, **kwargs):
        proveedores = Proveedor.objects.all().order_by('ruc')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE PROVEEDORES'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'RUC'
        ws['C3'] = 'RAZON_SOCIAL'
        ws['D3'] = 'DIRECCION'
        ws['E3'] = 'TELEFONO'
        ws['F3'] = 'CORREO'
        ws['G3'] = 'ESTADO_SUNAT'
        ws['H3'] = 'CONDICIÓN'
        ws['I3'] = 'REPRESENTANTE'
        ws['J3'] = 'CIIU'
        ws['K3'] = 'FECHA_ALTA'
        cont=4
        for proveedor in proveedores:
            ws.cell(row=cont,column=2).value = proveedor.ruc
            ws.cell(row=cont,column=3).value = proveedor.razon_social
            ws.cell(row=cont,column=4).value = proveedor.direccion
            ws.cell(row=cont,column=5).value = proveedor.telefono
            ws.cell(row=cont,column=6).value = proveedor.correo
            ws.cell(row=cont,column=7).value = proveedor.estado_sunat
            ws.cell(row=cont,column=8).value = proveedor.condicion
            try:
                ws.cell(row=cont,column=9).value = proveedor.representante.nombre
            except:
                ws.cell(row=cont,column=9).value = '-'
            ws.cell(row=cont,column=10).value = proveedor.ciiu
            ws.cell(row=cont,column=11).value = proveedor.fecha_alta
            ws.cell(row=cont,column=11).number_format = 'dd/mm/yyyy' 
            cont = cont + 1
        nombre_archivo ="ListadoProveedores.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
class ReporteExcelOrdenesServiciosFecha(FormView):
    form_class = FormularioReporteOrdenesFecha
    template_name = "compras/reporte_ordenes.html"    

    def form_valid(self, form):
        data = form.cleaned_data
        tipo_busqueda = data['tipo_busqueda']
        wb = Workbook()
        ws = wb.active
        if tipo_busqueda=='F':
            p_fecha_inicio = data['fecha_inicio']
            p_fecha_final = data['fecha_fin']
            anio = int(p_fecha_inicio[6:])        
            mes = int(p_fecha_inicio[3:5])
            dia = int(p_fecha_inicio[0:2])
            fecha_inicio = datetime.datetime(anio,mes,dia,23,59,59)
            anio = int(p_fecha_final[6:])        
            mes = int(p_fecha_final[3:5])
            dia = int(p_fecha_final[0:2])
            fecha_final = datetime.datetime(anio,mes,dia,23,59,59)            
            ws['B2'] = 'REPORTE DE ORDENES DE SERVICIOS POR FECHA'
            ws.merge_cells('B2:H2')
            ws['B3'] = 'DESDE'
            ws['C3'] = p_fecha_inicio
            ws['C3'].number_format = 'dd/mm/yyyy'
            ws['D3'] = 'HASTA'
            ws['E3'] = p_fecha_final
            ws['F3'].number_format = 'dd/mm/yyyy'        
            ordenes_servicios = OrdenServicios.objects.filter(fecha__range=[fecha_inicio, fecha_final])                                    
        elif tipo_busqueda=='M':
            mes = data['mes'].strip()
            annio = data['annio'].strip()            
            ws['B2'] = 'REPORTE DE ORDENES DE SERVICIOS POR MES'
            ws.merge_cells('B2:H2')
            ws['B3'] = 'MES'
            ws['C3'] = mes
            ws['D3'] = 'AÑO'
            ws['E3'] = annio                    
            ordenes_servicios = OrdenServicios.objects.filter(fecha__month=mes,fecha__year=annio)
        elif tipo_busqueda=='A':
            annio = data['annio'].strip()
            ws['B2'] = 'REPORTE DE ORDENES DE SERVICIOS POR AÑO'
            ws.merge_cells('B2:H2')
            ws['B3'] = 'AÑO'
            ws['C3'] = annio
            ordenes_servicios = OrdenServicios.objects.filter(fecha__year=annio)
        ws['B5'] = 'CODIGO'
        ws['C5'] = 'FECHA'
        ws['D5'] = 'PROVEEDOR'        
        ws['E5'] = 'IMPORTE'
        ws['F5'] = 'FORMA_PAGO'
        ws['G5'] = 'CREADO'
        ws['H5'] = 'ESTADO'
        cont=6
        for orden in ordenes_servicios:
            ws.cell(row=cont,column=2).value = orden.codigo
            ws.cell(row=cont,column=3).value = orden.fecha
            ws.cell(row=cont,column=3).number_format = 'dd/mm/yyyy'
            try:
                ws.cell(row=cont,column=4).value = orden.cotizacion.proveedor.razon_social
            except:
                ws.cell(row=cont,column=4).value = orden.proveedor.razon_social
            ws.cell(row=cont,column=5).value = orden.total
            ws.cell(row=cont,column=6).value = orden.forma_pago.descripcion
            ws.cell(row=cont,column=6).number_format = 'dd/mm/yyyy hh:mm:ss'            
            ws.cell(row=cont,column=7).value = orden.created
            ws.cell(row=cont,column=7).number_format = 'dd/mm/yyyy hh:mm:ss'
            ws.cell(row=cont,column=8).value = orden.get_estado_display()
            cont = cont + 1
        nombre_archivo ="ReporteOrdenesServicio.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
class ReporteExcelOrdenesCompraFecha(FormView):
    form_class = FormularioReporteOrdenesFecha
    template_name = "compras/reporte_ordenes.html"    

    def form_valid(self, form):
        data = form.cleaned_data
        tipo_busqueda = data['tipo_busqueda']
        wb = Workbook()
        ws = wb.active
        if tipo_busqueda=='F':
            p_fecha_inicio = data['fecha_inicio']
            p_fecha_final = data['fecha_fin']
            anio = int(p_fecha_inicio[6:])        
            mes = int(p_fecha_inicio[3:5])
            dia = int(p_fecha_inicio[0:2])
            fecha_inicio = datetime.datetime(anio,mes,dia,23,59,59)
            anio = int(p_fecha_final[6:])        
            mes = int(p_fecha_final[3:5])
            dia = int(p_fecha_final[0:2])
            fecha_final = datetime.datetime(anio,mes,dia,23,59,59)            
            ws['B2'] = 'REPORTE DE ORDENES DE COMPRA POR FECHA'
            ws.merge_cells('B2:H2')
            ws['B3'] = 'DESDE'
            ws['C3'] = p_fecha_inicio
            ws['C3'].number_format = 'dd/mm/yyyy'
            ws['D3'] = 'HASTA'
            ws['E3'] = p_fecha_final
            ws['F3'].number_format = 'dd/mm/yyyy'        
            ordenes_compra = OrdenCompra.objects.filter(fecha__range=[fecha_inicio, fecha_final])                                    
        elif tipo_busqueda=='M':
            mes = data['mes'].strip()
            annio = data['annio'].strip()            
            ws['B2'] = 'REPORTE DE ORDENES DE COMPRA POR MES'
            ws.merge_cells('B2:H2')
            ws['B3'] = 'MES'
            ws['C3'] = mes
            ws['D3'] = 'AÑO'
            ws['E3'] = annio                    
            ordenes_compra = OrdenCompra.objects.filter(fecha__month=mes,fecha__year=annio)
        elif tipo_busqueda=='A':
            annio = data['annio'].strip()
            ws['B2'] = 'REPORTE DE ORDENES DE COMPRA POR AÑO'
            ws.merge_cells('B2:H2')
            ws['B3'] = 'AÑO'
            ws['C3'] = annio
            ordenes_compra = OrdenCompra.objects.filter(fecha__year=annio)
        ws['B5'] = 'CODIGO'
        ws['C5'] = 'FECHA'
        ws['D5'] = 'PROVEEDOR'        
        ws['E5'] = 'IMPORTE'
        ws['F5'] = 'FORMA_PAGO'
        ws['G5'] = 'CREADO'
        ws['H5'] = 'ESTADO'
        cont=6
        for orden_compra in ordenes_compra:
            ws.cell(row=cont,column=2).value = orden_compra.codigo
            ws.cell(row=cont,column=3).value = orden_compra.fecha
            ws.cell(row=cont,column=3).number_format = 'dd/mm/yyyy'
            try:
                ws.cell(row=cont,column=4).value = orden_compra.cotizacion.proveedor.razon_social
            except:
                ws.cell(row=cont,column=4).value = orden_compra.proveedor.razon_social
            ws.cell(row=cont,column=5).value = orden_compra.total
            ws.cell(row=cont,column=6).value = orden_compra.forma_pago.descripcion
            ws.cell(row=cont,column=6).number_format = 'dd/mm/yyyy hh:mm:ss'            
            ws.cell(row=cont,column=7).value = orden_compra.created
            ws.cell(row=cont,column=7).number_format = 'dd/mm/yyyy hh:mm:ss'
            ws.cell(row=cont,column=8).value = orden_compra.get_estado_display()
            cont = cont + 1
        nombre_archivo ="ReporteOrdenesCompra.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class TransferenciaCotizacion(TemplateView):
    template_name = 'compras/transferencia_cotizacion.html'   
    
    def get_context_data(self, **kwargs):
        context = super(TransferenciaCotizacion, self).get_context_data(**kwargs)
        context['cotizaciones'] = Cotizacion.objects.filter(estado = Cotizacion.STATUS.PEND)
        return context
    
class TransferenciaOrdenCompra(TemplateView):
    template_name = 'compras/transferencia_orden_compra.html'   
    
    def get_context_data(self, **kwargs):
        context = super(TransferenciaOrdenCompra, self).get_context_data(**kwargs)
        context['ordenes'] = OrdenCompra.objects.filter(Q(estado=OrdenCompra.STATUS.PEND) | Q(estado = OrdenCompra.STATUS.ING_PARC))
        return context
    
class TransferenciaOrdenServicios(TemplateView):
    template_name = 'compras/transferencia_orden_servicios.html'   
    
    def get_context_data(self, **kwargs):
        context = super(TransferenciaOrdenServicios, self).get_context_data(**kwargs)
        trabajador = self.request.user.trabajador
        context['ordenes'] = OrdenServicios.objects.filter(estado=OrdenServicios.STATUS.PEND)
        return context  