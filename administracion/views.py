# -*- coding: utf-8 -*-
from django.shortcuts import render
from administracion.forms import OficinaForm, TrabajadorForm, PuestoForm, ModificacionPuestoForm, \
    ProfesionForm, NivelAprobacionForm, ProductorForm
from almacen.models import TipoMovimiento
from contabilidad.forms import UploadForm
from django.core.urlresolvers import reverse_lazy
from django.views.generic.edit import FormView, UpdateView, CreateView
from django.views.generic.list import ListView
from administracion.models import Oficina, Trabajador, Puesto, Profesion, \
    NivelAprobacion, Productor
from django.views.generic.base import View, TemplateView
from django.views.generic.detail import DetailView
from django.conf import settings
import csv
from django.http.response import HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.contrib.auth.models import User
from openpyxl import Workbook
from django.http import HttpResponse
import datetime
import os
from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
import simplejson

class Tablero(View):

    def get(self, request, *args, **kwargs):
        lista_notificaciones = []
        cant_oficinas = Oficina.objects.all().count()
        cant_trabajadores = Trabajador.objects.all().count()
        cant_puestos = Puesto.objects.all().count()
        cant_profesiones = Profesion.objects.all().count()
        cant_niveles = NivelAprobacion.objects.all().count()
        if cant_oficinas == 0:
            Oficina.objects.create(codigo='GGEN',
                                   nombre='GERENCIA GENERAL',
                                   es_gerencia = True)
            lista_notificaciones.append("Se ha creado la oficina de GERENCIA GENERAL")
        if cant_trabajadores == 0:
            lista_notificaciones.append("No se ha registrado ningún trabajador")
        if cant_puestos == 0:
            lista_notificaciones.append("No se ha registrado ningún puesto")
        if cant_profesiones == 0:
            lista_notificaciones.append("No se ha registrado ninguna profesión")
        if cant_niveles == 0:
            nivel_logistica = NivelAprobacion.objects.create(descripcion="LOGISTICA")
            nivel_usuario = NivelAprobacion.objects.create(descripcion = "USUARIO",
                                                           nivel_superior = nivel_logistica)
            lista_notificaciones.append("Se han creado los niveles de aprobación básicos")
        context = {'notificaciones': lista_notificaciones}
        return render(request, 'administracion/tablero_administracion.html', context)


class BusquedaReceptorDni(TemplateView):
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            dni = request.GET['dni']
            tipo_movimiento = TipoMovimiento.objects.get(pk=request.GET['tipo_movimiento'])
            if tipo_movimiento.es_venta:
                receptor = Productor.objects.get(dni=dni)
            else:
                receptor = Trabajador.objects.get(dni=dni)
            receptor_json = {}
            receptor_json['dni'] = receptor.dni
            receptor_json['nombre_completo'] = str(receptor.nombre_completo())
            data = simplejson.dumps(receptor_json)
            return HttpResponse(data, 'application/json')

class CargarOficinas(FormView):
    template_name = 'administracion/cargar_oficinas.html'
    form_class = UploadForm

    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT, 'archivos', str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        for fila in dataReader:
            Oficina.objects.get_or_create(codigo=fila[0],
                                          defaults={
                                              'nombre': unicode(fila[1], errors='ignore'),
                                              'dependencia': Oficina.objects.get(codigo=fila[2])},
                                          )
        return HttpResponseRedirect(reverse('administracion:maestro_oficinas'))


class CargarTrabajadores(FormView):
    template_name = 'administracion/cargar_trabajadores.html'
    form_class = UploadForm

    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT, 'archivos', str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        for fila in dataReader:
            usuario, creado = User.objects.get_or_create(username=fila[0],
                                                         defaults={'email': fila[5]}, )
            if creado:
                usuario.set_password('123456789')
                usuario.save()
            trabajador, creado = Trabajador.objects.get_or_create(usuario=usuario,
                                                                  defaults={'dni': fila[1],
                                                                            'apellido_paterno': unicode(fila[2],
                                                                                                        errors='ignore'),
                                                                            'apellido_materno': unicode(fila[3],
                                                                                                        errors='ignore'),
                                                                            'nombres': unicode(fila[4],
                                                                                               errors='ignore')})
        return HttpResponseRedirect(reverse('administracion:maestro_trabajadores'))


class CargarPuestos(FormView):
    template_name = 'administracion/cargar_puestos.html'
    form_class = UploadForm

    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT, 'archivos', str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        for fila in dataReader:
            anio = int(fila[3][6:])
            mes = int(fila[3][3:5])
            dia = int(fila[3][0:2])
            fecha = datetime.datetime(anio, mes, dia)
            if fila[4] == 'SI':
                es_jefatura = True
            else:
                es_jefatura = False
            try:
                puesto, creado = Puesto.objects.get_or_create(nombre=unicode(fila[0], errors='ignore'),
                                                              defaults={'oficina': Oficina.objects.get(
                                                                  codigo=fila[1].strip()),
                                                                        'trabajador': Trabajador.objects.get(
                                                                            dni=fila[2].strip()),
                                                                        'fecha_inicio': fecha,
                                                                        'es_jefatura': es_jefatura})
            except:
                pass
        return HttpResponseRedirect(reverse('administracion:maestro_puestos'))


class CrearNivelAprobacion(CreateView):
    template_name = 'administracion/nivel_aprobacion.html'
    form_class = NivelAprobacionForm

    @method_decorator(
        permission_required('administracion.add_nivelaprobacion', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearNivelAprobacion, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_nivel_aprobacion', args=[self.object.pk])


class CrearProfesion(CreateView):
    template_name = 'administracion/profesion.html'
    form_class = ProfesionForm

    @method_decorator(permission_required('administracion.add_profesion', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearProfesion, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_profesion', args=[self.object.pk])


class CrearOficina(CreateView):
    template_name = 'administracion/oficina.html'
    form_class = OficinaForm

    @method_decorator(permission_required('administracion.add_oficina', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearOficina, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_oficina', args=[self.object.pk])


class CrearTrabajador(CreateView):
    template_name = 'administracion/trabajador.html'
    form_class = TrabajadorForm

    @method_decorator(permission_required('administracion.add_trabajador', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearTrabajador, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_trabajador', args=[self.object.pk])

class CrearProductor(CreateView):
    template_name = 'administracion/productor.html'
    form_class = ProductorForm

    @method_decorator(permission_required('administracion.add_productor', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearProductor, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_productor', args=[self.object.pk])


class CrearPuesto(CreateView):
    template_name = 'administracion/puesto.html'
    form_class = PuestoForm

    @method_decorator(permission_required('administracion.add_puesto', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearPuesto, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_puesto', args=[self.object.pk])


class DetalleOficina(DetailView):
    model = Oficina
    template_name = 'administracion/detalle_oficina.html'


class DetalleTrabajador(DetailView):
    model = Trabajador
    template_name = 'administracion/detalle_trabajador.html'

class DetalleProductor(DetailView):
    model = Productor
    template_name = 'administracion/detalle_productor.html'

class DetallePuesto(DetailView):
    model = Puesto
    template_name = 'administracion/detalle_puesto.html'


class DetalleProfesion(DetailView):
    model = Profesion
    template_name = 'administracion/detalle_profesion.html'


class DetalleNivelAprobacion(DetailView):
    model = NivelAprobacion
    template_name = 'administracion/detalle_nivel_aprobacion.html'


class ListadoOficinas(ListView):
    model = Oficina
    template_name = 'administracion/oficinas.html'
    context_object_name = 'oficinas'
    queryset = Oficina.objects.all().order_by('nombre')


class ListadoTrabajadores(ListView):
    model = Trabajador
    template_name = 'administracion/trabajadores.html'
    context_object_name = 'trabajadores'

class ListadoProductores(ListView):
    model = Productor
    template_name = 'administracion/productores.html'
    context_object_name = 'productores'


class ListadoPuestos(ListView):
    model = Puesto
    template_name = 'administracion/puestos.html'
    context_object_name = 'puestos'
    queryset = Puesto.objects.filter(estado=True)


class ListadoProfesiones(ListView):
    model = Profesion
    template_name = 'administracion/profesiones.html'
    context_object_name = 'profesiones'


class ListadoNivelesAprobacion(ListView):
    model = NivelAprobacion
    template_name = 'administracion/niveles_aprobacion.html'
    context_object_name = 'niveles'


class ModificarNivelAprobacion(UpdateView):
    model = NivelAprobacion
    template_name = 'administracion/nivel_aprobacion.html'
    form_class = NivelAprobacionForm

    @method_decorator(
        permission_required('administracion.change_nivelaprobacion', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarNivelAprobacion, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_nivel_aprobacion', args=[self.object.pk])


class ModificarProfesion(UpdateView):
    model = Profesion
    template_name = 'administracion/profesion.html'
    form_class = ProfesionForm

    @method_decorator(
        permission_required('administracion.change_profesion', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarProfesion, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_profesion', args=[self.object.pk])


class ModificarOficina(UpdateView):
    model = Oficina
    template_name = 'administracion/oficina.html'
    form_class = OficinaForm
    success_url = reverse_lazy('administracion:maestro_oficinas')

    @method_decorator(permission_required('administracion.change_oficina', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarOficina, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_oficina', args=[self.object.pk])


class ModificarTrabajador(UpdateView):
    model = Trabajador
    template_name = 'administracion/trabajador.html'
    form_class = TrabajadorForm

    @method_decorator(
        permission_required('administracion.change_trabajador', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarTrabajador, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_trabajador', args=[self.object.pk])

class ModificarProductor(UpdateView):
    model = Productor
    template_name = 'administracion/productor.html'
    form_class = ProductorForm

    @method_decorator(permission_required('administracion.change_productor', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarProductor, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('administracion:detalle_productor', args=[self.object.pk])

class ModificarPuesto(UpdateView):
    model = Puesto
    template_name = 'administracion/puesto.html'
    form_class = ModificacionPuestoForm

    @method_decorator(permission_required('administracion.change_puesto', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarPuesto, self).dispatch(*args, **kwargs)

    def get_initial(self):
        initial = super(ModificarPuesto, self).get_initial()
        initial['fecha_inicio'] = self.object.fecha_inicio.strftime('%d/%m/%Y')
        if self.object.fecha_fin is not None:
            initial['fecha_fin'] = self.object.fecha_fin.strftime('%d/%m/%Y')
        return initial

    def get_success_url(self):
        return reverse('administracion:detalle_puesto', args=[self.object.pk])


class ReporteExcelOficinas(TemplateView):
    def get(self, request, *args, **kwargs):
        oficinas = Oficina.objects.filter(estado=True).order_by('codigo')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE OFICINAS'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'DEPENDENCIA'
        ws['E3'] = 'GERENCIA'
        cont = 4
        for oficina in oficinas:
            try:
                ws.cell(row=cont, column=2).value = oficina.codigo
                ws.cell(row=cont, column=3).value = oficina.nombre
                ws.cell(row=cont, column=4).value = oficina.dependencia.nombre
                ws.cell(row=cont, column=5).value = oficina.gerencia.nombre
                cont = cont + 1
            except:
                print oficina
        nombre_archivo = "Oficinas.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class ReporteExcelProfesiones(TemplateView):
    def get(self, request, *args, **kwargs):
        profesiones = Profesion.objects.filter(estado=True)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE PROFESIONES'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'ABREVIATURA'
        ws['C3'] = 'DESCRIPCION'
        ws['D3'] = 'ESTADO'
        cont = 4
        for profesion in profesiones:
            ws.cell(row=cont, column=2).value = profesion.abreviatura
            ws.cell(row=cont, column=3).value = profesion.descripcion
            ws.cell(row=cont, column=4).value = profesion.estado
            cont = cont + 1
        nombre_archivo = "Profesiones.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class ReporteExcelPuestos(TemplateView):
    def get(self, request, *args, **kwargs):
        puestos = Puesto.objects.filter(estado=True)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE PUESTOS'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'NOMBRE'
        ws['C3'] = 'OFICINA'
        ws['D3'] = 'TRABAJADOR'
        ws['E3'] = 'FECHA INICIO'
        ws['F3'] = 'FECHA FIN'
        ws['G3'] = 'ES JEFATURA'
        ws['H3'] = 'ESTADO'
        cont = 4
        for puesto in puestos:
            ws.cell(row=cont, column=2).value = puesto.nombre
            ws.cell(row=cont, column=3).value = puesto.oficina.nombre
            ws.cell(row=cont, column=4).value = puesto.trabajador.nombre_completo()
            ws.cell(row=cont, column=5).value = puesto.fecha_inicio.strftime('%d/%m/%Y')
            ws.cell(row=cont, column=6).value = puesto.fecha_fin
            if puesto.es_jefatura:
                ws.cell(row=cont, column=7).value = "SI"
            else:
                ws.cell(row=cont, column=7).value = "NO"
            ws.cell(row=cont, column=8).value = puesto.estado
            cont = cont + 1
        nombre_archivo = "Puestos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class ReporteExcelTrabajadores(TemplateView):
    def get(self, request, *args, **kwargs):
        trabajadores = Trabajador.objects.filter(estado=True)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE TRABAJADORES'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'USUARIO'
        ws['C3'] = 'DNI'
        ws['D3'] = 'APELLIDO_PATERNO'
        ws['E3'] = 'APELLIDO_MATERNO'
        ws['F3'] = 'NOMBRES'
        ws['G3'] = 'EMAIL'
        ws['H3'] = 'ESTADO'
        cont = 4
        for trabajador in trabajadores:
            ws.cell(row=cont, column=2).value = trabajador.usuario.username
            ws.cell(row=cont, column=3).value = trabajador.dni
            ws.cell(row=cont, column=4).value = trabajador.apellido_paterno
            ws.cell(row=cont, column=5).value = trabajador.apellido_materno
            ws.cell(row=cont, column=6).value = trabajador.nombres
            ws.cell(row=cont, column=7).value = trabajador.usuario.email
            ws.cell(row=cont, column=8).value = trabajador.estado
            cont = cont + 1
        nombre_archivo = "Trabajadores.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response