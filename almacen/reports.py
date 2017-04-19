# -*- coding: utf-8 -*-
import math
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Image, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER,TA_LEFT, TA_JUSTIFY, TA_RIGHT
from reportlab.platypus import Table
from reportlab.lib import colors
from reportlab.lib.pagesizes import cm
from reportlab.platypus.flowables import Spacer, PageBreak
from django.conf import settings
import os
from io import BytesIO
from almacen.models import DetalleMovimiento, Kardex, Movimiento
from almacen.settings import EMPRESA, OFICINA_ADMINISTRACION, LOGISTICA
from django.db.models import Sum
import datetime
from productos.models import Producto, GrupoProductos
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Q
class ReporteMovimiento():
    
    def __init__(self, pagesize, movimiento):
        self.movimiento = movimiento
        self.buffer = BytesIO()
        if pagesize == 'A4':
            self.pagesize = A4
        elif pagesize == 'Letter':
            self.pagesize = letter
        self.width, self.height = self.pagesize
        
    def tabla_encabezado(self, styles):
        movimiento = self.movimiento        
        sp = ParagraphStyle('parrafos',
                            alignment = TA_CENTER,
                            fontSize = 14,
                            fontName="Times-Roman")
        try:
            archivo_imagen = os.path.join(settings.MEDIA_ROOT,str(EMPRESA.logo))
            imagen = Image(archivo_imagen, width=90, height=50,hAlign='LEFT')
        except:
            imagen = Paragraph(u"LOGO", sp)
        
        if movimiento.tipo_movimiento.incrementa:
            nota = Paragraph(u"NOTA DE INGRESO N°", sp)
        else:
            nota = Paragraph(u"NOTA DE SALIDA N°", sp)
        id_movimiento = Paragraph(movimiento.id_movimiento, sp)
        fecha = Paragraph("FECHA: "+movimiento.fecha_operacion.strftime('%d/%m/%y'), sp)        
        encabezado = [[imagen,nota,fecha],
                      ['',id_movimiento,'']
                      ]
        tabla_encabezado = Table(encabezado,colWidths=[4 * cm, 9 * cm, 6 * cm])
        tabla_encabezado.setStyle(TableStyle(
            [
                ('VALIGN',(0,0),(2,0),'CENTER'),
                ('VALIGN',(1,1),(2,1),'TOP'),                
                ('SPAN',(0,0),(0,1)),                        
            ]
        ))
        return tabla_encabezado

    def tabla_datos(self, styles):
        movimiento = self.movimiento
        izquierda = ParagraphStyle('parrafos',
                            alignment = TA_LEFT,
                            fontSize = 10,
                            fontName="Times-Roman")
        try:
            if movimiento.referencia.cotizacion is not None:
                proveedor = Paragraph(u"PROVEEDOR: "+movimiento.referencia.cotizacion.proveedor.razon_social,izquierda)
            else:
                proveedor = Paragraph(u"PROVEEDOR: "+movimiento.referencia.proveedor.razon_social,izquierda)
        except:
            proveedor = Paragraph(u"PROVEEDOR:",izquierda)
        operacion = Paragraph(u"OPERACIÓN: "+movimiento.tipo_movimiento.descripcion,izquierda)
        almacen = Paragraph(u"ALMACÉN: "+movimiento.almacen.codigo+"-"+movimiento.almacen.descripcion,izquierda)
        try:
            orden_compra = Paragraph(u"ORDEN DE COMPRA: "+movimiento.referencia.codigo,izquierda)
        except:
            orden_compra = Paragraph(u"REFERENCIA: -",izquierda)
        try:
            documento = Paragraph(u"DOCUMENTO: "+movimiento.tipo_documento.descripcion + " SERIE:" + movimiento.serie + u" NÚMERO:" + movimiento.numero,
                                  izquierda)
        except:
            documento = ""
        try:
            pedido = Paragraph(u"PEDIDO: "+movimiento.pedido.codigo, izquierda)
        except:
            pedido = ""
        encabezado = [[operacion,''],
                      [almacen,''],
                      [proveedor,''],
                      [orden_compra,''],
                      [documento,''],
                      [pedido,'']]
        tabla_datos = Table(encabezado,colWidths=[11 * cm, 9 * cm])
        tabla_datos.setStyle(TableStyle(
            [

            ]
        ))
        return tabla_datos
    
    def tabla_detalle(self):
        movimiento = self.movimiento
        encabezados = ['Item', 'Cantidad', 'Unidad', u'Descripción','Precio','Total']
        detalles = DetalleMovimiento.objects.filter(movimiento=movimiento).order_by('pk')
        sp = ParagraphStyle('parrafos')
        sp.alignment = TA_JUSTIFY 
        sp.fontSize = 8
        sp.fontName="Times-Roman"        
        lista_detalles = []
        for detalle in detalles:
            tupla_producto = [str(detalle.nro_detalle),
                              format(detalle.cantidad,'.5f'),
                              str(detalle.producto.unidad_medida.codigo),
                              detalle.producto.descripcion,
                              format(detalle.precio,'.5f'),
                              format(detalle.valor,'.5f')]
            lista_detalles.append(tupla_producto)
        adicionales = [('','','','','')] * (15-len(lista_detalles))
        tabla_detalle = Table([encabezados] + lista_detalles,colWidths=[1.5 * cm, 2.5 * cm, 1.5 * cm,10* cm, 2 * cm, 2.5 * cm])
        style = TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.black), 
                ('FONTSIZE', (0, 0), (-1, -1), 8),  
                ('ALIGN',(4,0),(-1,-1),'RIGHT'),            
            ]
        )
        tabla_detalle.setStyle(style)
        return tabla_detalle
    
    def tabla_total(self):
        movimiento = self.movimiento
        izquierda = ParagraphStyle('parrafos',
                            alignment = TA_LEFT,
                            fontSize = 10,
                            fontName="Times-Roman")

        texto_total = Paragraph("Total: ",izquierda)
        total = Paragraph(str(round(movimiento.total,2)),izquierda)
        total = [['',texto_total, total]]
        tabla_total = Table(total,colWidths=[15.5 * cm,2 * cm,2.5 * cm])
        tabla_total.setStyle(TableStyle(
            [
                ('GRID', (2, 0), (2, 0), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN',(0,0),(-1,-1),'RIGHT'),
            ]
        ))
        return tabla_total
    
    def tabla_observaciones(self):
        movimiento = self.movimiento
        p = ParagraphStyle('parrafos',
                           alignment = TA_JUSTIFY,
                           fontSize = 8,
                           fontName="Times-Roman")
        obs=Paragraph("OBSERVACIONES: "+movimiento.observaciones,p)
        observaciones = [[obs]]
        tabla_observaciones = Table(observaciones,colWidths=[20 * cm], rowHeights=1.8 * cm)
        tabla_observaciones.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (0, 2), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN',(0,0),(-1,-1),'LEFT'),
                ('VALIGN',(0,0),(-1,-1),'TOP'),
            ]
        ))
        return tabla_observaciones  
    
    def tabla_firmas(self):
        movimiento = self.movimiento
        izquierda = ParagraphStyle('parrafos',
                            alignment = TA_CENTER,
                            fontSize = 8,
                            fontName="Times-Roman")
        nombre_oficina_administracion = Paragraph(OFICINA_ADMINISTRACION.nombre,izquierda)
        nombre_oficina_logistica = Paragraph(LOGISTICA.nombre,izquierda)
        if movimiento.tipo_movimiento.incrementa:            
            total = [[nombre_oficina_administracion,'', nombre_oficina_logistica]]
            tabla_firmas = Table(total,colWidths=[7 * cm,4 * cm,7 * cm])
            tabla_firmas.setStyle(TableStyle(
                [
                    ("LINEABOVE", (0,0), (0,0), 1, colors.black),
                    ("LINEABOVE", (2,0), (2,0), 1, colors.black),
                    ('VALIGN',(0,0),(-1,-1),'TOP'),
                ]
            ))
        else:
            solicitante = Paragraph('SOLICITANTE',izquierda)
            total = [[nombre_oficina_administracion,'',nombre_oficina_logistica,'',solicitante]]
            tabla_firmas = Table(total,colWidths=[5 * cm, 1 * cm, 5 * cm, 1 * cm, 5 * cm])
            tabla_firmas.setStyle(TableStyle(
                [
                    ("LINEABOVE", (0,0), (0,0), 1, colors.black),
                    ("LINEABOVE", (2,0), (2,0), 1, colors.black),
                    ("LINEABOVE", (4,0), (4,0), 1, colors.black),
                    #("LINEABOVE", (2,0), (2,0), 1, colors.black),
                    ('VALIGN',(0,0),(-1,-1),'TOP'),
                ]
            ))
        
        return tabla_firmas
    
    def pie_pagina(self, canvas, doc):
        canvas.saveState()
        canvas.setFont('Times-Roman',10)
        canvas.drawCentredString(300, 20, EMPRESA.direccion())
        canvas.restoreState()
    
    def imprimir(self):
        y=300
        buffer = self.buffer
        izquierda = ParagraphStyle('parrafos',
                                   alignment = TA_LEFT,
                                   fontSize = 10,
                                   fontName="Times-Roman")
        doc = SimpleDocTemplate(buffer,
                                rightMargin=50,
                                leftMargin=50,
                                topMargin=20,
                                bottomMargin=50,
                                pagesize=self.pagesize)
 
        elements = [] 
        styles = getSampleStyleSheet()        
        elements.append(self.tabla_encabezado(styles))
        elements.append(Spacer(1, 0.25 * cm))
        elements.append(self.tabla_datos(styles))        
        elements.append(Spacer(1, 0.25 * cm))
        elements.append(self.tabla_detalle())
        elements.append(Spacer(1,0.25 * cm))
        elements.append(self.tabla_total())   
        elements.append(Spacer(1, 0.25 * cm))     
        elements.append(self.tabla_observaciones())
        elements.append(Spacer(1, 4 * cm))
        elements.append(self.tabla_firmas())        
        doc.build(elements, onFirstPage=self.pie_pagina, onLaterPages=self.pie_pagina)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf


class ReporteKardexPDF():

    def __init__(self, pagesize, desde, hasta, almacen, grupos):
        self.desde = desde
        self.hasta = hasta
        self.almacen = almacen
        self.grupos = grupos
        self.total_paginas = 0
        self.buffer = BytesIO()
        if pagesize == 'A4':
            self.pagesize = landscape(A4)
        elif pagesize == 'Letter':
            self.pagesize = letter
        self.width, self.height = self.pagesize

    def tabla_encabezado(self,valorizado):
        sp = ParagraphStyle('parrafos',
                            alignment=TA_CENTER,
                            fontSize=14,
                            fontName="Times-Roman")
        try:
            archivo_imagen = os.path.join(settings.MEDIA_ROOT, str(EMPRESA.logo))
            imagen = Image(archivo_imagen, width=90, height=50, hAlign='LEFT')
        except:
            imagen = Paragraph(u"LOGO", sp)
        if valorizado:
            titulo = Paragraph(u"REGISTRO DEL INVENTARIO PERMANENTE VALORIZADO", sp)
        else:
            titulo = Paragraph(u"REGISTRO DEL INVENTARIO PERMANENTE EN UNIDADES FÍSICAS", sp)

        encabezado = [[imagen,titulo]]
        tabla_encabezado = Table(encabezado, colWidths=[2 * cm, 23 * cm])
        return tabla_encabezado

    def tabla_encabezado_consolidado(self, grupos):
        sp = ParagraphStyle('parrafos',
                            alignment=TA_CENTER,
                            fontSize=14,
                            fontName="Times-Roman")
        try:
            archivo_imagen = os.path.join(settings.MEDIA_ROOT, str(EMPRESA.logo))
            imagen = Image(archivo_imagen, width=90, height=50, hAlign='LEFT')
        except:
            imagen = Paragraph(u"LOGO", sp)
        if grupos:
            titulo = Paragraph(u"RESUMEN MENSUAL DE ALMACÉN POR GRUPOS Y CUENTAS", sp)
        else:
            titulo = Paragraph(u"RESUMEN MENSUAL DE ALMACÉN POR PRODUCTOS", sp)

        encabezado = [[imagen,titulo]]
        tabla_encabezado = Table(encabezado, colWidths=[2 * cm, 23 * cm])
        style = TableStyle(
            [
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        )
        tabla_encabezado.setStyle(style)
        return tabla_encabezado

    def tabla_detalle_unidades_fisicas(self, producto, desde, hasta, almacen):
        tabla = []
        encab_prim = [u"DOCUMENTO DE TRASLADO, COMPROBANTE DE PAGO,\n DOCUMENTO INTERNO O SIMILAR",
                        "",
                        "",
                        "",
                       u"TIPO DE \n OPERACIÓN \n (TABLA 12)",
                       u"ENTRADAS",
                       u"SALIDAS",
                       u"SALDO FINAL"]
        tabla.append(encab_prim)
        encab_seg = ["FECHA", "TIPO (TABLA 10)","SERIE","NÚMERO","","","",""]
        tabla.append(encab_seg)
        try:
            kardex_inicial = Kardex.objects.filter(producto=producto,
                                                   almacen=almacen,
                                                   fecha_operacion__lt=desde).latest('fecha_operacion')
            cant_saldo_inicial = kardex_inicial.cantidad_total
        except:
            cant_saldo_inicial = 0
        saldo_inicial = [desde.strftime('%d/%m/%Y'),'00','SALDO','INICIAL','16',format(0,'.2f'),format(0,'.2f'),format(cant_saldo_inicial,'.2f')]
        cantidad_total = cant_saldo_inicial
        tabla.append(saldo_inicial)
        listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = producto.obtener_kardex(
            almacen,
            desde,
            hasta)

        for kardex in listado_kardex:
            try:
                tipo_documento = kardex.movimiento.tipo_documento.codigo_sunat
            except:
                tipo_documento = '-'
            try:
                tipo_movimiento = kardex.movimiento.tipo_movimiento.codigo_sunat
            except:
                tipo_movimiento = "-"

            cantidad_total = kardex.cantidad_total

            tabla.append([kardex.fecha_operacion.strftime('%d/%m/%Y'),
                          tipo_documento,
                          kardex.movimiento.serie,
                          kardex.movimiento.numero,
                          tipo_movimiento,
                          format(kardex.cantidad_ingreso,'.2f'),
                          format(kardex.cantidad_salida,'.2f'),
                          format(cantidad_total,'.2f')])
        totales = ['','','','',"TOTALES",format(cantidad_ingreso,'.2f'),format(cantidad_salida,'.2f'),format(cantidad_total,'.2f')]
        tabla.append(totales)

        self.total_paginas += 1;
        total_registros_producto=listado_kardex.count()+2
        if total_registros_producto >11:
            self.total_paginas += int(math.ceil((total_registros_producto-11) / 21.0))

        tabla_detalle = Table(tabla, repeatRows=2, colWidths=[3 * cm, 4 * cm,3 * cm, 3 * cm,3 * cm, 3.5 * cm,3.5 * cm, 3.5 * cm])
        style = TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('TEXTFONT', (0, 1), (-1, 1), 'Times-Roman'),
                ('ALIGN', (5, 1), (7, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (7, 1), 'CENTER'),
                ('VALIGN', (0, 0), (7, 1), 'MIDDLE'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('SPAN', (0, 0), (3, 0)),
                ('SPAN', (4, 0), (4, 1)),
                ('SPAN', (5, 0), (5, 1)),
                ('SPAN', (6, 0), (6, 1)),
                ('SPAN', (7, 0), (7, 1)),
            ]
        )
        tabla_detalle.setStyle(style)
        return tabla_detalle

    def tabla_detalle_consolidado_productos(self, productos):
        almacen = self.almacen
        desde = self.desde
        hasta = self.hasta
        tabla = []
        encabezado1 = ["CODIGO", u"DENOMINACIÓN",u"UNIDAD",u"CUENTA", u"SALDO INICIAL", u"", u"INGRESOS", u"", u"SALIDAS", u"", u"SALDO",
                       u""]
        encabezado2 = [u"", u"",u"", u"", u"CANTIDAD", u"VALOR", u"CANTIDAD", u"VALOR",
                       u"CANTIDAD", u"VALOR", u"CANTIDAD", u"VALOR"]
        tabla.append(encabezado1)
        tabla.append(encabezado2)
        total_cant_saldo_inicial = 0
        total_valor_saldo_inicial = 0
        total_cantidad_ingreso = 0
        total_valor_ingreso = 0
        total_cantidad_salida = 0
        total_valor_salida = 0
        total_cantidad_total = 0
        total_valor_total = 0
        self.total_paginas=int(math.ceil(productos.count()/22.0))
        for producto in productos:
            try:
                kardex_inicial = Kardex.objects.filter(producto=producto,
                                                       almacen=almacen,
                                                       fecha_operacion__lt=desde).latest('fecha_operacion')
                cant_saldo_inicial = kardex_inicial.cantidad_total
                valor_saldo_inicial = kardex_inicial.valor_total
            except:
                cant_saldo_inicial = 0
                valor_saldo_inicial = 0

            listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = producto.obtener_kardex(
                almacen,
                desde,
                hasta)
            cantidad_total = cant_saldo_inicial + cantidad_ingreso - cantidad_salida
            valor_total = valor_saldo_inicial + valor_ingreso - valor_salida

            total_cant_saldo_inicial += cant_saldo_inicial
            total_valor_saldo_inicial += valor_saldo_inicial
            total_cantidad_ingreso += cantidad_ingreso
            total_valor_ingreso += valor_ingreso
            total_cantidad_salida += cantidad_salida
            total_valor_salida += valor_salida
            total_cantidad_total += cantidad_total
            total_valor_total += valor_total

            temp_valor_saldo_inicial = format(valor_saldo_inicial, '.3f')
            if temp_valor_saldo_inicial == '-0.000':
                valor_saldo_inicial = format(abs(valor_saldo_inicial), '.3f')
            else:
                valor_saldo_inicial = format(valor_saldo_inicial, '.3f')

            temp_valor_total = format(valor_total, '.3f')
            if temp_valor_total == '-0.000':
                valor_total = format(abs(valor_total), '.3f')
            else:
                valor_total = format(valor_total, '.3f')

            registro=[producto.codigo,
                      producto.descripcion,
                      producto.unidad_medida.descripcion,
                      producto.grupo_productos.ctacontable,
                      format(cant_saldo_inicial,'.3f'),
                      valor_saldo_inicial,
                      format(cantidad_ingreso,'.3f'),
                      format(valor_ingreso,'.3f'),
                      format(cantidad_salida,'.3f'),
                      format(valor_salida,'.3f'),
                      format(cantidad_total,'.3f'),
                      valor_total]
            tabla.append(registro)

        totales = ["","","",  "TOTALES",
                   format(total_cant_saldo_inicial,'.3f'), format(total_valor_saldo_inicial,'.3f'),
                   format(total_cantidad_ingreso,'.3f'),format(total_valor_ingreso,'.3f'),
                   format(total_cantidad_salida,'.3f'),format(total_valor_salida,'.3f'),
                   format(total_cantidad_total,'.3f'),format(total_valor_total, '.3f')]
        tabla.append(totales)
        tabla_detalle = Table(tabla,repeatRows=2)#,colWidths=[2 * cm, 7 * cm, 2.2 * cm, 1.5 * cm,2.3 * cm, 2.3 * cm,2.3 * cm, 2.4 * cm,2.3 * cm, 2.5 * cm])
        style = TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (11, 1), 'CENTER'),
                ('ALIGN', (4, 2), (11, -1), 'RIGHT'),
                ('TEXTFONT', (0, 0), (-1, -1), 'Times-Roman'),
                ('FONTSIZE', (0, 0), (-1, -1), 6.5),
                ('SPAN', (0, 0), (0, 1)),
                ('SPAN', (1, 0), (1, 1)),
                ('SPAN', (1, 0), (1, 1)),
                ('SPAN', (2, 0), (2, 1)),
                ('SPAN', (3, 0), (3, 1)),
                ('SPAN', (4, 0), (5, 0)),
                ('SPAN', (6, 0), (7, 0)),
                ('SPAN', (8, 0), (9, 0)),
                ('SPAN', (10, 0), (11, 0)),
            ]
        )
        tabla_detalle.setStyle(style)
        return tabla_detalle

    def tabla_detalle_consolidado_grupo(self, grupos):
        almacen = self.almacen
        desde = self.desde
        hasta = self.hasta
        tabla = []
        encabezado1 = ["CODIGO","NOMBRE","CUENTA",u"SALDO INICIAL", u"", u"INGRESOS", u"", u"SALIDAS", u"",u"SALDO",u""]
        encabezado2 = [u"", u"", u"", u"CANTIDAD", u"VALOR", u"CANTIDAD", u"VALOR",
                      u"CANTIDAD", u"VALOR", u"CANTIDAD", u"VALOR"]
        tabla.append(encabezado1)
        tabla.append(encabezado2)
        total_cant_saldo_inicial = 0
        total_valor_saldo_inicial = 0
        total_cantidad_ingreso = 0
        total_valor_ingreso = 0
        total_cantidad_salida = 0
        total_valor_salida = 0
        total_cantidad_total = 0
        total_valor_total = 0
        self.total_paginas = int(math.ceil(grupos.count() / 22.0))
        for grupo in grupos:
            cant_saldo_inicial = 0
            valor_saldo_inicial = 0
            productos = Producto.objects.filter(grupo_productos=grupo)
            for producto in productos:
                try:
                    kardex_inicial = Kardex.objects.filter(producto=producto,
                                                           almacen=almacen,
                                                           fecha_operacion__lt=desde).latest('fecha_operacion')
                    cant_saldo_inicial_producto = kardex_inicial.cantidad_total
                    valor_saldo_inicial_producto = kardex_inicial.valor_total
                except:
                    cant_saldo_inicial_producto = 0
                    valor_saldo_inicial_producto = 0
                cant_saldo_inicial += cant_saldo_inicial_producto
                valor_saldo_inicial += valor_saldo_inicial_producto

            listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = grupo.obtener_kardex(
                almacen,
                desde,
                hasta)
            cantidad_total = cant_saldo_inicial + cantidad_ingreso - cantidad_salida
            valor_total = valor_saldo_inicial + valor_ingreso - valor_salida

            total_cant_saldo_inicial += cant_saldo_inicial
            total_valor_saldo_inicial += valor_saldo_inicial
            total_cantidad_ingreso += cantidad_ingreso
            total_valor_ingreso += valor_ingreso
            total_cantidad_salida += cantidad_salida
            total_valor_salida += valor_salida
            total_cantidad_total += cantidad_total
            total_valor_total += valor_total

            temp_valor_saldo_inicial = format(valor_saldo_inicial, '.5f')
            if temp_valor_saldo_inicial == '-0.00000':
                valor_saldo_inicial = format(abs(valor_saldo_inicial), '.5f')
            else:
                valor_saldo_inicial = format(valor_saldo_inicial, '.5f')

            temp_valor_total = format(valor_total, '.5f')
            if temp_valor_total == '-0.00000':
                valor_total = format(abs(valor_total), '.5f')
            else:
                valor_total = format(valor_total, '.5f')

            registro=[grupo.codigo,
                      grupo.descripcion,
                      grupo.ctacontable.cuenta,
                      format(cant_saldo_inicial,'.5f'),
                      valor_saldo_inicial,
                      format(cantidad_ingreso,'.5f'),
                      format(valor_ingreso,'.5f'),
                      format(cantidad_salida,'.5f'),
                      format(valor_salida,'.5f'),
                      format(cantidad_total,'.5f'),
                      valor_total]
            tabla.append(registro)

        totales = ["", "", "TOTALES",
                   format(total_cant_saldo_inicial,'.5f'), format(total_valor_saldo_inicial,'.5f'),
                   format(total_cantidad_ingreso,'.5f'),format(total_valor_ingreso,'.5f'),
                   format(total_cantidad_salida,'.5f'),format(total_valor_salida,'.5f'),
                   format(total_cantidad_total,'.5f'),format(total_valor_total,'.5f')]
        tabla.append(totales)
        tabla_detalle = Table(tabla,repeatRows=2,
                              colWidths=[1.4 * cm, 7 * cm, 1.8 * cm, 2.2 * cm, 2.3 * cm,2.3 * cm, 2.3 * cm,2.3 * cm, 2.4 * cm,2.3 * cm, 2.5 * cm])
        style = TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (10, 1), 'CENTER'),
                ('ALIGN', (3, 2), (10, -1), 'RIGHT'),
                ('TEXTFONT', (0, 0), (-1, -1), 'Times-Roman'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('SPAN', (0, 0), (0, 1)),
                ('SPAN', (1, 0), (1, 1)),
                ('SPAN', (2, 0), (2, 1)),
                ('SPAN', (3, 0), (4, 0)),
                ('SPAN', (5, 0), (6, 0)),
                ('SPAN', (7, 0), (8, 0)),
                ('SPAN', (9, 0), (10, 0)),
            ]
        )
        tabla_detalle.setStyle(style)
        return tabla_detalle

    def tabla_detalle_valorizado(self, producto, desde, hasta, almacen):
        tabla = []
        encab_prim = [u"DOCUMENTO DE TRASLADO, COMPROBANTE DE PAGO,\n DOCUMENTO INTERNO O SIMILAR",
                        "","","",
                       u"TIPO DE \n OPERACIÓN \n (TABLA 12)",
                       u"ENTRADAS","","",
                       u"SALIDAS","","",
                       u"SALDO FINAL","",""]
        tabla.append(encab_prim)
        encab_seg = ["", "", "", "","",
                      "CANTIDAD", "COSTO\nUNITARIO", "TOTAL",
                      "CANTIDAD", "COSTO\nUNITARIO", "TOTAL",
                      "CANTIDAD", "COSTO\nUNITARIO", "TOTAL"]
        tabla.append(encab_seg)
        encab_terc = ["FECHA", "TIPO\n(TABLA 10)","SERIE","NÚMERO",
                     "","","",
                     "","","",
                     "","",""]
        tabla.append(encab_terc)

        try:
            kardex_inicial = Kardex.objects.filter(producto=producto,
                                                   almacen=almacen,
                                                   fecha_operacion__lt=desde).latest('fecha_operacion')
            cant_saldo_inicial = kardex_inicial.cantidad_total
            precio_saldo_inicial = kardex_inicial.precio_total
            valor_saldo_inicial = kardex_inicial.valor_total
        except:
            cant_saldo_inicial = 0
            precio_saldo_inicial = 0
            valor_saldo_inicial = 0

        cant_saldo_inicial = format(cant_saldo_inicial, '.2f')
        precio_saldo_inicial = format(precio_saldo_inicial, '.2f')
        temp_valor_saldo_inicial = format(valor_saldo_inicial, '.2f')
        if temp_valor_saldo_inicial == '-0.00':
            valor_saldo_inicial = format(abs(valor_saldo_inicial), '.2f')
        else:
            valor_saldo_inicial = format(valor_saldo_inicial, '.2f')

        saldo_inicial = [desde.strftime('%d/%m/%Y'),'00','SALDO','INICIAL','16',
                         format(0,'.2f'),format(0,'.2f'),format(0,'.2f'),
                         format(0,'.2f'),format(0,'.2f'),format(0,'.2f'),
                         cant_saldo_inicial,precio_saldo_inicial,valor_saldo_inicial]
        tabla.append(saldo_inicial)

        cantidad_total = cant_saldo_inicial
        precio_total = precio_saldo_inicial
        valor_total = valor_saldo_inicial

        listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = producto.obtener_kardex(
            almacen,
            desde,
            hasta)

        for kardex in listado_kardex:
            try:
                tipo_documento = kardex.movimiento.tipo_documento.codigo_sunat
            except:
                tipo_documento = '-'
            try:
                tipo_movimiento = kardex.movimiento.tipo_movimiento.codigo_sunat
            except:
                tipo_movimiento = "-"

            cantidad_total = format(kardex.cantidad_total,'.2f')
            precio_total = format(kardex.precio_total,'.2f')
            valor_total = format(kardex.valor_total,'.2f')
            if valor_total == '-0.00':
                valor_total = format(abs(kardex.valor_total), '.2f')

            tabla.append([kardex.fecha_operacion.strftime('%d/%m/%Y'),
                          tipo_documento,
                          kardex.movimiento.serie,
                          kardex.movimiento.numero,
                          tipo_movimiento,
                          format(kardex.cantidad_ingreso,'.2f'),
                          format(kardex.precio_ingreso, '.2f'),
                          format(kardex.valor_ingreso, '.2f'),
                          format(kardex.cantidad_salida,'.2f'),
                          format(kardex.precio_salida, '.2f'),
                          format(kardex.valor_salida, '.2f'),
                          cantidad_total,
                          precio_total,
                          valor_total])

        totales = ['','','','',"TOTALES",
                   format(cantidad_ingreso,'.2f'),"",format(valor_ingreso,'.2f'),
                   format(cantidad_salida,'.2f'),"",format(valor_salida,'.2f'),
                   cantidad_total,precio_total,valor_total]
        tabla.append(totales)

        self.total_paginas += 1;
        total_registros_producto = listado_kardex.count() + 2 
        if total_registros_producto > 10:
            self.total_paginas += int(math.ceil((total_registros_producto - 10) / 20.0))

        tabla_detalle = Table(tabla,repeatRows=3)
        style = TableStyle(
            [
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('TEXTFONT', (0, 1), (-1, 1), 'Times-Roman'),
                ('ALIGN', (5, 1), (13, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (13, 2), 'CENTER'),
                ('VALIGN', (0, 0), (13, 2), 'MIDDLE'),
                ('FONTSIZE', (0, 0), (-1, -1), 8.5),
                ('SPAN', (0, 0), (3, 1)),
                ('SPAN', (4, 0), (4, 2)),
                ('SPAN', (5, 0), (7, 0)),
                ('SPAN', (8, 0), (10, 0)),
                ('SPAN', (11, 0), (13, 0)),
                ('SPAN', (5, 1), (5, 2)),
                ('SPAN', (6, 1), (6, 2)),
                ('SPAN', (7, 1), (7, 2)),
                ('SPAN', (8, 1), (8, 2)),
                ('SPAN', (9, 1), (9, 2)),
                ('SPAN', (10, 1), (10, 2)),
                ('SPAN', (11, 1), (11, 2)),
                ('SPAN', (12, 1), (12, 2)),
                ('SPAN', (13, 1), (13, 2)),
            ]
        )
        tabla_detalle.setStyle(style)
        return tabla_detalle

    def imprimir_formato_sunat_unidades_fisicas_producto(self, producto):
        desde = self.desde
        hasta = self.hasta
        almacen = self.almacen
        buffer = self.buffer
        self.valorizado = False
        izquierda = ParagraphStyle('parrafos',
                                   alignment=TA_LEFT,
                                   fontSize=11,
                                   fontName="Times-Roman")
        doc = SimpleDocTemplate(buffer,
                                rightMargin=50,
                                leftMargin=50,
                                topMargin=100,
                                bottomMargin=50,
                                pagesize=self.pagesize)

        elements = []
        periodo = Paragraph("PERIODO: " + desde.strftime('%d/%m/%Y') + ' - ' + hasta.strftime('%d/%m/%Y'), izquierda)
        elements.append(periodo)
        elements.append(Spacer(1, 0.25 * cm))
        ruc = Paragraph(u"RUC:" + EMPRESA.ruc, izquierda)
        elements.append(ruc)
        elements.append(Spacer(1, 0.25 * cm))
        razon_social = Paragraph(u"APELLIDOS Y NOMBRES, DENOMINACIÓN O RAZÓN SOCIAL: " + EMPRESA.razon_social, izquierda)
        elements.append(razon_social)
        elements.append(Spacer(1, 0.25 * cm))
        direccion = Paragraph(u"ESTABLECIMIENTO (1): " + EMPRESA.direccion(), izquierda)
        elements.append(direccion)
        elements.append(Spacer(1, 0.25 * cm))
        codigo = Paragraph(u"CÓDIGO DE LA EXISTENCIA: " + producto.codigo, izquierda)
        elements.append(codigo)
        elements.append(Spacer(1, 0.25 * cm))
        tipo = Paragraph(u"TIPO: B - EXISTENCIA", izquierda)
        """tipo = Paragraph(u"TIPO (TABLA 5): " + producto.tipo_existencia.codigo_sunat + " - " + producto.tipo_existencia.descripcion,
                         izquierda)"""
        elements.append(tipo)
        elements.append(Spacer(1, 0.25 * cm))
        descripcion = Paragraph(u"DESCRIPCIÓN: " + producto.descripcion, izquierda)
        elements.append(descripcion)
        elements.append(Spacer(1, 0.25 * cm))
        unidad = Paragraph(u"CÓDIGO DE LA UNIDAD DE MEDIDA (TABLA 6): " + producto.unidad_medida.codigo_sunat + " - " + producto.unidad_medida.descripcion,
                           izquierda)
        elements.append(unidad)
        elements.append(Spacer(1, 0.25 * cm))
        unidad = Paragraph(u"MÉTODO DE VALUACIÓN: PEPS",
                           izquierda)
        elements.append(unidad)
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(self.tabla_detalle_unidades_fisicas(producto, desde, hasta, almacen))
        doc.build(elements, onFirstPage=self._header, onLaterPages=self._header)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf

    def imprimir_formato_sunat_valorizado_producto(self, producto):
        desde = self.desde
        hasta = self.hasta
        almacen = self.almacen
        buffer = self.buffer
        self.valorizado = True
        izquierda = ParagraphStyle('parrafos',
                                   alignment=TA_LEFT,
                                   fontSize=12,
                                   fontName="Times-Roman")
        doc = SimpleDocTemplate(buffer,
                                rightMargin=50,
                                leftMargin=50,
                                topMargin=100,
                                bottomMargin=50,
                                pagesize=self.pagesize)

        elements = []
        periodo = Paragraph("PERIODO: " + desde.strftime('%d/%m/%Y') + ' - ' + hasta.strftime('%d/%m/%Y'), izquierda)
        elements.append(periodo)
        elements.append(Spacer(1, 0.25 * cm))
        ruc = Paragraph(u"RUC:" + EMPRESA.ruc, izquierda)
        elements.append(ruc)
        elements.append(Spacer(1, 0.25 * cm))
        razon_social = Paragraph(u"APELLIDOS Y NOMBRES, DENOMINACIÓN O RAZÓN SOCIAL: " + EMPRESA.razon_social, izquierda)
        elements.append(razon_social)
        elements.append(Spacer(1, 0.25 * cm))
        direccion = Paragraph(u"ESTABLECIMIENTO (1): " + EMPRESA.direccion(), izquierda)
        elements.append(direccion)
        elements.append(Spacer(1, 0.25 * cm))
        codigo = Paragraph(u"CÓDIGO DE LA EXISTENCIA: " + producto.codigo, izquierda)
        elements.append(codigo)
        elements.append(Spacer(1, 0.25 * cm))
        tipo = Paragraph(u"TIPO: B - EXISTENCIA", izquierda)
        """tipo = Paragraph(u"TIPO (TABLA 5): " + producto.tipo_existencia.codigo_sunat + " - " + producto.tipo_existencia.descripcion,
                         izquierda)"""
        elements.append(tipo)
        elements.append(Spacer(1, 0.25 * cm))
        descripcion = Paragraph(u"DESCRIPCIÓN: " + producto.descripcion, izquierda)
        elements.append(descripcion)
        elements.append(Spacer(1, 0.25 * cm))
        unidad = Paragraph(u"CÓDIGO DE LA UNIDAD DE MEDIDA (TABLA 6): " + producto.unidad_medida.codigo_sunat + " - " + producto.unidad_medida.descripcion,
                           izquierda)
        elements.append(unidad)
        elements.append(Spacer(1, 0.25 * cm))
        unidad = Paragraph(u"MÉTODO DE VALUACIÓN: PEPS",
                           izquierda)
        elements.append(unidad)
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(self.tabla_detalle_valorizado(producto, desde, hasta, almacen))
        doc.build(elements, onFirstPage=self._header, onLaterPages=self._header)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf

    def imprimir_formato_sunat_unidades_fisicas_todos(self):
        desde = self.desde
        hasta = self.hasta
        almacen = self.almacen
        buffer = self.buffer
        self.valorizado = False
        izquierda = ParagraphStyle('parrafos',
                                   alignment=TA_LEFT,
                                   fontSize=12,
                                   fontName="Times-Roman")
        doc = SimpleDocTemplate(buffer,
                                rightMargin=50,
                                leftMargin=50,
                                topMargin=100,
                                bottomMargin=50,
                                pagesize=self.pagesize)

        elements = []
        productos_kardex = Kardex.objects.exclude(cantidad_ingreso=0,cantidad_salida=0).order_by().values('producto').distinct()
        productos = Producto.objects.filter(pk__in=productos_kardex).order_by('descripcion')
        for producto in productos:
            periodo = Paragraph("PERIODO: " + desde.strftime('%d/%m/%Y') + ' - ' + hasta.strftime('%d/%m/%Y'), izquierda)
            elements.append(periodo)
            elements.append(Spacer(1, 0.25 * cm))
            ruc = Paragraph(u"RUC:" + EMPRESA.ruc, izquierda)
            elements.append(ruc)
            elements.append(Spacer(1, 0.25 * cm))
            razon_social = Paragraph(u"APELLIDOS Y NOMBRES, DENOMINACIÓN O RAZÓN SOCIAL: " + EMPRESA.razon_social, izquierda)
            elements.append(razon_social)
            elements.append(Spacer(1, 0.25 * cm))
            direccion = Paragraph(u"ESTABLECIMIENTO (1): " + EMPRESA.direccion(), izquierda)
            elements.append(direccion)
            elements.append(Spacer(1, 0.25 * cm))
            codigo = Paragraph(u"CÓDIGO DE LA EXISTENCIA: " + producto.codigo, izquierda)
            elements.append(codigo)
            elements.append(Spacer(1, 0.25 * cm))
            tipo = Paragraph(u"TIPO: B - EXISTENCIA", izquierda)
            """tipo = Paragraph(u"TIPO (TABLA 5): " + producto.tipo_existencia.codigo_sunat + " - " + producto.tipo_existencia.descripcion,
                             izquierda)"""
            elements.append(tipo)
            elements.append(Spacer(1, 0.25 * cm))
            descripcion = Paragraph(u"DESCRIPCIÓN: " + producto.descripcion, izquierda)
            elements.append(descripcion)
            elements.append(Spacer(1, 0.25 * cm))
            unidad = Paragraph(u"CÓDIGO DE LA UNIDAD DE MEDIDA (TABLA 6): " + producto.unidad_medida.codigo_sunat + " - " + producto.unidad_medida.descripcion,
                               izquierda)
            elements.append(unidad)
            elements.append(Spacer(1, 0.25 * cm))
            unidad = Paragraph(u"MÉTODO DE VALUACIÓN: PEPS",
                               izquierda)
            elements.append(unidad)
            elements.append(Spacer(1, 0.5 * cm))
            elements.append(self.tabla_detalle_unidades_fisicas(producto, desde, hasta, almacen))
            elements.append(PageBreak())
        doc.build(elements, onFirstPage=self._header, onLaterPages=self._header)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf

    def imprimir_formato_consolidado_productos(self):
        buffer = self.buffer
        doc = SimpleDocTemplate(buffer,
                                rightMargin=50,
                                leftMargin=50,
                                topMargin=100,
                                bottomMargin=50,
                                pagesize=self.pagesize)

        elements = []
        productos = Producto.objects.all().order_by('descripcion')
        elements.append(self.tabla_detalle_consolidado_productos(productos))
        doc.build(elements, onFirstPage=self._header_footer, onLaterPages=self._header_footer)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf

    def _header_footer(self, canvas, doc):
        canvas.saveState()
        sp = ParagraphStyle('parrafos',
                            alignment=TA_CENTER,
                            fontSize=14,
                            fontName="Times-Roman")
        try:
            archivo_imagen = os.path.join(settings.MEDIA_ROOT, str(EMPRESA.logo))
            imagen = Image(archivo_imagen, width=90, height=50, hAlign='LEFT')
        except:
            imagen = Paragraph(u"LOGO", sp)
        ruc_empresa = "RUC: " + EMPRESA.ruc
        if self.grupos:
            titulo = Paragraph(u"RESUMEN MENSUAL DE ALMACÉN POR GRUPOS Y CUENTAS", sp)
        else:
            titulo = Paragraph(u"RESUMEN MENSUAL DE ALMACÉN", sp)
        periodo = "PERIODO: " + self.desde.strftime('%d/%m/%Y') + ' - ' + self.hasta.strftime('%d/%m/%Y')
        pagina  = u"Página " + str(doc.page) + " de " + str(self.total_paginas)
        encabezado = [[imagen, titulo, pagina],[ruc_empresa,periodo,""]]
        tabla_encabezado = Table(encabezado, colWidths=[3 * cm, 20 * cm, 3 * cm])
        style = TableStyle(
            [
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        )
        tabla_encabezado.setStyle(style)
        tabla_encabezado.wrapOn(canvas, 50, 510)
        tabla_encabezado.drawOn(canvas, 50, 510)
        canvas.restoreState()

    def _header(self, canvas, doc):
        canvas.saveState()

        sp = ParagraphStyle('parrafos',
                            alignment=TA_CENTER,
                            fontSize=14,
                            fontName="Times-Roman")
        try:
            archivo_imagen = os.path.join(settings.MEDIA_ROOT, str(EMPRESA.logo))
            imagen = Image(archivo_imagen, width=90, height=50, hAlign='LEFT')
        except:
            imagen = Paragraph(u"LOGO", sp)
        ruc_empresa = "RUC: " + EMPRESA.ruc
        if self.valorizado:
            titulo = Paragraph(u"REGISTRO DEL INVENTARIO PERMANENTE VALORIZADO", sp)
        else:
            titulo = Paragraph(u"REGISTRO DEL INVENTARIO PERMANENTE EN UNIDADES FÍSICAS",sp)
        pagina = u"Página " + str(doc.page) + " de " + str(self.total_paginas)
        encabezado = [[imagen, titulo, pagina], [ruc_empresa, "", ""]]
        tabla_encabezado = Table(encabezado, colWidths=[3 * cm, 20 * cm, 3 * cm])
        style = TableStyle(
            [
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        )
        tabla_encabezado.setStyle(style)
        tabla_encabezado.wrapOn(canvas, 50, 510)
        tabla_encabezado.drawOn(canvas, 50, 510)
        canvas.restoreState()

    def imprimir_formato_consolidado_grupos(self):
        buffer = self.buffer
        doc = SimpleDocTemplate(buffer,
                                rightMargin=50,
                                leftMargin=50,
                                topMargin=100,
                                bottomMargin=50,
                                pagesize=self.pagesize)

        elements = []
        grupos = GrupoProductos.objects.filter(estado=True,
                                               son_productos=True).order_by('descripcion')
        elements.append(self.tabla_detalle_consolidado_grupo(grupos))

        doc.build(elements, onFirstPage=self._header_footer, onLaterPages=self._header_footer)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf

    def imprimir_formato_sunat_valorizado_todos(self):
        desde = self.desde
        hasta = self.hasta
        almacen = self.almacen
        buffer = self.buffer
        self.valorizado = True
        izquierda = ParagraphStyle('parrafos',
                                   alignment=TA_LEFT,
                                   fontSize=12,
                                   fontName="Times-Roman")
        doc = SimpleDocTemplate(buffer,
                                rightMargin=50,
                                leftMargin=50,
                                topMargin=100,
                                bottomMargin=50,
                                pagesize=self.pagesize)

        elements = []
        productos_kardex = Kardex.objects.exclude(cantidad_ingreso=0,
                                                  cantidad_salida=0).order_by().values('producto').distinct()
        productos = Producto.objects.filter(pk__in=productos_kardex).order_by('descripcion')
        for producto in productos:
            periodo = Paragraph("PERIODO: " + desde.strftime('%d/%m/%Y') + ' - ' + hasta.strftime('%d/%m/%Y'), izquierda)
            elements.append(periodo)
            elements.append(Spacer(1, 0.25 * cm))
            ruc = Paragraph(u"RUC:" + EMPRESA.ruc, izquierda)
            elements.append(ruc)
            elements.append(Spacer(1, 0.25 * cm))
            razon_social = Paragraph(u"APELLIDOS Y NOMBRES, DENOMINACIÓN O RAZÓN SOCIAL: " + EMPRESA.razon_social, izquierda)
            elements.append(razon_social)
            elements.append(Spacer(1, 0.25 * cm))
            direccion = Paragraph(u"ESTABLECIMIENTO (1): " + EMPRESA.direccion(), izquierda)
            elements.append(direccion)
            elements.append(Spacer(1, 0.25 * cm))
            codigo = Paragraph(u"CÓDIGO DE LA EXISTENCIA: " + producto.codigo, izquierda)
            elements.append(codigo)
            elements.append(Spacer(1, 0.25 * cm))
            tipo = Paragraph(u"TIPO: B - EXISTENCIA",izquierda)
            """tipo = Paragraph(u"TIPO (TABLA 5): " + producto.tipo_existencia.codigo_sunat + " - " + producto.tipo_existencia.descripcion,
                             izquierda)"""
            elements.append(tipo)
            elements.append(Spacer(1, 0.25 * cm))
            descripcion = Paragraph(u"DESCRIPCIÓN: " + producto.descripcion, izquierda)
            elements.append(descripcion)
            elements.append(Spacer(1, 0.25 * cm))
            unidad = Paragraph(u"CÓDIGO DE LA UNIDAD DE MEDIDA (TABLA 6): " + producto.unidad_medida.codigo_sunat + " - " + producto.unidad_medida.descripcion,
                               izquierda)
            elements.append(unidad)
            elements.append(Spacer(1, 0.25 * cm))
            unidad = Paragraph(u"MÉTODO DE VALUACIÓN: PEPS",
                               izquierda)
            elements.append(unidad)
            elements.append(Spacer(1, 0.5 * cm))
            elements.append(self.tabla_detalle_valorizado(producto, desde, hasta, almacen))
            elements.append(PageBreak())
        doc.build(elements, onFirstPage=self._header, onLaterPages=self._header)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf


class ReporteKardexExcel():

    def obtener_formato_sunat_unidades_fisicas_producto(self, producto, desde, hasta, almacen):
        wb = Workbook()
        ws = wb.active
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 15
        ws.column_dimensions["L"].width = 15
        ws.column_dimensions["M"].width = 15
        ws.column_dimensions["N"].width = 15
        ws.column_dimensions["O"].width = 15

        ws['D1'] = u'REGISTRO DEL INVENTARIO PERMANENTE EN UNIDADES FÍSICAS'
        ws.merge_cells('D1:G1')
        ws['B3'] = "PERIODO: "+ desde.strftime('%d/%m/%Y')+' - '+ hasta.strftime('%d/%m/%Y')
        ws.merge_cells('B3:E3')
        ws['B4'] = u"RUC:" + EMPRESA.ruc
        ws.merge_cells('B4:E4')
        ws['B5'] = u"APELLIDOS Y NOMBRES, DENOMINACIÓN O RAZÓN SOCIAL: " + EMPRESA.razon_social
        ws.merge_cells('B5:K5')
        ws['B6'] = u"ESTABLECIMIENTO (1): " + EMPRESA.direccion()
        ws.merge_cells('B6:E6')
        ws['B7'] = u"CÓDIGO DE LA EXISTENCIA: " + producto.codigo
        ws.merge_cells('B7:E7')
        ws['B8'] = u"TIPO (TABLA 5): " + producto.tipo_existencia.codigo_sunat +" - "+ producto.tipo_existencia.descripcion
        ws.merge_cells('B8:G8')
        ws['B9'] = u"DESCRIPCIÓN: " + producto.descripcion
        ws.merge_cells('B9:E9')
        ws['B10'] = u"CÓDIGO DE LA UNIDAD DE MEDIDA (TABLA 6): " + producto.unidad_medida.codigo +" - "+ producto.unidad_medida.descripcion
        ws.merge_cells('B10:H10')
        ws['B11'] = u"MÉTODO DE VALUACIÓN: PEPS"
        ws.merge_cells('B11:E11')

        ws['B13'] = u"DOCUMENTO DE TRASLADO, COMPROBANTE DE PAGO,\n DOCUMENTO INTERNO O SIMILAR"
        ws['B13'].border = thin_border
        ws.merge_cells('B13:E14')
        ws['B15'] = u"FECHA"
        ws['B15'].border = thin_border
        ws['C15'] = u"TIPO (TABLA 10)"
        ws['C15'].border = thin_border
        ws['D15'] = u"SERIE"
        ws['D15'].border = thin_border
        ws['E15'] = u"NÚMERO"
        ws['E15'].border = thin_border
        ws['F13'] = u"TIPO DE \n OPERACIÓN \n (TABLA 12)"
        ws['F13'].border = thin_border
        ws.merge_cells('F13:F15')

        ws['G13'] = u"ENTRADAS"
        ws['G13'].border = thin_border
        ws.merge_cells('G13:G15')
        ws['H13'] = u"SALIDAS"
        ws['H13'].border = thin_border
        ws.merge_cells('H13:H15')
        ws['I13'] = u"SALDO FINAL"
        ws['I13'].border = thin_border
        ws.merge_cells('I13:I15')

        try:
            kardex_inicial = Kardex.objects.filter(producto=producto,
                                                   almacen=almacen,
                                                   fecha_operacion__lt=desde).latest('fecha_operacion')
            cant_saldo_inicial = kardex_inicial.cantidad_total
        except:
            cant_saldo_inicial = 0
        cont = 16
        ws.cell(row=cont, column=2).border = thin_border
        ws.cell(row=cont, column=3).value = '00'
        ws.cell(row=cont, column=3).border = thin_border
        ws.cell(row=cont, column=4).value = 'SALDO'
        ws.cell(row=cont, column=4).border = thin_border
        ws.cell(row=cont, column=5).value = 'INICIAL'
        ws.cell(row=cont, column=5).border = thin_border
        ws.cell(row=cont, column=6).value = '16'
        ws.cell(row=cont, column=6).border = thin_border
        ws.cell(row=cont, column=7).value = 0
        ws.cell(row=cont, column=7).number_format = '#.00000'
        ws.cell(row=cont, column=7).border = thin_border
        ws.cell(row=cont, column=8).value = 0
        ws.cell(row=cont, column=8).number_format = '#.00000'
        ws.cell(row=cont, column=8).border = thin_border
        ws.cell(row=cont, column=9).value = cant_saldo_inicial
        ws.cell(row=cont, column=9).number_format = '#.00000'
        ws.cell(row=cont, column=9).border = thin_border

        listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = producto.obtener_kardex(
            almacen,
            desde,
            hasta)
        for kardex in listado_kardex:
            cont = cont + 1
            ws.cell(row=cont, column=2).value = kardex.fecha_operacion.strftime('%d/%m/%Y')
            ws.cell(row=cont, column=2).border = thin_border
            try:
                ws.cell(row=cont, column=3).value = kardex.movimiento.tipo_documento.codigo_sunat
            except:
                ws.cell(row=cont, column=3).value = '-'
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=4).value = kardex.movimiento.serie
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=5).value = kardex.movimiento.numero
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=6).value = kardex.movimiento.tipo_movimiento.codigo_sunat
            ws.cell(row=cont, column=6).border = thin_border
            ws.cell(row=cont, column=7).value = kardex.cantidad_ingreso
            ws.cell(row=cont, column=7).number_format = '#.00000'
            ws.cell(row=cont, column=7).border = thin_border
            ws.cell(row=cont, column=8).value = kardex.cantidad_salida
            ws.cell(row=cont, column=8).number_format = '#.00000'
            ws.cell(row=cont, column=8).border = thin_border
            ws.cell(row=cont, column=9).value = kardex.cantidad_total
            ws.cell(row=cont, column=9).number_format = '#.00000'
            ws.cell(row=cont, column=9).border = thin_border
        cont = cont + 1
        ws.cell(row=cont, column=6).value = "TOTALES"
        ws.cell(row=cont, column=6).border = thin_border
        ws.cell(row=cont, column=7).value = cantidad_ingreso
        ws.cell(row=cont, column=7).number_format = '#.00000'
        ws.cell(row=cont, column=7).border = thin_border
        ws.cell(row=cont, column=8).value = cantidad_salida
        ws.cell(row=cont, column=8).number_format = '#.00000'
        ws.cell(row=cont, column=8).border = thin_border
        ws.cell(row=cont, column=9).value = ""
        ws.cell(row=cont, column=9).number_format = '#.00000'
        ws.cell(row=cont, column=9).border = thin_border
        return wb

    def obtener_formato_normal_producto(self, producto, desde, hasta, almacen):
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 12
        ws.column_dimensions["K"].width = 12
        ws.column_dimensions["L"].width = 12
        ws.column_dimensions["M"].width = 15
        ws['E1'] = u'Almacén: ' + almacen.descripcion
        ws.merge_cells('E1:G1')
        ws['H1'] = 'Periodo: ' + desde.strftime('%d/%m/%Y') + ' - ' + hasta.strftime('%d/%m/%Y')
        ws.merge_cells('H1:J1')
        cont = 3
        ws.cell(row=cont, column=2).value = 'Codigo: ' + producto.codigo
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=3)
        ws.cell(row=cont, column=4).value = u" Denominación: " + producto.descripcion
        ws.merge_cells(start_row=cont, start_column=4, end_row=cont, end_column=10)
        ws.cell(row=cont, column=11).value = " Unidad: " + producto.unidad_medida.descripcion
        ws.merge_cells(start_row=cont, start_column=11, end_row=cont, end_column=12)
        cont = cont + 1
        try:
            kardex_inicial = Kardex.objects.filter(producto = producto,
                                            almacen = almacen,
                                            fecha_operacion__lt = desde).latest('fecha_operacion')
            cant_saldo_inicial = kardex_inicial.cantidad_total
            valor_saldo_inicial = kardex_inicial.valor_total
        except:
            cant_saldo_inicial = 0
            valor_saldo_inicial = 0
        ws.cell(row=cont,column=8).value = "SALDO INICIAL:"
        ws.merge_cells(start_row=cont, start_column=8, end_row=cont, end_column=9)
        ws.cell(row=cont,column=10).value = "Cantidad: "
        ws.cell(row=cont,column=11).value = cant_saldo_inicial
        ws.cell(row=cont,column=11).number_format = '#.00000'
        ws.cell(row=cont,column=12).value = "Valor: "
        ws.cell(row=cont,column=13).value = valor_saldo_inicial
        ws.cell(row=cont,column=13).number_format = '#.00000'
        ws['B5'] = 'FECHA'
        ws['C5'] = 'NRO_DOC'
        ws['D5']= 'TIPO_MOV'
        ws['E5'] = 'CANT. ENT'
        ws['F5'] = 'PRE. ENT'
        ws['G5'] = 'VALOR. ENT'
        ws['H5'] = 'CANT. SAL'
        ws['I5'] = 'PRE. SAL'
        ws['J5'] = 'VALOR. SAL'
        ws['K5'] = 'CANT. TOT'
        ws['L5'] = 'PRE. TOT'
        ws['M5'] = 'VALOR. TOT'
        cont = cont + 2
        listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = producto.obtener_kardex(almacen,desde,hasta)
        if len(listado_kardex)>0:
            for kardex in listado_kardex:
                ws.cell(row=cont,column=2).value = kardex.fecha_operacion
                ws.cell(row=cont,column=2).number_format = 'dd/mm/yyyy'
                ws.cell(row=cont,column=3).value = kardex.movimiento.id_movimiento
                ws.cell(row=cont,column=4).value = kardex.movimiento.tipo_movimiento.codigo
                ws.cell(row=cont,column=5).value = kardex.cantidad_ingreso
                ws.cell(row=cont, column=5).number_format = '#.00000'
                ws.cell(row=cont,column=6).value = kardex.precio_ingreso
                ws.cell(row=cont,column=6).number_format = '#.00000'
                ws.cell(row=cont,column=7).value = kardex.valor_ingreso
                ws.cell(row=cont,column=7).number_format = '#.00000'
                ws.cell(row=cont,column=8).value = kardex.cantidad_salida
                ws.cell(row=cont, column=8).number_format = '#.00000'
                ws.cell(row=cont,column=9).value = kardex.precio_salida
                ws.cell(row=cont,column=9).number_format = '#.00000'
                ws.cell(row=cont,column=10).value = kardex.valor_salida
                ws.cell(row=cont,column=10).number_format = '#.00000'
                ws.cell(row=cont,column=11).value = kardex.cantidad_total
                ws.cell(row=cont, column=11).number_format = '#.00000'
                ws.cell(row=cont,column=12).value = kardex.precio_total
                ws.cell(row=cont,column=12).number_format = '#.00000'
                ws.cell(row=cont,column=13).value = kardex.valor_total
                ws.cell(row=cont,column=13).number_format = '#.00000'
                cont = cont + 1
            ws.cell(row=cont,column=5).value = cantidad_ingreso
            ws.cell(row=cont,column=7).value = valor_ingreso
            ws.cell(row=cont,column=7).number_format = '#.00000'
            ws.cell(row=cont,column=8).value = cantidad_salida
            ws.cell(row=cont,column=10).value = valor_salida
            ws.cell(row=cont,column=10).number_format = '#.00000'
            ws.cell(row=cont,column=11).value = kardex.cantidad_total
            ws.cell(row=cont,column=13).value = kardex.valor_total
            ws.cell(row=cont,column=13).number_format = '#.00000'
            cont = cont + 2
        else:
            ws.cell(row=cont,column=5).value = 0
            ws.cell(row=cont,column=6).value = 0
            ws.cell(row=cont,column=6).number_format = '#.00000'
            ws.cell(row=cont,column=7).value = 0
            ws.cell(row=cont,column=7).number_format = '#.00000'
            ws.cell(row=cont,column=8).value = 0
            ws.cell(row=cont,column=9).value = 0
            ws.cell(row=cont,column=9).number_format = '#.00000'
            ws.cell(row=cont,column=10).value = 0
            ws.cell(row=cont,column=10).number_format = '#.00000'
            ws.cell(row=cont,column=11).value = 0
            ws.cell(row=cont,column=12).value = 0
            ws.cell(row=cont,column=12).number_format = '#.00000'
            ws.cell(row=cont,column=13).value = 0
            ws.cell(row=cont,column=13).number_format = '#.00000'
            cont = cont + 1
            ws.cell(row=cont,column=5).value = 0
            ws.cell(row=cont,column=7).value = 0
            ws.cell(row=cont,column=7).number_format = '#.00000'
            ws.cell(row=cont,column=8).value = 0
            ws.cell(row=cont,column=10).value = 0
            ws.cell(row=cont,column=10).number_format = '#.00000'
            ws.cell(row=cont,column=11).value = cant_saldo_inicial
            ws.cell(row=cont,column=13).value = valor_saldo_inicial
            ws.cell(row=cont,column=13).number_format = '#.00000'
            cont = cont + 2
        return wb


    def obtener_formato_sunat_valorizado_producto(self, producto, desde, hasta, almacen):
        wb = Workbook()
        ws = wb.active
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 15
        ws.column_dimensions["L"].width = 15
        ws.column_dimensions["M"].width = 15
        ws.column_dimensions["N"].width = 15
        ws.column_dimensions["O"].width = 15

        ws['H1'] = u'REGISTRO DE INVENTARIO PERMANENTE VALORIZADO'
        ws.merge_cells('H1:K1')
        ws['B3'] = "PERIODO: "+ desde.strftime('%d/%m/%Y')+' - '+ hasta.strftime('%d/%m/%Y')
        ws.merge_cells('B3:E3')
        ws['B4'] = u"RUC:" + EMPRESA.ruc
        ws.merge_cells('B4:E4')
        ws['B5'] = u"APELLIDOS Y NOMBRES, DENOMINACIÓN O RAZÓN SOCIAL: " + EMPRESA.razon_social
        ws.merge_cells('B5:K5')
        ws['B6'] = u"ESTABLECIMIENTO (1): " + EMPRESA.direccion()
        ws.merge_cells('B6:E6')
        ws['B7'] = u"CÓDIGO DE LA EXISTENCIA: " + producto.codigo
        ws.merge_cells('B7:E7')
        ws['B8'] = u"TIPO (TABLA 5): " + producto.tipo_existencia.codigo_sunat +" - "+ producto.tipo_existencia.descripcion
        ws.merge_cells('B8:G8')
        ws['B9'] = u"DESCRIPCIÓN: " + producto.descripcion
        ws.merge_cells('B9:E9')
        ws['B10'] = u"CÓDIGO DE LA UNIDAD DE MEDIDA (TABLA 6): " + producto.unidad_medida.codigo +" - "+ producto.unidad_medida.descripcion
        ws.merge_cells('B10:H10')
        ws['B11'] = u"MÉTODO DE VALUACIÓN: PEPS"
        ws.merge_cells('B11:E11')

        ws['B13'] = u"DOCUMENTO DE TRASLADO, COMPROBANTE DE PAGO,\n DOCUMENTO INTERNO O SIMILAR"
        ws['B13'].border = thin_border
        ws.merge_cells('B13:E14')
        ws['B15'] = u"FECHA"
        ws['B15'].border = thin_border
        ws['C15'] = u"TIPO (TABLA 10)"
        ws['C15'].border = thin_border
        ws['D15'] = u"SERIE"
        ws['D15'].border = thin_border
        ws['E15'] = u"NÚMERO"
        ws['E15'].border = thin_border
        ws['F13'] = u"TIPO DE \n OPERACIÓN \n (TABLA 12)"
        ws['F13'].border = thin_border
        ws.merge_cells('F13:F15')

        ws['G13'] = u"ENTRADAS"
        ws['G13'].border = thin_border
        ws['I13'].border = thin_border
        ws.merge_cells('G13:I13')
        ws['G14'] = u"CANTIDAD"
        ws['G14'].border = thin_border
        ws.merge_cells('G14:G15')
        ws['H14'] = u"COSTO UNITARIO"
        ws['H14'].border = thin_border
        ws.merge_cells('H14:H15')
        ws['I14'] = u"COSTO TOTAL"
        ws['I14'].border = thin_border
        ws.merge_cells('I14:I15')

        ws['J13'] = u"SALIDAS"
        ws['J13'].border = thin_border
        ws['L13'].border = thin_border
        ws.merge_cells('J13:L13')
        ws['J14'] = u"CANTIDAD"
        ws['J14'].border = thin_border
        ws.merge_cells('J14:J15')
        ws['K14'] = u"COSTO UNITARIO"
        ws['K14'].border = thin_border
        ws.merge_cells('K14:K15')
        ws['L14'] = u"COSTO TOTAL"
        ws['L14'].border = thin_border
        ws.merge_cells('L14:L15')

        ws['M13'] = u"SALDO FINAL"
        ws['M13'].border = thin_border
        ws['O13'].border = thin_border
        ws.merge_cells('M13:O13')
        ws['M14'] = u"CANTIDAD"
        ws['M14'].border = thin_border
        ws.merge_cells('M14:M15')
        ws['N14'] = u"COSTO UNITARIO"
        ws['N14'].border = thin_border
        ws.merge_cells('N14:N15')
        ws['O14'] = u"COSTO TOTAL"
        ws['O14'].border = thin_border
        ws.merge_cells('O14:O15')

        try:
            kardex_inicial = Kardex.objects.filter(producto=producto,
                                                   almacen=almacen,
                                                   fecha_operacion__lt=desde).latest('fecha_operacion')
            cant_saldo_inicial = kardex_inicial.cantidad_total
            valor_saldo_inicial = kardex_inicial.valor_total
        except:
            cant_saldo_inicial = 0
            valor_saldo_inicial = 0
        cont = 16
        ws.cell(row=cont, column=2).border = thin_border
        ws.cell(row=cont, column=3).value = '00'
        ws.cell(row=cont, column=3).border = thin_border
        ws.cell(row=cont, column=4).value = 'SALDO'
        ws.cell(row=cont, column=4).border = thin_border
        ws.cell(row=cont, column=5).value = 'INICIAL'
        ws.cell(row=cont, column=5).border = thin_border
        ws.cell(row=cont, column=6).value = '16'
        ws.cell(row=cont, column=6).border = thin_border
        ws.cell(row=cont, column=7).value = 0
        ws.cell(row=cont, column=7).number_format = '#.00000'
        ws.cell(row=cont, column=7).border = thin_border
        ws.cell(row=cont, column=8).value = 0
        ws.cell(row=cont, column=8).number_format = '#.00000'
        ws.cell(row=cont, column=9).border = thin_border
        ws.cell(row=cont, column=9).value = 0
        ws.cell(row=cont, column=9).number_format = '#.00000'
        ws.cell(row=cont, column=10).value = 0
        ws.cell(row=cont, column=10).number_format = '#.00000'
        ws.cell(row=cont, column=10).border = thin_border
        ws.cell(row=cont, column=11).value = 0
        ws.cell(row=cont, column=11).number_format = '#.00000'
        ws.cell(row=cont, column=11).border = thin_border
        ws.cell(row=cont, column=12).value = 0
        ws.cell(row=cont, column=12).number_format = '#.00000'
        ws.cell(row=cont, column=12).border = thin_border
        ws.cell(row=cont, column=13).value = cant_saldo_inicial
        ws.cell(row=cont, column=13).number_format = '#.00000'
        ws.cell(row=cont, column=13).border = thin_border
        try:
            ws.cell(row=cont, column=14).value = valor_saldo_inicial / cant_saldo_inicial
        except:
            ws.cell(row=cont, column=14).value = 0
        ws.cell(row=cont, column=14).number_format = '#.00000'
        ws.cell(row=cont, column=14).border = thin_border
        ws.cell(row=cont, column=15).value = valor_saldo_inicial
        ws.cell(row=cont, column=15).number_format = '#.00000'
        ws.cell(row=cont, column=15).border = thin_border

        listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = producto.obtener_kardex(
            almacen,
            desde,
            hasta)
        for kardex in listado_kardex:
            cont = cont + 1
            ws.cell(row=cont, column=2).value = kardex.fecha_operacion.strftime('%d/%m/%Y')
            ws.cell(row=cont, column=2).border = thin_border
            try:
                ws.cell(row=cont, column=3).value = kardex.movimiento.tipo_documento.codigo_sunat
            except:
                ws.cell(row=cont, column=3).value = '-'
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=4).value = kardex.movimiento.serie
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=5).value = kardex.movimiento.numero
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=6).value = kardex.movimiento.tipo_movimiento.codigo_sunat
            ws.cell(row=cont, column=6).border = thin_border
            ws.cell(row=cont, column=7).value = kardex.cantidad_ingreso
            ws.cell(row=cont, column=7).number_format = '#.00000'
            ws.cell(row=cont, column=7).border = thin_border
            ws.cell(row=cont, column=8).value = kardex.precio_ingreso
            ws.cell(row=cont, column=8).number_format = '#.00000'
            ws.cell(row=cont, column=8).border = thin_border
            ws.cell(row=cont, column=9).value = kardex.valor_ingreso
            ws.cell(row=cont, column=9).number_format = '#.00000'
            ws.cell(row=cont, column=9).border = thin_border
            ws.cell(row=cont, column=10).value = kardex.cantidad_salida
            ws.cell(row=cont, column=10).number_format = '#.00000'
            ws.cell(row=cont, column=10).border = thin_border
            ws.cell(row=cont, column=11).value = kardex.precio_salida
            ws.cell(row=cont, column=11).number_format = '#.00000'
            ws.cell(row=cont, column=11).border = thin_border
            ws.cell(row=cont, column=12).value = kardex.valor_salida
            ws.cell(row=cont, column=12).number_format = '#.00000'
            ws.cell(row=cont, column=12).border = thin_border
            ws.cell(row=cont, column=13).value = kardex.cantidad_total
            ws.cell(row=cont, column=13).number_format = '#.00000'
            ws.cell(row=cont, column=13).border = thin_border
            ws.cell(row=cont, column=14).value = kardex.precio_total
            ws.cell(row=cont, column=14).number_format = '#.00000'
            ws.cell(row=cont, column=14).border = thin_border
            ws.cell(row=cont, column=15).value = kardex.valor_total
            ws.cell(row=cont, column=15).number_format = '#.00000'
            ws.cell(row=cont, column=15).border = thin_border
        cont = cont + 1

        ws.cell(row=cont, column=6).value = "TOTALES"
        ws.cell(row=cont, column=6).border = thin_border
        ws.cell(row=cont, column=7).value = cantidad_ingreso
        ws.cell(row=cont, column=7).number_format = '#.00000'
        ws.cell(row=cont, column=7).border = thin_border
        ws.cell(row=cont, column=8).value = ""
        ws.cell(row=cont, column=8).number_format = '#.00000'
        ws.cell(row=cont, column=8).border = thin_border
        ws.cell(row=cont, column=9).value = valor_ingreso
        ws.cell(row=cont, column=9).number_format = '#.00000'
        ws.cell(row=cont, column=9).border = thin_border
        ws.cell(row=cont, column=10).value = cantidad_salida
        ws.cell(row=cont, column=10).number_format = '#.00000'
        ws.cell(row=cont, column=10).border = thin_border
        ws.cell(row=cont, column=11).value = ""
        ws.cell(row=cont, column=11).number_format = '#.00000'
        ws.cell(row=cont, column=11).border = thin_border
        ws.cell(row=cont, column=12).value = valor_salida
        ws.cell(row=cont, column=12).number_format = '#.00000'
        ws.cell(row=cont, column=12).border = thin_border
        ws.cell(row=cont, column=13).value = ""
        ws.cell(row=cont, column=13).number_format = '#.00000'
        ws.cell(row=cont, column=13).border = thin_border
        ws.cell(row=cont, column=14).value = ""
        ws.cell(row=cont, column=14).number_format = '#.00000'
        ws.cell(row=cont, column=14).border = thin_border
        ws.cell(row=cont, column=15).value = ""
        ws.cell(row=cont, column=15).number_format = '#.00000'
        ws.cell(row=cont, column=15).border = thin_border
        return wb


    def obtener_formato_sunat_unidades_fisicas_excel_por_producto(self, ws, thin_border, cont, producto, desde, hasta, almacen):
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 15
        ws.column_dimensions["L"].width = 15
        ws.column_dimensions["M"].width = 15
        ws.column_dimensions["N"].width = 15
        ws.column_dimensions["O"].width = 15
        ws.cell(row=cont, column=4).value = u'REGISTRO DEL INVENTARIO PERMANENTE EN UNIDADES FÍSICAS'
        ws.merge_cells(start_row=cont, start_column=4, end_row=cont, end_column=8)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = "PERIODO: " + desde.strftime('%d/%m/%Y') + ' - ' + hasta.strftime('%d/%m/%Y')
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"RUC:" + EMPRESA.ruc
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"APELLIDOS Y NOMBRES, DENOMINACIÓN O RAZÓN SOCIAL: " + EMPRESA.razon_social
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=11)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"ESTABLECIMIENTO (1): " + EMPRESA.direccion()
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"CÓDIGO DE LA EXISTENCIA: " + producto.codigo
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"TIPO (TABLA 5): " + producto.tipo_existencia.codigo_sunat + " - " + producto.tipo_existencia.descripcion
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=7)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"DESCRIPCIÓN: " + producto.descripcion
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"CÓDIGO DE LA UNIDAD DE MEDIDA (TABLA 6): " + producto.unidad_medida.codigo + " - " + producto.unidad_medida.descripcion
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=8)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"MÉTODO DE VALUACIÓN: PEPS"
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 2
        ws.cell(row=cont, column=2).value = u"DOCUMENTO DE TRASLADO, COMPROBANTE DE PAGO,\n DOCUMENTO INTERNO O SIMILAR"
        ws.cell(row=cont, column=2).border = thin_border
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont + 1, end_column=5)
        ws.cell(row=cont, column=6).value = u"TIPO DE \n OPERACIÓN \n (TABLA 12)"
        ws.cell(row=cont, column=6).border = thin_border
        ws.merge_cells(start_row=cont, start_column=6, end_row=cont + 2, end_column=6)
        ws.cell(row=cont, column=7).value = u"ENTRADAS"
        ws.cell(row=cont, column=7).border = thin_border
        ws.merge_cells(start_row=cont, start_column=7, end_row=cont + 2, end_column=7)
        ws.cell(row=cont, column=8).value = u"SALIDAS"
        ws.cell(row=cont, column=8).border = thin_border
        ws.merge_cells(start_row=cont, start_column=8, end_row=cont + 2, end_column=8)
        ws.cell(row=cont, column=9).value = u"SALDO FINAL"
        ws.cell(row=cont, column=9).border = thin_border
        ws.merge_cells(start_row=cont, start_column=9, end_row=cont + 2, end_column=9)
        cont = cont + 2
        ws.cell(row=cont, column=2).value = u"FECHA"
        ws.cell(row=cont, column=2).border = thin_border
        ws.cell(row=cont, column=3).value = u"TIPO (TABLA 10)"
        ws.cell(row=cont, column=3).border = thin_border
        ws.cell(row=cont, column=4).value = u"SERIE"
        ws.cell(row=cont, column=4).border = thin_border
        ws.cell(row=cont, column=5).value = u"NÚMERO"
        ws.cell(row=cont, column=5).border = thin_border

        try:
            kardex_inicial = Kardex.objects.filter(producto=producto,
                                                   almacen=almacen,
                                                   fecha_operacion__lt=desde).latest('fecha_operacion')
            cant_saldo_inicial = kardex_inicial.cantidad_total
        except:
            cant_saldo_inicial = 0
        cont = cont + 1
        ws.cell(row=cont, column=2).border = thin_border
        ws.cell(row=cont, column=3).value = '00'
        ws.cell(row=cont, column=3).border = thin_border
        ws.cell(row=cont, column=4).value = 'SALDO'
        ws.cell(row=cont, column=4).border = thin_border
        ws.cell(row=cont, column=5).value = 'INICIAL'
        ws.cell(row=cont, column=5).border = thin_border
        ws.cell(row=cont, column=6).value = '16'
        ws.cell(row=cont, column=6).border = thin_border
        ws.cell(row=cont, column=7).value = 0
        ws.cell(row=cont, column=7).number_format = '#.00000'
        ws.cell(row=cont, column=7).border = thin_border
        ws.cell(row=cont, column=8).value = 0
        ws.cell(row=cont, column=8).number_format = '#.00000'
        ws.cell(row=cont, column=8).border = thin_border
        ws.cell(row=cont, column=9).value = cant_saldo_inicial
        ws.cell(row=cont, column=9).number_format = '#.00000'
        ws.cell(row=cont, column=9).border = thin_border

        listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = producto.obtener_kardex(almacen, desde, hasta)
        for kardex in listado_kardex:
            cont = cont + 1
            ws.cell(row=cont, column=2).value = kardex.fecha_operacion.strftime('%d/%m/%Y')
            ws.cell(row=cont, column=2).border = thin_border
            try:
                ws.cell(row=cont, column=3).value = kardex.movimiento.tipo_documento.codigo_sunat
            except:
                ws.cell(row=cont, column=3).value = '-'
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=4).value = kardex.movimiento.serie
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=5).value = kardex.movimiento.numero
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=6).value = kardex.movimiento.tipo_movimiento.codigo_sunat
            ws.cell(row=cont, column=6).border = thin_border
            ws.cell(row=cont, column=7).value = kardex.cantidad_ingreso
            ws.cell(row=cont, column=7).number_format = '#.00000'
            ws.cell(row=cont, column=7).border = thin_border
            ws.cell(row=cont, column=8).value = kardex.cantidad_salida
            ws.cell(row=cont, column=8).number_format = '#.00000'
            ws.cell(row=cont, column=8).border = thin_border
            ws.cell(row=cont, column=9).value = kardex.cantidad_total
            ws.cell(row=cont, column=9).number_format = '#.00000'
            ws.cell(row=cont, column=9).border = thin_border
        cont = cont + 1
        ws.cell(row=cont, column=6).value = "TOTALES"
        ws.cell(row=cont, column=6).border = thin_border
        ws.cell(row=cont, column=7).value = cantidad_ingreso
        ws.cell(row=cont, column=7).number_format = '#.00000'
        ws.cell(row=cont, column=7).border = thin_border
        ws.cell(row=cont, column=8).value = cantidad_salida
        ws.cell(row=cont, column=8).number_format = '#.00000'
        ws.cell(row=cont, column=8).border = thin_border
        ws.cell(row=cont, column=9).value = ""
        ws.cell(row=cont, column=9).number_format = '#.00000'
        ws.cell(row=cont, column=9).border = thin_border
        return ws

    def obtener_formato_sunat_unidades_fisicas_todos(self, desde, hasta, almacen):
        productos = Producto.objects.all().order_by('descripcion')
        wb = Workbook()
        ws = wb.active
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cont = 1
        for producto in productos:
            ws.title = producto.codigo
            self.obtener_formato_sunat_unidades_fisicas_excel_por_producto(ws,thin_border,cont,producto,desde,hasta,almacen)
            ws = wb.create_sheet("Hoja")
        return wb

    def obtener_formato_sunat_valorizado_excel_por_producto(self, ws, thin_border, cont, producto, desde, hasta, almacen):
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 15
        ws.column_dimensions["L"].width = 15
        ws.column_dimensions["M"].width = 15
        ws.column_dimensions["N"].width = 15
        ws.column_dimensions["O"].width = 15
        ws.cell(row=cont, column=4).value = u'REGISTRO DE INVENTARIO PERMANENTE VALORIZADO'
        ws.merge_cells(start_row=cont, start_column=4, end_row=cont, end_column=8)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = "PERIODO: " + desde.strftime('%d/%m/%Y') + ' - ' + hasta.strftime('%d/%m/%Y')
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"RUC:" + EMPRESA.ruc
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"APELLIDOS Y NOMBRES, DENOMINACIÓN O RAZÓN SOCIAL: " + EMPRESA.razon_social
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=11)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"ESTABLECIMIENTO (1): " + EMPRESA.direccion()
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"CÓDIGO DE LA EXISTENCIA: " + producto.codigo
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"TIPO (TABLA 5): " + producto.tipo_existencia.codigo_sunat + " - " + producto.tipo_existencia.descripcion
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=7)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"DESCRIPCIÓN: " + producto.descripcion
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"CÓDIGO DE LA UNIDAD DE MEDIDA (TABLA 6): " + producto.unidad_medida.codigo + " - " + producto.unidad_medida.descripcion
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=8)
        cont = cont + 1
        ws.cell(row=cont, column=2).value = u"MÉTODO DE VALUACIÓN: PEPS"
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=5)
        cont = cont + 2

        ws.cell(row=cont, column=2).value = u"DOCUMENTO DE TRASLADO, COMPROBANTE DE PAGO,\n DOCUMENTO INTERNO O SIMILAR"
        ws.cell(row=cont, column=2).border = thin_border
        ws.merge_cells(start_row=cont, start_column=2, end_row=cont + 1, end_column=5)
        ws.cell(row=cont, column=6).value = u"TIPO DE \n OPERACIÓN \n (TABLA 12)"
        ws.cell(row=cont, column=6).border = thin_border
        ws.merge_cells(start_row=cont, start_column=6, end_row=cont + 2, end_column=6)
        ws.cell(row=cont, column=7).value = u"ENTRADAS"
        ws.cell(row=cont, column=7).border = thin_border
        ws.merge_cells(start_row=cont, start_column=7, end_row=cont, end_column=9)
        ws.cell(row=cont, column=10).value = u"SALIDAS"
        ws.cell(row=cont, column=10).border = thin_border
        ws.merge_cells(start_row=cont, start_column=10, end_row=cont, end_column=12)
        ws.cell(row=cont, column=13).value = u"SALDO FINAL"
        ws.merge_cells(start_row=cont, start_column=13, end_row=cont, end_column=15)
        ws.cell(row=cont, column=13).border = thin_border
        cont = cont + 2
        ws.cell(row=cont, column=2).value = u"FECHA"
        ws.cell(row=cont, column=2).border = thin_border
        ws.cell(row=cont, column=3).value = u"TIPO (TABLA 10)"
        ws.cell(row=cont, column=3).border = thin_border
        ws.cell(row=cont, column=4).value = u"SERIE"
        ws.cell(row=cont, column=4).border = thin_border
        ws.cell(row=cont, column=5).value = u"NÚMERO"
        ws.cell(row=cont, column=5).border = thin_border

        ws.cell(row=cont - 1, column=7).value = u"CANTIDAD"
        ws.merge_cells(start_row=cont - 1, start_column=7, end_row=cont, end_column=7)
        ws.cell(row=cont - 1, column=7).border = thin_border
        ws.cell(row=cont - 1, column=8).value = u"COSTO UNITARIO"
        ws.merge_cells(start_row=cont - 1, start_column=8, end_row=cont, end_column=8)
        ws.cell(row=cont - 1, column=8).border = thin_border
        ws.cell(row=cont - 1, column=9).value = u"COSTO TOTAL"
        ws.merge_cells(start_row=cont - 1, start_column=9, end_row=cont, end_column=9)
        ws.cell(row=cont - 1, column=9).border = thin_border
        ws.cell(row=cont - 1, column=10).value = u"CANTIDAD"
        ws.merge_cells(start_row=cont - 1, start_column=10, end_row=cont, end_column=10)
        ws.cell(row=cont - 1, column=10).border = thin_border
        ws.cell(row=cont - 1, column=11).value = u"COSTO UNITARIO"
        ws.merge_cells(start_row=cont - 1, start_column=11, end_row=cont, end_column=11)
        ws.cell(row=cont - 1, column=11).border = thin_border
        ws.cell(row=cont - 1, column=12).value = u"COSTO TOTAL"
        ws.merge_cells(start_row=cont - 1, start_column=12, end_row=cont, end_column=12)
        ws.cell(row=cont - 1, column=12).border = thin_border
        ws.cell(row=cont - 1, column=13).value = u"CANTIDAD"
        ws.merge_cells(start_row=cont - 1, start_column=13, end_row=cont, end_column=13)
        ws.cell(row=cont - 1, column=13).border = thin_border
        ws.cell(row=cont - 1, column=14).value = u"COSTO UNITARIO"
        ws.merge_cells(start_row=cont - 1, start_column=14, end_row=cont, end_column=14)
        ws.cell(row=cont - 1, column=14).border = thin_border
        ws.cell(row=cont - 1, column=15).value = u"COSTO TOTAL"
        ws.merge_cells(start_row=cont - 1, start_column=15, end_row=cont, end_column=15)
        ws.cell(row=cont - 1, column=15).border = thin_border
        try:
            kardex_inicial = Kardex.objects.filter(producto=producto,
                                                   almacen=almacen,
                                                   fecha_operacion__lt=desde).latest('fecha_operacion')
            cant_saldo_inicial = kardex_inicial.cantidad_total
            valor_saldo_inicial = kardex_inicial.valor_total
        except:
            cant_saldo_inicial = 0
            valor_saldo_inicial = 0
        cont = cont + 1
        ws.cell(row=cont, column=2).border = thin_border
        ws.cell(row=cont, column=3).value = '00'
        ws.cell(row=cont, column=3).border = thin_border
        ws.cell(row=cont, column=4).value = 'SALDO'
        ws.cell(row=cont, column=4).border = thin_border
        ws.cell(row=cont, column=5).value = 'INICIAL'
        ws.cell(row=cont, column=5).border = thin_border
        ws.cell(row=cont, column=6).value = '16'
        ws.cell(row=cont, column=6).border = thin_border
        ws.cell(row=cont, column=7).value = 0
        ws.cell(row=cont, column=7).number_format = '#.00000'
        ws.cell(row=cont, column=7).border = thin_border
        ws.cell(row=cont, column=8).value = 0
        ws.cell(row=cont, column=8).number_format = '#.00000'
        ws.cell(row=cont, column=8).border = thin_border
        ws.cell(row=cont, column=9).value = 0
        ws.cell(row=cont, column=9).number_format = '#.00000'
        ws.cell(row=cont, column=9).border = thin_border
        ws.cell(row=cont, column=10).value = 0
        ws.cell(row=cont, column=10).number_format = '#.00000'
        ws.cell(row=cont, column=10).border = thin_border
        ws.cell(row=cont, column=11).value = 0
        ws.cell(row=cont, column=11).number_format = '#.00000'
        ws.cell(row=cont, column=11).border = thin_border
        ws.cell(row=cont, column=12).value = 0
        ws.cell(row=cont, column=12).number_format = '#.00000'
        ws.cell(row=cont, column=12).border = thin_border
        ws.cell(row=cont, column=13).value = cant_saldo_inicial
        ws.cell(row=cont, column=13).number_format = '#.00000'
        ws.cell(row=cont, column=13).border = thin_border
        try:
            ws.cell(row=cont, column=14).value = valor_saldo_inicial / cant_saldo_inicial
        except:
            ws.cell(row=cont, column=14).value = 0
        ws.cell(row=cont, column=14).number_format = '#.00000'
        ws.cell(row=cont, column=14).border = thin_border
        ws.cell(row=cont, column=15).value = valor_saldo_inicial
        ws.cell(row=cont, column=15).number_format = '#.00000'
        ws.cell(row=cont, column=15).border = thin_border

        listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = producto.obtener_kardex(
            almacen,
            desde,
            hasta)
        for kardex in listado_kardex:
            cont = cont + 1
            ws.cell(row=cont, column=2).value = kardex.fecha_operacion.strftime('%d/%m/%Y')
            ws.cell(row=cont, column=2).border = thin_border
            try:
                ws.cell(row=cont, column=3).value = kardex.movimiento.tipo_documento.codigo_sunat
            except:
                ws.cell(row=cont, column=3).value = '-'
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=4).value = kardex.movimiento.serie
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=5).value = kardex.movimiento.numero
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=6).value = kardex.movimiento.tipo_movimiento.codigo_sunat
            ws.cell(row=cont, column=6).border = thin_border
            ws.cell(row=cont, column=7).value = kardex.cantidad_ingreso
            ws.cell(row=cont, column=7).number_format = '#.00000'
            ws.cell(row=cont, column=7).border = thin_border
            ws.cell(row=cont, column=8).value = kardex.precio_ingreso
            ws.cell(row=cont, column=8).number_format = '#.00000'
            ws.cell(row=cont, column=8).border = thin_border
            ws.cell(row=cont, column=9).value = kardex.valor_ingreso
            ws.cell(row=cont, column=9).number_format = '#.00000'
            ws.cell(row=cont, column=9).border = thin_border
            ws.cell(row=cont, column=10).value = kardex.cantidad_salida
            ws.cell(row=cont, column=10).number_format = '#.00000'
            ws.cell(row=cont, column=10).border = thin_border
            ws.cell(row=cont, column=11).value = kardex.precio_salida
            ws.cell(row=cont, column=11).number_format = '#.00000'
            ws.cell(row=cont, column=11).border = thin_border
            ws.cell(row=cont, column=12).value = kardex.valor_salida
            ws.cell(row=cont, column=12).number_format = '#.00000'
            ws.cell(row=cont, column=12).border = thin_border
            ws.cell(row=cont, column=13).value = kardex.cantidad_total
            ws.cell(row=cont, column=13).number_format = '#.00000'
            ws.cell(row=cont, column=13).border = thin_border
            ws.cell(row=cont, column=14).value = kardex.precio_total
            ws.cell(row=cont, column=14).number_format = '#.00000'
            ws.cell(row=cont, column=14).border = thin_border
            ws.cell(row=cont, column=15).value = kardex.valor_total
            ws.cell(row=cont, column=15).number_format = '#.00000'
            ws.cell(row=cont, column=15).border = thin_border
        cont = cont + 1

        ws.cell(row=cont, column=6).value = "TOTALES"
        ws.cell(row=cont, column=6).border = thin_border
        ws.cell(row=cont, column=7).value = cantidad_ingreso
        ws.cell(row=cont, column=7).number_format = '#.00000'
        ws.cell(row=cont, column=7).border = thin_border
        ws.cell(row=cont, column=8).value = ""
        ws.cell(row=cont, column=8).number_format = '#.00000'
        ws.cell(row=cont, column=8).border = thin_border
        ws.cell(row=cont, column=9).value = valor_ingreso
        ws.cell(row=cont, column=9).number_format = '#.00000'
        ws.cell(row=cont, column=9).border = thin_border
        ws.cell(row=cont, column=10).value = cantidad_salida
        ws.cell(row=cont, column=10).number_format = '#.00000'
        ws.cell(row=cont, column=10).border = thin_border
        ws.cell(row=cont, column=11).value = ""
        ws.cell(row=cont, column=11).number_format = '#.00000'
        ws.cell(row=cont, column=11).border = thin_border
        ws.cell(row=cont, column=12).value = valor_salida
        ws.cell(row=cont, column=12).number_format = '#.00000'
        ws.cell(row=cont, column=12).border = thin_border
        ws.cell(row=cont, column=13).value = ""
        ws.cell(row=cont, column=13).number_format = '#.00000'
        ws.cell(row=cont, column=13).border = thin_border
        ws.cell(row=cont, column=14).value = ""
        ws.cell(row=cont, column=14).number_format = '#.00000'
        ws.cell(row=cont, column=14).border = thin_border
        ws.cell(row=cont, column=15).value = ""
        ws.cell(row=cont, column=15).number_format = '#.00000'
        ws.cell(row=cont, column=15).border = thin_border
        return ws

    def obtener_formato_sunat_valorizado_todos(self, desde, hasta, almacen):
        productos = Producto.objects.all().order_by('descripcion')
        wb = Workbook()
        ws = wb.active
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        cont = 1
        for producto in productos:
            ws.title = producto.codigo
            self.obtener_formato_sunat_valorizado_excel_por_producto(ws,thin_border,cont,producto,desde,hasta,almacen)
            ws = wb.create_sheet("Hoja")
        return wb

    def obtener_consolidado_grupos(self, desde, hasta, almacen):
        grupos = GrupoProductos.objects.filter(estado=True,
                                               son_productos=True)
        wb = Workbook()
        ws = wb.active
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 14
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 14
        ws.column_dimensions["L"].width = 14
        ws['D1'] = u'Almacén: ' + almacen.descripcion
        ws.merge_cells('D1:F1')
        ws['H1'] = 'Periodo: ' + desde.strftime('%d/%m/%Y') + '-' + hasta.strftime('%d/%m/%Y')
        ws.merge_cells('H1:J1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'CTA_CONT.'
        ws['E3'] = 'CANT INICIAL'
        ws['F3'] = 'VALOR INICIAL'
        ws['G3'] = 'CANT. ENT'
        ws['H3'] = 'VALOR. ENT'
        ws['I3'] = 'CANT. SAL'
        ws['J3'] = 'VALOR. SAL'
        ws['K3'] = 'CANT. TOT'
        ws['L3'] = 'VALOR. TOT'
        ws['B3'].border = thin_border
        ws['C3'].border = thin_border
        ws['D3'].border = thin_border
        ws['E3'].border = thin_border
        ws['F3'].border = thin_border
        ws['G3'].border = thin_border
        ws['H3'].border = thin_border
        ws['I3'].border = thin_border
        ws['J3'].border = thin_border
        ws['K3'].border = thin_border
        ws['L3'].border = thin_border
        cont = 4
        for grupo in grupos:
            ws.cell(row=cont, column=2).value = grupo.codigo
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=3).value = grupo.descripcion
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=4).value = grupo.ctacontable.cuenta
            ws.cell(row=cont, column=4).border = thin_border
            try:
                kardex_inicial = Kardex.objects.filter(producto__grupo_productos=grupo,
                                                       almacen=almacen,
                                                       fecha_operacion__lt=desde).latest('fecha_operacion')
                cant_saldo_inicial = kardex_inicial.cantidad_total
                valor_saldo_inicial = kardex_inicial.valor_total
            except:
                cant_saldo_inicial = 0
                valor_saldo_inicial = 0
            ws.cell(row=cont, column=5).value = cant_saldo_inicial
            ws.cell(row=cont, column=5).number_format = '#.00000'
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=6).value = valor_saldo_inicial
            ws.cell(row=cont, column=6).number_format = '#.00000'
            ws.cell(row=cont, column=6).border = thin_border
            listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = grupo.obtener_kardex(
                almacen,
                desde,
                hasta)
            cantidad_total = cant_saldo_inicial + cantidad_ingreso - cantidad_salida
            valor_total = valor_saldo_inicial + valor_ingreso - valor_salida
            ws.cell(row=cont, column=7).value = cantidad_ingreso
            ws.cell(row=cont, column=7).number_format = '#.00000'
            ws.cell(row=cont, column=7).border = thin_border
            ws.cell(row=cont, column=8).value = valor_ingreso
            ws.cell(row=cont, column=8).number_format = '#.00000'
            ws.cell(row=cont, column=8).border = thin_border
            ws.cell(row=cont, column=9).value = cantidad_salida
            ws.cell(row=cont, column=9).number_format = '#.00000'
            ws.cell(row=cont, column=9).border = thin_border
            ws.cell(row=cont, column=10).value = valor_salida
            ws.cell(row=cont, column=10).number_format = '#.00000'
            ws.cell(row=cont, column=10).border = thin_border
            ws.cell(row=cont, column=11).value = cantidad_total
            ws.cell(row=cont, column=11).number_format = '#.00000'
            ws.cell(row=cont, column=11).border = thin_border
            ws.cell(row=cont, column=12).value = valor_total
            ws.cell(row=cont, column=12).number_format = '#.00000'
            ws.cell(row=cont, column=12).border = thin_border
            cont += 1
        return wb

    def obtener_consolidado_productos(self, desde, hasta, almacen):
        productos = Producto.objects.all().order_by('descripcion')
        wb = Workbook()
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ws = wb.active
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 12
        ws.column_dimensions["K"].width = 12
        ws['D1'] = u'Almacén: ' + almacen.descripcion
        ws.merge_cells('D1:F1')
        ws['G1'] = 'Periodo: ' + desde.strftime('%d/%m/%Y') + ' - ' + hasta.strftime('%d/%m/%Y')
        ws.merge_cells('G1:I1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'CANT INICIAL'
        ws['E3'] = 'VALOR INICIAL'
        ws['F3'] = 'CANT. ENT'
        ws['G3'] = 'VALOR. ENT'
        ws['H3'] = 'CANT. SAL'
        ws['I3'] = 'VALOR. SAL'
        ws['J3'] = 'CANT. TOT'
        ws['K3'] = 'VALOR. TOT'
        ws['B3'].border = thin_border
        ws['C3'].border = thin_border
        ws['D3'].border = thin_border
        ws['E3'].border = thin_border
        ws['F3'].border = thin_border
        ws['G3'].border = thin_border
        ws['H3'].border = thin_border
        ws['I3'].border = thin_border
        ws['J3'].border = thin_border
        ws['K3'].border = thin_border
        cont = 4
        for producto in productos:
            ws.cell(row=cont, column=2).value = producto.codigo
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=3).value = producto.descripcion
            ws.cell(row=cont, column=3).border = thin_border
            try:
                kardex_inicial = Kardex.objects.filter(producto=producto,
                                                       almacen=almacen,
                                                       fecha_operacion__lt=desde).latest('fecha_operacion')
                cant_saldo_inicial = kardex_inicial.cantidad_total
                valor_saldo_inicial = kardex_inicial.valor_total
            except:
                cant_saldo_inicial = 0
                valor_saldo_inicial = 0
            ws.cell(row=cont, column=4).value = cant_saldo_inicial
            ws.cell(row=cont, column=4).number_format = '#.00000'
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=5).value = valor_saldo_inicial
            ws.cell(row=cont, column=5).number_format = '#.00000'
            ws.cell(row=cont, column=5).border = thin_border
            listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida, cantidad_total, valor_total = producto.obtener_kardex(
                almacen,
                desde,
                hasta)
            cantidad_total = cant_saldo_inicial + cantidad_ingreso - cantidad_salida
            valor_total = valor_saldo_inicial + valor_ingreso - valor_salida
            ws.cell(row=cont, column=6).value = cantidad_ingreso
            ws.cell(row=cont, column=6).number_format = '#.00000'
            ws.cell(row=cont, column=6).border = thin_border
            ws.cell(row=cont, column=7).value = valor_ingreso
            ws.cell(row=cont, column=7).number_format = '#.00000'
            ws.cell(row=cont, column=7).border = thin_border
            ws.cell(row=cont, column=8).value = cantidad_salida
            ws.cell(row=cont, column=8).number_format = '#.00000'
            ws.cell(row=cont, column=8).border = thin_border
            ws.cell(row=cont, column=9).value = valor_salida
            ws.cell(row=cont, column=9).number_format = '#.00000'
            ws.cell(row=cont, column=9).border = thin_border
            ws.cell(row=cont, column=10).value = cantidad_total
            ws.cell(row=cont, column=10).number_format = '#.00000'
            ws.cell(row=cont, column=10).border = thin_border
            ws.cell(row=cont, column=11).value = valor_total
            ws.cell(row=cont, column=11).number_format = '#.00000'
            ws.cell(row=cont, column=11).border = thin_border
            cont += 1
        return wb

    def obtener_formato_normal_todos(self, desde, hasta, almacen):
        productos = Kardex.objects.filter(almacen=almacen).order_by('producto').distinct('producto__codigo')
        wb = Workbook()
        ws = wb.active
        ws['E1'] = u'Almacén: ' + almacen.descripcion
        ws.merge_cells('E1:G1')
        ws['H1'] = 'Periodo: ' + desde.strftime('%d/%m/%Y') + ' - ' + hasta.strftime('%d/%m/%Y')
        ws.merge_cells('H1:J1')
        ws['B3'] = 'FECHA'
        ws['C3'] = 'NRO_DOC'
        ws['D3'] = 'TIPO_MOV'
        ws['E3'] = 'CANT. ENT'
        ws['F3'] = 'PRE. ENT'
        ws['G3'] = 'VALOR. ENT'
        ws['H3'] = 'CANT. SAL'
        ws['I3'] = 'PRE. SAL'
        ws['J3'] = 'VALOR. SAL'
        ws['K3'] = 'CANT. TOT'
        ws['L3'] = 'PRE. TOT'
        ws['M3'] = 'VALOR. TOT'
        cont = 4
        for prod in productos:
            producto = prod.producto
            ws.cell(row=cont, column=2).value = 'Codigo: ' + producto.codigo
            ws.merge_cells(start_row=cont, start_column=2, end_row=cont, end_column=3)
            ws.cell(row=cont, column=4).value = u" Denominación: " + producto.descripcion
            ws.merge_cells(start_row=cont, start_column=4, end_row=cont, end_column=10)
            ws.cell(row=cont, column=11).value = " Unidad: " + producto.unidad_medida.descripcion
            ws.merge_cells(start_row=cont, start_column=11, end_row=cont, end_column=12)
            cont += 1
            try:
                kardex_inicial = Kardex.objects.filter(producto=producto,
                                                       almacen=almacen,
                                                       fecha_operacion__lt=desde).latest('fecha_operacion')
                cant_saldo_inicial = kardex_inicial.cantidad_total
                valor_saldo_inicial = kardex_inicial.valor_total
            except:
                cant_saldo_inicial = 0
                valor_saldo_inicial = 0
            ws.cell(row=cont, column=8).value = "SALDO INICIAL:"
            ws.merge_cells(start_row=cont, start_column=8, end_row=cont, end_column=9)
            ws.cell(row=cont, column=10).value = "Cantidad: "
            ws.cell(row=cont, column=11).value = cant_saldo_inicial
            ws.cell(row=cont, column=11).number_format = '#.00000'
            ws.cell(row=cont, column=12).value = "Valor: "
            ws.cell(row=cont, column=13).value = valor_saldo_inicial
            ws.cell(row=cont, column=13).number_format = '#.00000'
            cont += 1
            listado_kardex, cantidad_ingreso, valor_ingreso, cantidad_salida, valor_salida = producto.obtener_kardex(
                almacen,
                desde,
                hasta)
            if len(listado_kardex) > 0:
                for kardex in listado_kardex:
                    ws.cell(row=cont, column=2).value = kardex.fecha_operacion
                    ws.cell(row=cont, column=2).number_format = 'dd/mm/yyyy'
                    ws.cell(row=cont, column=3).value = kardex.movimiento.id_movimiento
                    ws.cell(row=cont, column=4).value = kardex.movimiento.tipo_movimiento.codigo
                    ws.cell(row=cont, column=5).value = kardex.cantidad_ingreso
                    ws.cell(row=cont, column=6).value = kardex.precio_ingreso
                    ws.cell(row=cont, column=6).number_format = '#.00000'
                    ws.cell(row=cont, column=7).value = kardex.valor_ingreso
                    ws.cell(row=cont, column=7).number_format = '#.00000'
                    ws.cell(row=cont, column=8).value = kardex.cantidad_salida
                    ws.cell(row=cont, column=9).value = kardex.precio_salida
                    ws.cell(row=cont, column=9).number_format = '#.00000'
                    ws.cell(row=cont, column=10).value = kardex.valor_salida
                    ws.cell(row=cont, column=10).number_format = '#.00000'
                    ws.cell(row=cont, column=11).value = kardex.cantidad_total
                    ws.cell(row=cont, column=12).value = kardex.precio_total
                    ws.cell(row=cont, column=12).number_format = '#.00000'
                    ws.cell(row=cont, column=13).value = kardex.valor_total
                    ws.cell(row=cont, column=13).number_format = '#.00000'
                    cont += 1
                ws.cell(row=cont, column=5).value = cantidad_ingreso
                ws.cell(row=cont, column=7).value = valor_ingreso
                ws.cell(row=cont, column=7).number_format = '#.00000'
                ws.cell(row=cont, column=8).value = cantidad_salida
                ws.cell(row=cont, column=10).value = valor_salida
                ws.cell(row=cont, column=10).number_format = '#.00000'
                ws.cell(row=cont, column=11).value = kardex.cantidad_total
                ws.cell(row=cont, column=13).value = kardex.valor_total
                ws.cell(row=cont, column=13).number_format = '#.00000'
                cont += 2
            else:
                ws.cell(row=cont, column=5).value = 0
                ws.cell(row=cont, column=6).value = 0
                ws.cell(row=cont, column=6).number_format = '#.00000'
                ws.cell(row=cont, column=7).value = 0
                ws.cell(row=cont, column=7).number_format = '#.00000'
                ws.cell(row=cont, column=8).value = 0
                ws.cell(row=cont, column=9).value = 0
                ws.cell(row=cont, column=9).number_format = '#.00000'
                ws.cell(row=cont, column=10).value = 0
                ws.cell(row=cont, column=10).number_format = '#.00000'
                ws.cell(row=cont, column=11).value = 0
                ws.cell(row=cont, column=12).value = 0
                ws.cell(row=cont, column=12).number_format = '#.00000'
                ws.cell(row=cont, column=13).value = 0
                ws.cell(row=cont, column=13).number_format = '#.00000'
                cont += 1
                ws.cell(row=cont, column=5).value = 0
                ws.cell(row=cont, column=7).value = 0
                ws.cell(row=cont, column=7).number_format = '#.00000'
                ws.cell(row=cont, column=8).value = 0
                ws.cell(row=cont, column=10).value = 0
                ws.cell(row=cont, column=10).number_format = '#.00000'
                ws.cell(row=cont, column=11).value = cant_saldo_inicial
                ws.cell(row=cont, column=13).value = valor_saldo_inicial
                ws.cell(row=cont, column=13).number_format = '#.00000'
                cont += 2
        return wb
