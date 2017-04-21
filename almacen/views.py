# -*- coding: utf-8 -*- 
from django.shortcuts import render
from openpyxl.styles import Border
from openpyxl.styles import Side

from almacen.models import Almacen, Movimiento, Kardex,TipoMovimiento,  DetalleMovimiento, ControlProductoAlmacen,\
    Pedido, DetallePedido
from django.http import HttpResponse, HttpResponseRedirect
from django.core.urlresolvers import reverse, reverse_lazy
import datetime
from django.views.generic import TemplateView, FormView, View, ListView
from almacen.forms import AlmacenForm, TipoStockForm, TipoSalidaForm, TipoMovimientoForm, FormularioReporteMovimientos,\
    FormularioKardexProducto, CargarInventarioInicialForm, FormularioReporteStock, MovimientoForm,\
    DetalleIngresoFormSet, DetalleSalidaFormSet, PedidoForm, DetallePedidoFormSet,\
    AprobacionPedidoForm, FormularioReprocesoPrecio,\
    FormularioMovimientosProducto, FormularioConsultaStock
from django.db.models import Sum
from decimal import Decimal
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table
from django.http import JsonResponse
from compras.models import DetalleOrdenCompra
from openpyxl import Workbook
import simplejson
import json
from django.conf import settings
from django.views.generic.detail import DetailView
from django.views.generic.edit import UpdateView, CreateView
from administracion.models import Puesto
import locale
from contabilidad.models import Tipo, TipoDocumento
import csv
from contabilidad.forms import UploadForm
import os
from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
from django.db.models import Q
from django.db import transaction, IntegrityError
from django.contrib import messages
from productos.models import Producto, GrupoProductos
from almacen.mail import correo_creacion_pedido
from almacen.reports import ReporteMovimiento, ReporteKardexPDF, ReporteKardexExcel
from almacen.settings import EMPRESA, LOGISTICA
from datetime import date

locale.setlocale(locale.LC_ALL,"")

class Tablero(View):
    
    def get(self, request, *args, **kwargs):
        cod_mov_invent_ini = 'I00'
        cod_mov_ingreso_compra = 'I01'
        cod_mov_salida_pedido = 'S01'
        lista_notificaciones = []
        cant_almacenes = Almacen.objects.count()        
        cant_tipos_movimientos_ingreso = TipoMovimiento.objects.filter(incrementa=True).exclude(codigo=cod_mov_invent_ini).count()
        cant_tipos_movimientos_salida = TipoMovimiento.objects.filter(incrementa=False).count()        
        tipo_movimiento, creado = TipoMovimiento.objects.get_or_create(codigo = cod_mov_invent_ini,
                                                                       defaults = {'descripcion':'INVENTARIO INICIAL',
                                                                                   'codigo_sunat': '16',
                                                                                   'incrementa':True,
                                                                                   'estado':True})
        if creado:
            lista_notificaciones.append("Se ha creado el tipo de movimiento inventario inicial")
        tipo_movimiento, creado = TipoMovimiento.objects.get_or_create(codigo = cod_mov_ingreso_compra,
                                                                       defaults = {'descripcion':'INGRESO POR COMPRA',
                                                                                   'codigo_sunat': '02',
                                                                                   'incrementa':True,
                                                                                   'pide_referencia':True,
                                                                                   'estado':True})
        if creado:
            lista_notificaciones.append("Se ha creado el tipo de movimiento Ingreso por Compra")
        tipo_movimiento, creado = TipoMovimiento.objects.get_or_create(codigo = cod_mov_salida_pedido,
                                                                       defaults = {'descripcion':'SALIDA POR PEDIDO',
                                                                                   'codigo_sunat': '10',
                                                                                   'incrementa':False,
                                                                                   'pide_referencia':True,
                                                                                   'estado':True})
        inventario_inicial = Movimiento.objects.filter(tipo_movimiento__codigo=cod_mov_invent_ini).count()
        if creado:
            lista_notificaciones.append("Se ha creado el tipo de movimiento Salida por Pedido")                            
        if cant_almacenes==0:
            lista_notificaciones.append("No se ha creado ningún almacen")            
        if cant_tipos_movimientos_ingreso == 0:
            lista_notificaciones.append("No se ha creado ningún tipo de movimiento de ingreso")            
        if cant_tipos_movimientos_salida == 0:
            lista_notificaciones.append("No se ha creado ningún tipo de movimiento de salida")        
        if inventario_inicial == 0:
            lista_notificaciones.append("No se ha realizado el inventario inicial")            
        context = {'notificaciones':lista_notificaciones}
        return render(request, 'almacen/tablero_almacen.html', context)
    
class AprobarPedido(CreateView):
    form_class = AprobacionPedidoForm
    template_name = 'almacen/aprobar_pedido.html'    
    model = Movimiento
    
    @method_decorator(permission_required('almacen.aprobar_pedido',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        self.codigo = kwargs['codigo']
        return super(AprobarPedido, self).dispatch(*args, **kwargs)
        
    def get_form_kwargs(self):
        kwargs = super(AprobarPedido, self).get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_initial(self):
        initial = super(AprobarPedido, self).get_initial()
        initial['cod_pedido'] = self.codigo
        return initial
    
    def get_context_data(self, **kwargs):
        pedido = Pedido.objects.get(codigo = self.codigo)
        context = super(AprobarPedido, self).get_context_data(**kwargs)
        context['pedido'] = pedido
        return context
    
    def get(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)        
        pedido = Pedido.objects.get(codigo = self.codigo)
        try:
            trabajador = self.request.user.trabajador
        except:
            return HttpResponseRedirect(reverse('administracion:crear_trabajador'))        
        try:
            puestos = trabajador.puesto_set.all().filter(estado=True)
            if trabajador.firma == '':
                return HttpResponseRedirect(reverse('administracion:modificar_trabajador'))
            if puestos[0].es_jefatura and puestos[0].oficina == LOGISTICA:
                form_class = self.get_form_class()
                form = self.get_form(form_class)                
                detalles = DetallePedido.objects.filter(pedido=pedido,estado=DetallePedido.STATUS.PEND)
                detalles_data = []
                for detalle in detalles:                    
                    d = {'pedido': detalle.id,
                         'codigo': detalle.producto.codigo,
                         'nombre': detalle.producto.descripcion,
                         'unidad': detalle.producto.unidad_medida.codigo,
                         'cantidad': detalle.cantidad
                        }
                    detalles_data.append(d)
                detalle_salida_formset = DetalleSalidaFormSet(initial=detalles_data)
                return self.render_to_response(self.get_context_data(form=form,
                                                                     detalle_salida_formset=detalle_salida_formset))
            else:
                return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))
        except:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))        
    
    def post(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_salida_formset = DetalleSalidaFormSet(request.POST)
        if form.is_valid() and detalle_salida_formset.is_valid():
            return self.form_valid(form, detalle_salida_formset)
        else:
            return self.form_invalid(form, detalle_salida_formset)
        
    def form_valid(self, form, detalle_salida_formset):
        try:
            with transaction.atomic():
                self.object = form.save()
                pedido = self.object.pedido
                detalles = []
                cont = 1                
                for detalle_salida_form in detalle_salida_formset:
                    detalle_pedido = detalle_salida_form.cleaned_data.get('pedido')
                    codigo = detalle_salida_form.cleaned_data.get('codigo') 
                    cantidad = detalle_salida_form.cleaned_data.get('cantidad')
                    precio = detalle_salida_form.cleaned_data.get('precio')
                    valor = detalle_salida_form.cleaned_data.get('valor')
                    if cantidad and precio and valor:
                        detalle_movimiento = DetalleMovimiento(nro_detalle = cont,
                                                               movimiento = self.object,
                                                               producto = Producto.objects.get(pk=codigo),
                                                               detalle_pedido = DetallePedido.objects.get(pk=detalle_pedido),
                                                               cantidad = cantidad,
                                                               precio = precio)
                        detalles.append(detalle_movimiento)                        
                        cont = cont + 1
                DetalleMovimiento.objects.bulk_create(detalles, None, pedido) 
                return HttpResponseRedirect(reverse('almacen:detalle_movimiento', args=[self.object.id_movimiento]))
        except IntegrityError:
            messages.error(self.request, 'Error guardando la cotizacion.')
        
    def form_invalid(self, form, detalle_salida_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_salida_formset = detalle_salida_formset))
        
class BusquedaProductosAlmacen(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            lista_productos = []
            descripcion = request.GET['descripcion']
            almacen = request.GET['almacen']
            kardex_ant = Kardex.objects.filter(producto__descripcion__icontains = descripcion,
                                               almacen__id=almacen).order_by('producto').distinct('producto__codigo')[:20]
            for kardex in kardex_ant:
                control = Kardex.objects.filter(producto = kardex.producto,
                                                almacen__id=almacen).latest('fecha_operacion')
                producto_json = {}
                producto_json['label'] = control.producto.descripcion
                producto_json['codigo'] = control.producto.codigo
                producto_json['descripcion'] = control.producto.descripcion
                producto_json['unidad'] = control.producto.unidad_medida.descripcion
                try:
                    precio=round(control.valor_total / control.cantidad_total, 5)
                except:
                    precio = 0
                producto_json['precio'] = str(precio)
                lista_productos.append(producto_json)
            data = json.dumps(lista_productos)
            return HttpResponse(data, 'application/json')

class CargarAlmacenes(FormView):
    template_name = 'almacen/cargar_almacenes.html'
    form_class = UploadForm
    
    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']            
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT,'archivos',str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        for fila in dataReader:
            Almacen.objects.create(codigo=fila[0],
                                   descripcion=unicode(fila[1], errors='ignore'))                        
        return HttpResponseRedirect(reverse('almacen:almacenes'))
    
class CargarInventarioInicial(FormView):
    template_name = 'almacen/cargar_inventario_inicial.html'
    form_class = CargarInventarioInicialForm
    
    def obtener_fecha_hora(self,r_fecha,r_hora):
        r_hora = r_hora.replace(" ","")
        anio = int(r_fecha[6:])
        mes = int(r_fecha[3:5])
        dia = int(r_fecha[0:2])
        horas = int(r_hora[0:2])
        minutos = int(r_hora[3:5])
        segundos = int(r_hora[6:8])
        fecha = datetime.datetime(anio,mes,dia,horas,minutos,segundos)
        return fecha
    
    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']
        fecha = data['fecha']
        hora = data['hora']
        almacen = data['almacenes']
        form.save()
        fecha_operacion = self.obtener_fecha_hora(fecha, hora)
        usuario = self.request.user
        csv_filepathname = os.path.join(settings.MEDIA_ROOT,'archivos',str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        tipo_movimiento = TipoMovimiento.objects.get(codigo='I00')
        with transaction.atomic():
            tipo_documento = TipoDocumento.objects.get(codigo_sunat='PEC')            
            movimiento = Movimiento.objects.create(tipo_movimiento = tipo_movimiento,
                                                   tipo_documento = tipo_documento,
                                                   almacen = almacen,
                                                   fecha_operacion=fecha_operacion,
                                                   observaciones = 'INVENTARIO INICIAL',
                                                   serie = 'SALDO',
                                                   numero = 'INICIAL')            
            cont_detalles = 1
            detalles = []
            total = 0
            for fila in dataReader:
                try:
                    producto = Producto.objects.get(descripcion=unicode(fila[0].strip(), errors='ignore'))
                    cantidad = Decimal(fila[1])
                    try:
                        precio = Decimal(fila[2])
                    except:
                        precio = ''
                    valor = Decimal(fila[3])
                    if precio == '':
                        try:
                            precio = valor / cantidad
                        except:
                            precio = 0
                    if valor == '':
                        valor = cantidad * precio
                    detalle_movimiento = DetalleMovimiento(nro_detalle=cont_detalles,
                                                           movimiento=movimiento,
                                                           producto=producto,
                                                           cantidad=cantidad,
                                                           precio=precio,
                                                           valor=valor)
                    detalles.append(detalle_movimiento)
                except Producto.DoesNotExist:
                    pass                 
                total = total + valor
                cont_detalles = cont_detalles + 1
            DetalleMovimiento.objects.bulk_create(detalles,None, None)
            movimiento.save()
        return HttpResponseRedirect(reverse('almacen:detalle_movimiento', args=[movimiento.id_movimiento]))
    
class CrearTipoMovimiento(CreateView):
    template_name = 'almacen/tipo_movimiento.html'
    form_class = TipoMovimientoForm
    success_url = reverse_lazy('almacen:tipos_movimientos')

    @method_decorator(permission_required('almacen.add_tipomovimiento',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearTipoMovimiento, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('almacen:detalle_tipo_movimiento', args=[self.object.pk])
    
class CrearTipoSalida(FormView):
    template_name = 'almacen/crear_tipo_salida.html'
    form_class = TipoSalidaForm
    success_url = reverse_lazy('almacen:crear_tipo_salida')

    def form_valid(self, form):
        form.save()
        return super(CrearTipoSalida, self).form_valid(form)  
    
class CrearTipoStock(View):
    
    def get(self, request, *args, **kwargs):
        form = TipoStockForm()        
        return render(request, 'crear_tipo_stock.html', {'form': form})
    
    '''template_name = 'almacen/crear_tipo_stock.html'
    form_class = TipoStockForm
    success_url = reverse_lazy('almacen:crear_tipo_stock')

    def form_valid(self, form):
        form.save()
        return super(CrearTipoStock, self).form_valid(form)'''

class CrearAlmacen(FormView):
    template_name = 'almacen/almacen.html'
    form_class = AlmacenForm
    success_url = reverse_lazy('almacen:almacenes')

    def form_valid(self, form):
        form.save()
        return super(CrearAlmacen, self).form_valid(form)
    
'''class CrearDetalleSalida(FormView):
    template_name = 'almacen/crear_detalle_salida.html'
    form_class = FormularioDetalleMovimiento
    success_url = reverse_lazy('almacen:crear_detalle_salida')
    
    def get(self, request, *args, **kwargs):
        self.almacen = kwargs['almacen']
        return super(CrearDetalleSalida, self).get(request, *args, **kwargs)
    
    def get_initial(self):
        initial = super(CrearDetalleSalida, self).get_initial()        
        initial['almacen'] = self.almacen       
        return initial

    def form_valid(self, form):
        form.save()
        return super(CrearDetalleSalida, self).form_valid(form)'''
    
class CrearDetalleSalida(TemplateView):
        
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            lista_detalles = []            
            det = {}
            det['codigo'] = ''               
            det['nombre'] = ''                    
            det['cantidad'] = '0'
            det['precio'] = '0'
            det['unidad'] = ''
            det['valor'] = '0'
            lista_detalles.append(det)
            formset = DetalleSalidaFormSet(initial=lista_detalles)
            lista_json = []
            for form in formset:
                detalle_json = {}    
                detalle_json['codigo'] = str(form['codigo'])
                detalle_json['nombre'] = str(form['nombre'])
                detalle_json['cantidad'] = str(form['cantidad'])
                detalle_json['precio'] = str(form['precio'])
                detalle_json['unidad'] = str(form['unidad'])                
                detalle_json['valor'] = str(form['valor'])
                lista_json.append(detalle_json)                                
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')
        
class CrearDetallePedido(TemplateView):
        
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            lista_detalles = []            
            det = {}
            det['codigo'] = ''               
            det['nombre'] = ''                    
            det['cantidad'] = '0'
            det['unidad'] = ''            
            lista_detalles.append(det)
            formset = DetallePedidoFormSet(initial=lista_detalles)
            lista_json = []
            for form in formset:
                detalle_json = {}    
                detalle_json['codigo'] = str(form['codigo'])
                detalle_json['nombre'] = str(form['nombre'])
                detalle_json['cantidad'] = str(form['cantidad'])
                detalle_json['unidad'] = str(form['unidad'])
                lista_json.append(detalle_json)
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')

class CrearDetalleIngreso(TemplateView):
        
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            lista_detalles = []            
            det = {}
            det['orden_compra'] = '0'
            det['codigo'] = ''               
            det['nombre'] = ''                    
            det['cantidad'] = '0'
            det['precio'] = '0'
            det['unidad'] = ''
            det['valor'] = '0'
            lista_detalles.append(det)
            formset = DetalleIngresoFormSet(initial=lista_detalles)
            lista_json = []
            for form in formset:
                detalle_json = {}    
                detalle_json['orden_compra'] = str(form['orden_compra'])
                detalle_json['codigo'] = str(form['codigo'])
                detalle_json['nombre'] = str(form['nombre'])
                detalle_json['cantidad'] = str(form['cantidad'])
                detalle_json['precio'] = str(form['precio'])
                detalle_json['unidad'] = str(form['unidad'])                
                detalle_json['valor'] = str(form['valor'])
                lista_json.append(detalle_json)                                
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')
    
class CrearPedido(CreateView):
    template_name = 'almacen/pedido.html'
    form_class = PedidoForm
    model = Pedido
    
    @method_decorator(permission_required('almacen.add_pedido',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        try:
            trabajador = self.request.user.trabajador
        except:
            return HttpResponseRedirect(reverse('administracion:crear_trabajador'))
        if trabajador.firma == '':
            return HttpResponseRedirect(reverse('administracion:modificar_trabajador', args=[trabajador.pk]))
        puesto = trabajador.puesto
        if puesto is None:
            return HttpResponseRedirect(reverse('administracion:crear_puesto'))
        if puesto.es_jefatura or puesto.es_asistente:
            return super(CrearPedido, self).dispatch(*args, **kwargs)                
        else:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))        
    
    def get(self, request, *args, **kwargs):
        self.object = None        
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_pedido_formset=DetallePedidoFormSet()
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_pedido_formset = detalle_pedido_formset))
        
    def get_form_kwargs(self):
        kwargs = super(CrearPedido, self).get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def post(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_pedido_formset = DetallePedidoFormSet(request.POST)
        if form.is_valid() and detalle_pedido_formset.is_valid():
            return self.form_valid(form, detalle_pedido_formset)
        else:
            return self.form_invalid(form, detalle_pedido_formset)
    
    def form_valid(self, form, detalle_pedido_formset):
        try:
            with transaction.atomic():
                self.object = form.save()        
                detalles = []
                cont = 1
                for detalle_pedido_form in detalle_pedido_formset:                
                    codigo = detalle_pedido_form.cleaned_data.get('codigo')
                    cantidad = detalle_pedido_form.cleaned_data.get('cantidad')
                    if codigo and cantidad:                                
                        producto = Producto.objects.get(codigo=codigo)                
                        detalles.append(DetallePedido(pedido=self.object, 
                                                      nro_detalle=cont, 
                                                      producto=producto, 
                                                      cantidad=cantidad))
                        cont = cont + 1                    
                DetallePedido.objects.bulk_create(detalles)
                puesto_jefe_logistica = Puesto.objects.get(oficina = LOGISTICA, es_jefatura = True, estado = True)
                jefe_logistica = puesto_jefe_logistica.trabajador
                destinatario = jefe_logistica.usuario.email
                correo_creacion_pedido(destinatario,self.object)
                return HttpResponseRedirect(reverse('almacen:detalle_pedido', args=[self.object.pk]))
        except IntegrityError:
                messages.error(self.request, 'Error guardando el pedido.')
        
    def form_invalid(self, form, detalle_pedido_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_pedido_formset=detalle_pedido_formset))
    
class ConsultaStock(TemplateView):
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            almacen = request.GET['almacen']
            codigo = request.GET['codigo']            
            control_producto = Kardex.objects.filter(producto__codigo = codigo,
                                                     almacen__id=almacen).latest('fecha_operacion')
            producto_json = {}                
            producto_json['stock'] = control_producto.cantidad_total
            data = simplejson.dumps(producto_json)
            return HttpResponse(data, 'application/json')

class DetalleAlmacen(DetailView):
    model = Almacen
    template_name = 'almacen/detalle_almacen.html'
    
class DetalleTipoMovimiento(DetailView):
    model = TipoMovimiento
    template_name = 'almacen/detalle_tipo_movimiento.html'

class DetalleOperacionPedido(DetailView):
    model = Pedido
    template_name = 'almacen/detalle_pedido.html'    
    
class DetalleOperacionMovimiento(DetailView):
    model = Movimiento
    template_name = 'almacen/detalle_movimiento.html'

class EliminarAlmacen(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            almacen = Almacen.objects.get(pk=codigo)
            almacen_json = {}
            almacen_json['codigo'] = almacen.codigo
            almacen_json['descripcion'] = almacen.descripcion
            if len(almacen.movimiento_set.all())>0:
                almacen_json['relaciones'] = 'SI'
            else:
                almacen_json['relaciones'] = 'NO'
                Almacen.objects.filter(pk=codigo).update(estado = False)                
            data = simplejson.dumps(almacen_json)
            return HttpResponse(data, 'application/json')
    
class EliminarMovimiento(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            id_movimiento = request.GET['id_movimiento']
            movimiento = Movimiento.objects.get(pk = id_movimiento)
            orden = movimiento.referencia
            pedido = movimiento.pedido
            if orden is not None:
                movimiento.eliminar_referencia()                
            if pedido is not None:
                movimiento.eliminar_pedido()
            detalle_kardex = Kardex.objects.filter(movimiento = movimiento)
            for kardex in detalle_kardex:
                control = ControlProductoAlmacen.objects.get(producto = kardex.producto, almacen = kardex.almacen)
                if kardex.cantidad_ingreso>0:
                    control.stock = control.stock - kardex.cantidad_ingreso                                        
                elif kardex.cantidad_salida > 0:
                    control.stock = control.stock + kardex.cantidad_salida
                control.save()
                kardex.delete()                
            Movimiento.objects.filter(pk=id_movimiento).update(estado = Movimiento.STATUS.CANC, referencia=None)
            DetalleMovimiento.objects.filter(movimiento=movimiento).delete()
            movimiento_json = {}
            movimiento_json['id_movimiento'] = id_movimiento
            data = simplejson.dumps(movimiento_json)
            return HttpResponse(data, 'application/json')
        
class EliminarPedido(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            pedido = Pedido.objects.get(pk=codigo)            
            movimientos = pedido.movimiento_set.all()
            almacen_json = {}
            almacen_json['codigo'] = pedido.codigo
            if len(movimientos) > 0:
                almacen_json['movimientos'] = 'SI'
            else:
                almacen_json['movimientos'] = 'NO'
                with transaction.atomic():
                    Pedido.objects.filter(codigo=codigo).update(estado = Pedido.STATUS.CANC)
                    DetallePedido.objects.filter(pedido = pedido).delete()
            data = simplejson.dumps(almacen_json)
            return HttpResponse(data, 'application/json')

class InicioOperaciones(TemplateView):
    template_name = "inicio_operaciones.html"    
    
class ListadoAprobacionPedidos(ListView):
    model = Pedido
    template_name = 'almacen/listado_pedidos.html'
    context_object_name = 'pedidos'    
    
    @method_decorator(permission_required('almacen.ver_tabla_aprobacion_pedidos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        try:
            trabajador = self.request.user.trabajador
        except:
            return HttpResponseRedirect(reverse('administracion:crear_trabajador'))        
        try:
            puestos = trabajador.puesto_set.all().filter(estado=True)
            if trabajador.firma == '':
                return HttpResponseRedirect(reverse('administracion:modificar_trabajador'))
            if puestos[0].es_jefatura and puestos[0].oficina == LOGISTICA:
                return super(ListadoAprobacionPedidos, self).dispatch(*args, **kwargs)
            else:
                return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))
        except:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))
    
    def get_queryset(self):
        queryset = Pedido.objects.filter(~Q(estado = Pedido.STATUS.APROB))
        return queryset
    
class ListadoAlmacenes(ListView):
    model = Almacen
    template_name = 'almacen/almacenes.html'
    context_object_name = 'almacenes'
    queryset = Almacen.objects.all().order_by('descripcion')

class ListadoPedidos(ListView):
    model = Pedido
    template_name = 'almacen/listado_pedidos.html'
    context_object_name = 'pedidos'
    queryset = Pedido.objects.exclude(estado=Pedido.STATUS.CANC).order_by('codigo')
    
class ListadoTiposUnidadMedida(ListView):
    model = Tipo
    template_name = 'almacen/tipos.html'
    context_object_name = 'tipos'
    paginate_by = 10
    queryset = Tipo.objects.filter(tabla="tipo_unidad_medida",descripcion_campo="tipo_unidad_medida").order_by('descripcion_valor')

    def get_context_data(self, **kwargs):
        context = super(ListadoTiposUnidadMedida, self).get_context_data(**kwargs)
        context['tabla'] = 'tipo_unidad_medida'
        return context
    
class ListadoTiposStock(ListView):
    model = Tipo
    template_name = 'almacen/tipos.html'
    context_object_name = 'tipos'
    paginate_by = 10
    queryset = Tipo.objects.filter(tabla="tipo_stock",descripcion_campo="tipo_stock").order_by('descripcion_valor')

    def get_context_data(self, **kwargs):
        context = super(ListadoTiposStock, self).get_context_data(**kwargs)
        context['tabla'] = 'tipo_stock'
        return context

class ListadoTiposMovimiento(ListView):
    model = TipoMovimiento
    template_name = 'almacen/tipos_movimiento.html'
    context_object_name = 'tipos_movimiento'
    paginate_by = 10
    queryset = TipoMovimiento.objects.all().order_by('codigo')
    
class ListadoMovimientos(ListView):
    model = Movimiento
    template_name = 'almacen/movimientos.html'
    context_object_name = 'movimientos'
    queryset = Movimiento.objects.filter(estado=Movimiento.STATUS.ACT)    

class ListadoIngresos(ListView):
    model = Movimiento
    template_name = 'almacen/listado_ingresos.html'
    context_object_name = 'movimientos'
    queryset = Movimiento.objects.filter(estado=Movimiento.STATUS.ACT, tipo_movimiento__incrementa=True) 
    
class ListadoSalidas(ListView):
    model = Movimiento
    template_name = 'almacen/listado_salidas.html'
    context_object_name = 'movimientos'
    queryset = Movimiento.objects.filter(estado=Movimiento.STATUS.ACT, tipo_movimiento__incrementa=False) 
    
class ListadoMovimientosPorPedido(ListView):
    model = Movimiento
    template_name = 'almacen/movimientos.html'
    context_object_name = 'movimientos'    
    
    @method_decorator(permission_required('almacen.ver_tabla_movimientos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoMovimientosPorPedido, self).dispatch(*args, **kwargs)
    
    def get_queryset(self):
        pedido = Pedido.objects.get(pk=self.kwargs['pedido'])
        queryset = pedido.movimiento_set.all()
        return queryset

class ModificarMovimiento(TemplateView):
    
    @method_decorator(permission_required('almacen.change_movimiento',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarMovimiento, self).dispatch(*args, **kwargs)        
    
    def get(self, request, *args, **kwargs):  
        pk = kwargs['pk']
        movimiento = Movimiento.objects.get(pk=pk)
        if movimiento.estado == Movimiento.STATUS.CANC:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))    
        tipo_movimiento = movimiento.tipo_movimiento
        if tipo_movimiento.incrementa:
            return HttpResponseRedirect(reverse('almacen:modificar_ingreso_almacen', args=[movimiento.pk]))
        else:
            return HttpResponseRedirect(reverse('almacen:modificar_salida_almacen', args=[movimiento.pk]))
    
class ModificarIngresoAlmacen(UpdateView):
    template_name = 'almacen/ingreso_almacen.html'
    form_class = MovimientoForm
    model = Movimiento
    
    def get_form_kwargs(self):
        kwargs = super(ModificarIngresoAlmacen, self).get_form_kwargs()
        kwargs['tipo_movimiento'] = 'I'
        return kwargs
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.estado == Movimiento.STATUS.ACT:            
            form_class = self.get_form_class()
            form = self.get_form(form_class)
            detalles = DetalleMovimiento.objects.filter(movimiento=self.object).order_by('nro_detalle')
            detalles_data = []
            for detalle in detalles:
                if detalle.detalle_orden_compra is not None:
                    if detalle.detalle_orden_compra.detalle_cotizacion is not None:
                        d = {'orden_compra': detalle.detalle_orden_compra.pk,
                             'codigo': detalle.detalle_orden_compra.detalle_cotizacion.detalle_requerimiento.producto.codigo,
                             'nombre': detalle.detalle_orden_compra.detalle_cotizacion.detalle_requerimiento.producto.descripcion,
                             'unidad': detalle.detalle_orden_compra.detalle_cotizacion.detalle_requerimiento.producto.unidad_medida.codigo,
                             'cantidad': detalle.cantidad,
                             'precio': detalle.precio,
                             'valor': detalle.valor }
                    else:
                        d = {'orden_compra': detalle.detalle_orden_compra.pk,
                             'codigo': detalle.detalle_orden_compra.producto.codigo,
                             'nombre': detalle.detalle_orden_compra.producto.descripcion,
                             'unidad': detalle.detalle_orden_compra.producto.unidad_medida.codigo,
                             'cantidad': detalle.cantidad,
                             'precio': detalle.precio,
                             'valor': detalle.valor }
                else:
                    d = {'orden_compra': '0',
                         'codigo': detalle.producto.codigo,
                         'nombre': detalle.producto.descripcion,
                         'unidad': detalle.producto.unidad_medida.codigo,
                         'cantidad': detalle.cantidad,
                         'precio': detalle.precio,
                         'valor': detalle.valor }
                detalles_data.append(d)
            detalle_ingreso_formset = DetalleIngresoFormSet(initial=detalles_data)
            return self.render_to_response(self.get_context_data(form=form,
                                                                 detalle_ingreso_formset=detalle_ingreso_formset))            
        else:
            return HttpResponseRedirect(reverse('almacen:listado_ingresos'))
    
    def get_initial(self):
        initial = super(ModificarIngresoAlmacen, self).get_initial()
        movimiento = self.object
        initial['id_movimiento'] = movimiento.id_movimiento
        initial['fecha'] = movimiento.fecha_operacion.strftime('%d/%m/%Y')
        initial['hora'] = movimiento.fecha_operacion.strftime('%H : %M : %S')
        initial['almacen'] = movimiento.almacen
        initial['tipo_movimiento'] = movimiento.tipo_movimiento
        initial['doc_referencia'] = movimiento.referencia     
        initial['tipo_documento'] = movimiento.tipo_documento
        initial['serie'] = movimiento.serie
        initial['numero'] = movimiento.numero
        initial['total'] = movimiento.total
        initial['observaciones'] = movimiento.observaciones
        return initial 
        
    def get_context_data(self, **kwargs):
        movimiento = self.object        
        context = super(ModificarIngresoAlmacen, self).get_context_data(**kwargs)
        context['movimiento'] = movimiento
        return context
    
    def post(self, request, *args, **kwargs):                      
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_ingreso_formset = DetalleIngresoFormSet(request.POST)
        if form.is_valid() and detalle_ingreso_formset.is_valid():
            return self.form_valid(form, detalle_ingreso_formset)
        else:
            return self.form_invalid(form, detalle_ingreso_formset)
    
    def form_valid(self, form, detalle_ingreso_formset):
        try:
            with transaction.atomic():
                if self.object.referencia:
                    self.object.eliminar_referencia()
                self.object.eliminar_kardex()
                self.object.eliminar_detalles()
                self.object = form.save()
                referencia = self.object.referencia
                detalles = []
                cont = 1                
                for detalle_ingreso_form in detalle_ingreso_formset:
                    orden_compra = detalle_ingreso_form.cleaned_data.get('orden_compra')
                    codigo = detalle_ingreso_form.cleaned_data.get('codigo') 
                    cantidad = detalle_ingreso_form.cleaned_data.get('cantidad')
                    precio = detalle_ingreso_form.cleaned_data.get('precio')
                    valor = detalle_ingreso_form.cleaned_data.get('valor')                    
                    if cantidad and precio and valor:
                        try:
                            detalle_orden_compra = DetalleOrdenCompra.objects.get(pk=orden_compra)
                            detalle_movimiento = DetalleMovimiento(detalle_orden_compra=detalle_orden_compra,
                                                                   nro_detalle=cont,
                                                                   movimiento=self.object,
                                                                   producto=Producto.objects.get(pk=codigo),
                                                                   cantidad=cantidad,
                                                                   precio=precio,
                                                                   valor=valor)
                        except:
                            detalle_movimiento = DetalleMovimiento(nro_detalle=cont,
                                                                   movimiento=self.object,
                                                                   producto=Producto.objects.get(pk=codigo),
                                                                   cantidad=cantidad,
                                                                   precio=precio,
                                                                   valor=valor)
                        detalles.append(detalle_movimiento)                        
                        cont = cont + 1
                DetalleMovimiento.objects.bulk_create(detalles, referencia, None) 
                return HttpResponseRedirect(reverse('almacen:detalle_movimiento', args=[self.object.pk]))
        except IntegrityError:
            messages.error(self.request, 'Error guardando la cotizacion.')
    
    def form_invalid(self, form, detalle_ingreso_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_ingreso_formset=detalle_ingreso_formset))    
    
class ModificarSalidaAlmacen(UpdateView):
    template_name = 'almacen/salida_almacen.html'
    form_class = MovimientoForm
    model = Movimiento
    
    def get_form_kwargs(self):
        kwargs = super(ModificarSalidaAlmacen, self).get_form_kwargs()
        kwargs['tipo_movimiento'] = 'S'
        return kwargs
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalles = DetalleMovimiento.objects.filter(movimiento = self.object)
        detalles_data = []
        for detalle in detalles:
            try:
                d = {'pedido': detalle.detalle_pedido.pk,
                     'codigo': detalle.producto.pk,
                     'nombre': detalle.producto.descripcion,
                     'unidad': detalle.producto.unidad_medida,
                     'cantidad': detalle.cantidad,
                     'precio': detalle.precio,
                     'valor': detalle.valor}
            except:
                d = {'pedido': 0,
                     'codigo': detalle.producto.pk,
                     'nombre': detalle.producto.descripcion,
                     'unidad': detalle.producto.unidad_medida,
                     'cantidad': detalle.cantidad,
                     'precio': detalle.precio,
                     'valor': detalle.valor }
            detalles_data.append(d)
        detalle_salida_formset = DetalleSalidaFormSet(initial=detalles_data)
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_salida_formset = detalle_salida_formset))        
    
    def get_initial(self):
        initial = super(ModificarSalidaAlmacen, self).get_initial()
        movimiento = self.object
        self.detalles = DetalleMovimiento.objects.filter(movimiento=movimiento)
        initial['id_movimiento'] = movimiento.id_movimiento
        initial['fecha'] = movimiento.fecha_operacion.strftime('%d/%m/%Y')
        initial['hora'] = movimiento.fecha_operacion.strftime('%H : %M : %S')
        initial['almacenes'] = movimiento.almacen
        initial['tipos_salida'] = movimiento.tipo_movimiento
        initial['oficina'] = movimiento.oficina
        initial['referencia'] = movimiento.referencia     
        initial['doc_referencia'] = movimiento.tipo_documento
        initial['serie'] = movimiento.serie
        initial['numero'] = movimiento.numero
        initial['total'] = movimiento.total
        initial['observaciones'] = movimiento.observaciones 
        initial['cdetalles'] = self.detalles.count()
        return initial 
        
    def get_context_data(self, **kwargs):
        movimiento = self.object       
        context = super(ModificarSalidaAlmacen, self).get_context_data(**kwargs)
        context['movimiento'] = movimiento
        context['detalles'] = self.detalles        
        return context
               
    def post(self, request, *args, **kwargs):                      
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_salida_formset = DetalleSalidaFormSet(request.POST)
        if form.is_valid() and detalle_salida_formset.is_valid():
            return self.form_valid(form, detalle_salida_formset)
        else:
            return self.form_invalid(form, detalle_salida_formset)
    
    def form_valid(self, form, detalle_salida_formset):
        try:
            with transaction.atomic():
                if self.object.referencia:
                    self.object.eliminar_referencia()
                self.object.eliminar_detalles()
                self.object.eliminar_kardex()
                self.object = form.save()
                referencia = self.object.referencia
                detalles = []
                cont = 1                
                for detalle_salida_form in detalle_salida_formset:
                    detalle_pedido = detalle_salida_form.cleaned_data.get('pedido') 
                    codigo = detalle_salida_form.cleaned_data.get('codigo') 
                    cantidad = detalle_salida_form.cleaned_data.get('cantidad')
                    precio = detalle_salida_form.cleaned_data.get('precio')
                    valor = detalle_salida_form.cleaned_data.get('valor')
                    if cantidad and precio and valor:
                        try:
                            det_ped = DetallePedido.objects.get(pk = detalle_pedido)
                        except:
                            det_ped = None
                        detalle_movimiento = DetalleMovimiento(nro_detalle=cont,
                                                               movimiento=self.object,
                                                               detalle_pedido = det_ped,
                                                               producto=Producto.objects.get(pk=codigo),
                                                               cantidad=cantidad,
                                                               precio=precio,
                                                               valor = valor)
                        detalles.append(detalle_movimiento)                        
                        cont = cont + 1
                DetalleMovimiento.objects.bulk_create(detalles, referencia, self.object.pedido) 
                return HttpResponseRedirect(reverse('almacen:detalle_movimiento', args=[self.object.pk]))
        except IntegrityError:
            messages.error(self.request, 'Error guardando la cotizacion.')
        
    def form_invalid(self, form, detalle_salida_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_salida_formset = detalle_salida_formset))
    
class ModificarAlmacen(UpdateView):
    model = Almacen
    template_name = 'almacen/almacen.html'
    form_class = AlmacenForm
    success_url = reverse_lazy('almacen:almacenes')
    
class ModificarPedido(UpdateView):    
    template_name = 'almacen/pedido.html'
    form_class = PedidoForm
    model = Pedido
    
    @method_decorator(permission_required('almacen.change_pedido',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        pedido = self.get_object()
        if pedido.estado == Pedido.STATUS.PEND:
            return super(ModificarPedido, self).dispatch(*args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))  
    
    def get_form_kwargs(self):
        kwargs = super(ModificarPedido, self).get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_initial(self):
        initial = super(ModificarPedido, self).get_initial()
        pedido = self.object        
        initial['fecha'] = pedido.fecha.strftime('%d/%m/%Y')
        initial['observaciones'] = pedido.observaciones        
        return initial 
    
    def get_context_data(self, **kwargs):
        pedido = self.object
        detalles = DetallePedido.objects.filter(pedido=pedido).order_by('nro_detalle')
        cant_detalles = detalles.count()
        context = super(ModificarPedido, self).get_context_data(**kwargs)
        context['pedido'] = pedido
        context['detalles'] = detalles
        context['cant_detalles'] = cant_detalles
        return context
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.estado == Pedido.STATUS.PEND:            
            form_class = self.get_form_class()
            form = self.get_form(form_class)
            detalles = DetallePedido.objects.filter(pedido = self.object).order_by('nro_detalle')
            detalles_data = []
            for detalle in detalles:
                d = {'codigo': detalle.producto.codigo,
                     'nombre': detalle.producto.descripcion,
                     'unidad': detalle.producto.unidad_medida.codigo,
                     'cantidad': detalle.cantidad}
                detalles_data.append(d)
            detalle_pedido_formset=DetallePedidoFormSet(initial=detalles_data)
            return self.render_to_response(self.get_context_data(form=form,
                                                                 detalle_pedido_formset=detalle_pedido_formset))
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_pedido_formset = DetallePedidoFormSet(request.POST)
        if form.is_valid() and detalle_pedido_formset.is_valid():
            return self.form_valid(form, detalle_pedido_formset)
        else:
            return self.form_invalid(form, detalle_pedido_formset)
    
    def form_valid(self, form, detalle_pedido_formset):
        try:
            with transaction.atomic():
                self.object = form.save()
                DetallePedido.objects.filter(pedido=self.object).delete()
                detalles = []
                cont = 1                
                for detalle_pedido_form in detalle_pedido_formset:
                    codigo = detalle_pedido_form.cleaned_data.get('codigo')
                    cantidad = detalle_pedido_form.cleaned_data.get('cantidad')
                    if codigo and cantidad:
                        producto = Producto.objects.get(codigo=codigo) 
                        detalles.append(DetallePedido(pedido=self.object, 
                                                      nro_detalle=cont, 
                                                      producto=producto, 
                                                      cantidad=cantidad))
                        cont = cont + 1
                DetallePedido.objects.bulk_create(detalles)                
                return HttpResponseRedirect(reverse('almacen:detalle_pedido', args=[self.object.codigo]))
        except IntegrityError:
                messages.error(self.request, 'Error guardando la cotizacion.')
        
    def form_invalid(self, form, detalle_pedido_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_pedido_formset = detalle_pedido_formset))
    
class MovimientosPorProducto(FormView):
    template_name = 'almacen/movimientos_por_producto.html'
    form_class = FormularioMovimientosProducto
    
    def form_valid(self, form):
        data = form.cleaned_data
        desde = data['desde']
        hasta = data['hasta']
        almacen = data['almacen']
        producto = Producto.objects.get(codigo = data['producto'])        
        return self.obtener_movimientos(desde, hasta, almacen, producto)
    
    def obtener_movimientos(self, desde, hasta, almacen, producto):
        detalles = DetalleMovimiento.objects.filter(movimiento__almacen = almacen,
                                                    producto = producto,
                                                    movimiento__fecha_operacion__gte = desde,
                                                    movimiento__fecha_operacion__lte = hasta).order_by('movimiento__fecha_operacion')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = u'Producto: '+ producto.descripcion
        ws.merge_cells('B1:I1')
        ws['B2'] = u'Almacén: '+ almacen.descripcion
        ws.merge_cells('B2:D2')
        ws['E2'] = 'Periodo: Desde: '+ desde.strftime('%d/%m/%Y')+' Hasta: '+ hasta.strftime('%d/%m/%Y')
        ws.merge_cells('E2:H2')
        ws['B4'] = 'MOVIMIENTO'
        ws['C4'] = 'TIPO MOV.'
        ws['D4'] = 'ORDEN COMPRA'
        ws['E4'] = 'PEDIDO'
        ws['F4'] = 'FECHA OPERACION'
        ws['G4']= 'CANTIDAD'
        ws['H4'] = 'PRECIO'
        ws['I4'] = 'VALOR'        
        cont = 5
        for detalle in detalles:
            ws.cell(row=cont,column=2).value = str(detalle.movimiento)
            ws.cell(row=cont,column=3).value = str(detalle.movimiento.tipo_movimiento)
            if detalle.movimiento.referencia is not None:
                ws.cell(row=cont,column=4).value = str(detalle.movimiento.referencia)
            else:
                ws.cell(row=cont,column=4).value = ""
            if detalle.movimiento.pedido is not None:
                ws.cell(row=cont,column=5).value = str(detalle.movimiento.pedido)
            else:
                ws.cell(row=cont,column=5).value = ""
            ws.cell(row=cont,column=6).value = detalle.movimiento.fecha_operacion.strftime('%d/%m/%Y %H : %M : %S')
            ws.cell(row=cont,column=7).value = detalle.cantidad
            ws.cell(row=cont,column=8).value = detalle.precio
            ws.cell(row=cont,column=9).value = detalle.valor
            cont = cont + 1
        nombre_archivo ="MovimientosPorProducto.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class RegistrarIngresoAlmacen(CreateView):
    template_name = 'almacen/ingreso_almacen.html'
    form_class = MovimientoForm
    model = Movimiento
    
    def get_initial(self):
        initial = super(RegistrarIngresoAlmacen, self).get_initial()
        initial['fecha'] = date.today().strftime('%d/%m/%Y')
        initial['total'] = 0        
        return initial
    
    def get_form_kwargs(self):
        kwargs = super(RegistrarIngresoAlmacen, self).get_form_kwargs()
        kwargs['tipo_movimiento'] = 'I'
        return kwargs
    
    def get(self, request, *args, **kwargs):
        self.object = None
        cod_tipo_mov = 'I00'
        tipos_ingreso = TipoMovimiento.objects.filter(incrementa=True).exclude(codigo=cod_tipo_mov)        
        if not tipos_ingreso:
            return HttpResponseRedirect(reverse('almacen:crear_tipo_movimiento'))
        almacenes = Almacen.objects.all()
        cant_suministros = Producto.objects.count()
        if almacenes.count()>0:
            if cant_suministros > 0:                    
                form_class = self.get_form_class()
                form = self.get_form(form_class)
                detalle_ingreso_formset=DetalleIngresoFormSet()
                return self.render_to_response(self.get_context_data(form=form,
                                                                     detalle_ingreso_formset=detalle_ingreso_formset))                
        return HttpResponseRedirect(reverse('almacen:tablero'))
    
    def post(self, request, *args, **kwargs):                      
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_ingreso_formset = DetalleIngresoFormSet(request.POST)
        if form.is_valid() and detalle_ingreso_formset.is_valid():
            return self.form_valid(form, detalle_ingreso_formset)
        else:
            return self.form_invalid(form, detalle_ingreso_formset)
    
    def form_valid(self, form, detalle_ingreso_formset):
        try:
            with transaction.atomic():
                self.object = form.save()
                referencia = self.object.referencia
                detalles = []
                cont = 1                
                for detalle_ingreso_form in detalle_ingreso_formset:
                    orden_compra = detalle_ingreso_form.cleaned_data.get('orden_compra')
                    codigo = detalle_ingreso_form.cleaned_data.get('codigo') 
                    cantidad = detalle_ingreso_form.cleaned_data.get('cantidad')
                    precio = detalle_ingreso_form.cleaned_data.get('precio')
                    valor = detalle_ingreso_form.cleaned_data.get('valor')
                    if cantidad and precio and valor:
                        try:
                            detalle_orden_compra = DetalleOrdenCompra.objects.get(pk=orden_compra)
                            detalle_movimiento = DetalleMovimiento(detalle_orden_compra=detalle_orden_compra,
                                                                   nro_detalle=cont,
                                                                   movimiento=self.object,
                                                                   producto=Producto.objects.get(pk=codigo),
                                                                   cantidad=cantidad,
                                                                   precio=precio,
                                                                   valor = valor)
                        except:
                            detalle_movimiento = DetalleMovimiento(nro_detalle=cont,
                                                                   movimiento=self.object,
                                                                   producto=Producto.objects.get(pk=codigo),
                                                                   cantidad=cantidad,
                                                                   precio=precio,
                                                                   valor=valor)
                        detalles.append(detalle_movimiento)                        
                        cont = cont + 1
                DetalleMovimiento.objects.bulk_create(detalles, referencia, None) 
                return HttpResponseRedirect(reverse('almacen:detalle_movimiento', args=[self.object.pk]))
        except IntegrityError:
                messages.error(self.request, 'Error guardando la cotizacion.')
        
    def form_invalid(self, form, detalle_ingreso_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_ingreso_formset=detalle_ingreso_formset))
        
class RegistrarSalidaAlmacen(CreateView):
    form_class = MovimientoForm
    template_name = "almacen/salida_almacen.html"
    model = Movimiento
    
    def get_form_kwargs(self):
        kwargs = super(RegistrarSalidaAlmacen, self).get_form_kwargs()
        kwargs['tipo_movimiento'] = 'S'
        return kwargs
    
    def get_initial(self):
        initial = super(RegistrarSalidaAlmacen, self).get_initial()
        initial['total'] = 0
        initial['fecha'] = date.today().strftime('%d/%m/%Y')
        return initial
    
    def get(self, request, *args, **kwargs):
        self.object = None
        tipos_salida = TipoMovimiento.objects.filter(incrementa=False)
        if not tipos_salida:
            return HttpResponseRedirect(reverse('almacen:crear_tipo_movimiento'))    
        almacenes = Almacen.objects.filter()  
        if not almacenes:
            return HttpResponseRedirect(reverse('almacen:crear_almacen'))      
        else:
            form_class = self.get_form_class()
            form = self.get_form(form_class)
            detalle_salida_formset=DetalleSalidaFormSet()
            return self.render_to_response(self.get_context_data(form=form,
                                                                 detalle_salida_formset=detalle_salida_formset))
    
    def post(self, request, *args, **kwargs):                      
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_salida_formset = DetalleSalidaFormSet(request.POST)
        if form.is_valid() and detalle_salida_formset.is_valid():
            return self.form_valid(form, detalle_salida_formset)
        else:
            return self.form_invalid(form, detalle_salida_formset)
    
    def form_valid(self, form, detalle_salida_formset):
        try:
            with transaction.atomic():
                self.object = form.save()
                referencia = self.object.referencia
                detalles = []
                cont = 1                
                for detalle_salida_form in detalle_salida_formset:
                    codigo = detalle_salida_form.cleaned_data.get('codigo') 
                    cantidad = detalle_salida_form.cleaned_data.get('cantidad')
                    precio = detalle_salida_form.cleaned_data.get('precio')
                    valor = detalle_salida_form.cleaned_data.get('valor')
                    if cantidad and precio and valor:
                        detalle_movimiento = DetalleMovimiento(nro_detalle=cont,
                                                               movimiento=self.object,
                                                               producto=Producto.objects.get(pk=codigo),
                                                               cantidad=cantidad,
                                                               precio=precio,
                                                               valor=valor)
                        detalles.append(detalle_movimiento)                        
                        cont = cont + 1
                DetalleMovimiento.objects.bulk_create(detalles, referencia, None) 
                return HttpResponseRedirect(reverse('almacen:detalle_movimiento', args=[self.object.pk]))
        except IntegrityError:
                messages.error(self.request, 'Error guardando la cotizacion.')
        
    def form_invalid(self, form, detalle_salida_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_salida_formset = detalle_salida_formset))
    
class ReporteExcelAlmacenes(TemplateView):
    
    def get(self, request, *args, **kwargs):
        almacenes = Almacen.objects.filter(estado=True).order_by('codigo')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE ALMACENES'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'DESCRIPCIÓN'
        cont=4
        for almacen in almacenes:
            ws.cell(row=cont,column=2).value = almacen.codigo
            ws.cell(row=cont,column=3).value = almacen.descripcion
            cont = cont + 1
        nombre_archivo ="ListadoAlmacenes.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
class ReporteExcelTiposMovimientos(TemplateView):
    
    def get(self, request, *args, **kwargs):
        tipos = TipoMovimiento.objects.filter(estado=True).order_by('codigo')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE TIPOS DE MOVIMIENTOS'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'DESCRIPCIÓN'
        cont=4
        for tipo in tipos:
            ws.cell(row=cont,column=2).value = tipo.codigo
            ws.cell(row=cont,column=3).value = tipo.descripcion
            cont = cont + 1
        nombre_archivo ="MaestroTiposMovimientos.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteKardexProducto(FormView):
    template_name = 'almacen/reporte_kardex_producto.html'
    form_class = FormularioKardexProducto
    
    def form_valid(self, form):
        data = form.cleaned_data
        cod_prod = data.get('cod_producto')
        producto = Producto.objects.get(codigo = cod_prod)
        desde = data.get('desde')
        hasta = data.get('hasta')
        almacen = data.get('almacenes')
        formato_sunat = data.get('formato_sunat')
        formatos = data.get('formatos')
        if formatos == "XLS":
            if formato_sunat == 'S':
                return self.obtener_formato_sunat_unidades_fisicas_excel(producto,
                                                                         desde,
                                                                         hasta,
                                                                         almacen)
            elif formato_sunat == 'V':
                return self.obtener_formato_sunat_valorizado_excel(producto,
                                                                   desde,
                                                                   hasta,
                                                                   almacen)
            else:
                return self.obtener_formato_normal_excel(producto,
                                                         desde,
                                                         hasta,
                                                         almacen)
        elif formatos == "PDF":
            if formato_sunat == 'S':
                return self.obtener_formato_sunat_unidades_fisicas_pdf(producto,
                                                                       desde,
                                                                       hasta,
                                                                       almacen)
            elif formato_sunat == 'V':
                return self.obtener_formato_sunat_valorizado_pdf(producto,
                                                                 desde,
                                                                 hasta,
                                                                 almacen)
            else:
                return self.obtener_formato_normal_pdf(producto,
                                                       desde,
                                                       hasta,
                                                       almacen)

    def obtener_formato_sunat_unidades_fisicas_pdf(self, producto, desde, hasta, almacen):
        response = HttpResponse(content_type='application/pdf')
        reporte = ReporteKardexPDF('A4', desde, hasta, almacen, False)
        pdf = reporte.imprimir_formato_sunat_unidades_fisicas_producto(producto)
        response['Content-Disposition'] = 'attachment; filename=InventarioPermanenteUnidadesFisicas.pdf'
        response.write(pdf)
        return response

    def obtener_formato_sunat_valorizado_pdf(self, producto, desde, hasta, almacen):
        response = HttpResponse(content_type='application/pdf')
        reporte = ReporteKardexPDF('A4', desde, hasta, almacen, False)
        pdf = reporte.imprimir_formato_sunat_valorizado_producto(producto)
        response['Content-Disposition'] = 'attachment; filename=InventarioPermanenteValorizado.pdf'
        response.write(pdf)
        return response

    def obtener_formato_sunat_unidades_fisicas_excel(self, producto, desde, hasta, almacen):
        reporte = ReporteKardexExcel()
        excel = reporte.obtener_formato_sunat_unidades_fisicas_producto(producto, desde, hasta, almacen)
        nombre_archivo = "InventarioPermanenteUnidadesFisicas.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        excel.save(response)
        return response

    def obtener_formato_sunat_valorizado_excel(self, producto, desde, hasta, almacen):
        reporte = ReporteKardexExcel()
        excel = reporte.obtener_formato_sunat_valorizado_producto(producto, desde, hasta, almacen)
        nombre_archivo = "InventarioPermanenteValorizado.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        excel.save(response)
        return response

    def obtener_formato_normal_excel(self, producto, desde, hasta, almacen):
        reporte = ReporteKardexExcel()
        excel = reporte.obtener_formato_normal_producto(producto, desde, hasta, almacen)
        nombre_archivo = "ReporteExcelKardexProducto.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        excel.save(response)
        return response

class ReporteKardex(FormView):
    template_name = 'almacen/reporte_kardex.html'
    form_class = FormularioKardexProducto

    def form_valid(self, form):
        data = form.cleaned_data
        desde = data.get('desde')
        hasta = data['hasta']
        almacen = data.get('almacenes')
        formato_sunat = data.get('formato_sunat')
        formatos = data.get('formatos')
        consolidado = data['consolidado']

        if formatos == "XLS":
            if consolidado == 'P':
                return self.obtener_consolidado_productos_excel(desde, hasta, almacen)
            elif consolidado == 'G':
                return self.obtener_consolidado_grupos_excel(desde, hasta, almacen)
            else:
                if formato_sunat == 'S':
                    return self.obtener_formato_sunat_unidades_fisicas_excel(desde, hasta, almacen)
                elif formato_sunat == 'V':
                    return self.obtener_formato_sunat_valorizado_excel(desde, hasta, almacen)
                else:
                    return self.obtener_formato_normal_excel(desde, hasta, almacen)
        elif formatos == "PDF":
            if consolidado == 'P':
                return self.obtener_consolidado_productos_pdf(desde, hasta, almacen)
            elif consolidado == 'G':
                return self.obtener_consolidado_grupos_pdf(desde, hasta, almacen)
            else:
                if formato_sunat == 'S':
                    return self.obtener_formato_sunat_unidades_fisicas_pdf(desde, hasta, almacen)
                elif formato_sunat == 'V':
                    return self.obtener_formato_sunat_valorizado_pdf(desde, hasta, almacen)
                else:
                    return self.obtener_formato_normal_pdf(desde, hasta, almacen)

    def obtener_consolidado_productos_pdf(self, desde, hasta, almacen):
        response = HttpResponse(content_type='application/pdf')
        reporte = ReporteKardexPDF('A4',desde, hasta, almacen, False)
        pdf = reporte.imprimir_formato_consolidado_productos()
        response['Content-Disposition'] = 'attachment; filename=ResumenMensualDeAlmacen.pdf'
        response.write(pdf)
        return response

    def obtener_consolidado_grupos_pdf(self, desde, hasta, almacen):
        response = HttpResponse(content_type='application/pdf')
        reporte = ReporteKardexPDF('A4',desde, hasta, almacen, True)
        pdf = reporte.imprimir_formato_consolidado_grupos()
        response['Content-Disposition'] = 'attachment; filename=ResumenMensualDeAlmacenPorGruposYCuentas.pdf'
        response.write(pdf)
        return response

    def obtener_formato_sunat_unidades_fisicas_pdf(self, desde, hasta, almacen):
        response = HttpResponse(content_type='application/pdf')
        reporte = ReporteKardexPDF('A4',desde, hasta, almacen, False)
        pdf = reporte.imprimir_formato_sunat_unidades_fisicas_todos()
        response['Content-Disposition'] = 'attachment; filename=InventarioPermanenteUnidadesFisicas.pdf'
        response.write(pdf)
        return response

    def obtener_formato_sunat_valorizado_pdf(self, desde, hasta, almacen):
        response = HttpResponse(content_type='application/pdf')
        reporte = ReporteKardexPDF('A4',desde, hasta, almacen,False)
        pdf = reporte.imprimir_formato_sunat_valorizado_todos()
        response['Content-Disposition'] = 'attachment; filename=InventarioPermanenteValorizado.pdf'
        response.write(pdf)
        return response

    def obtener_formato_normal_excel(self, desde, hasta, almacen):
        reporte = ReporteKardexExcel()
        excel = reporte.obtener_formato_normal_todos(desde, hasta, almacen)
        nombre_archivo = "ReporteFormatoNormalKardexTodosLosProductos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        excel.save(response)
        return response

    def obtener_formato_sunat_unidades_fisicas_excel(self, desde, hasta, almacen):
        reporte = ReporteKardexExcel()
        excel = reporte.obtener_formato_sunat_unidades_fisicas_todos(desde, hasta, almacen)
        nombre_archivo = "InventarioPermanenteUnidadesFisicas.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        excel.save(response)
        return response

    def obtener_formato_sunat_valorizado_excel(self, desde, hasta, almacen):
        reporte = ReporteKardexExcel()
        excel = reporte.obtener_formato_sunat_valorizado_todos(desde, hasta, almacen)
        nombre_archivo = "InventarioPermanenteValorizado.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        excel.save(response)
        return response

    def obtener_consolidado_productos_excel(self, desde, hasta, almacen):
        reporte = ReporteKardexExcel()
        excel = reporte.obtener_consolidado_productos(desde, hasta, almacen)
        nombre_archivo = "ReporteConsolidadoKardexExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        excel.save(response)
        return response

    def obtener_consolidado_grupos_excel(self, desde, hasta, almacen):
        reporte = ReporteKardexExcel()
        excel = reporte.obtener_consolidado_grupos(desde, hasta, almacen)
        nombre_archivo = "ReporteConsolidadoCuentasContablesAlmacen.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        excel.save(response)
        return response

class ReporteStock(FormView):
    template_name = 'almacen/reporte_stock.html'
    form_class = FormularioReporteStock    
    
    def post(self, request, *args, **kwargs):
        r_almacen = request.POST['almacenes']
        return HttpResponseRedirect(reverse('almacen:listado_stock',args=[r_almacen]))
    
class ReprocesoPrecio(FormView):
    template_name = 'almacen/reproceso_precio.html'
    form_class = FormularioReprocesoPrecio
    
    def reprocesar_precio_producto(self, producto, almacen, desde):
        detalles = Kardex.objects.filter(producto = producto,
                                         almacen = almacen,
                                         fecha_operacion__gte=desde).order_by('fecha_operacion')
        indice = 0
        for detalle in detalles:
            try:
                anterior = detalles[indice - 1]
                cantidad_ant = anterior.cantidad_total                
                valor_ant = anterior.valor_total
                precio_ant = Decimal(round(valor_ant / cantidad_ant, 8))
            except:
                cantidad_ant = 0
                precio_ant = 0
                valor_ant = 0
            tipo_mov = detalle.movimiento.tipo_movimiento
            if tipo_mov.incrementa:
                detalle.cantidad_total = cantidad_ant + detalle.cantidad_ingreso  
                detalle.precio_total = detalle.precio_ingreso
                detalle.valor_total = valor_ant + detalle.valor_ingreso
            else:
                detalle.precio_salida = precio_ant
                detalle.valor_salida = detalle.cantidad_salida * detalle.precio_salida                
                detalle.cantidad_total = cantidad_ant - detalle.cantidad_salida
                detalle.valor_total = valor_ant - detalle.valor_salida
                try:
                    detalle.precio_total = detalle.valor_total / detalle.cantidad_total
                except:
                    detalle.precio_total = 0 
            detalle.save()   
            indice = indice + 1
    
    def form_valid(self, form):
        data = form.cleaned_data        
        desde = data['desde']
        almacen = data['almacen']
        seleccion = data['seleccion']
        if seleccion == 'P':
            cod_prod = data['producto']
            producto = Producto.objects.get(codigo=cod_prod)
            self.reprocesar_precio_producto(producto, almacen, desde)
        else:
            listado_kardex = Kardex.objects.filter(almacen = almacen).order_by('producto').distinct('producto__codigo')
            for kardex in listado_kardex:                
                self.reprocesar_precio_producto(kardex.producto, almacen, desde)
        return HttpResponseRedirect(reverse('almacen:tablero'))
    
class StockProductos(FormView):
    form_class = FormularioConsultaStock
    template_name = 'almacen/stock_productos.html'

    def get_initial(self):
        initial = super(StockProductos, self).get_initial()
        initial['desde'] = date.today().strftime('%d/%m/%Y')
        return initial

    def form_valid(self, form):
        data = form.cleaned_data
        desde = data['desde']
        almacen = data['almacen']
        descripcion = data['descripcion']
        productos = Producto.objects.filter(descripcion__icontains=descripcion).order_by('descripcion')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'STOCK DE PRODUCTOS'
        ws.merge_cells('B1:E1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'DESCRIPCION'
        ws['D3'] = 'UNIDAD'
        ws['E3'] = 'CANTIDAD'
        ws['F3'] = 'PRECIO'
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 40
        cont = 4
        for producto in productos:
            try:
                kardex = Kardex.objects.filter(producto=producto,
                                               almacen=almacen).latest('fecha_operacion')
                codigo = kardex.producto.codigo
                descripcion = kardex.producto.descripcion
                unidad_medida = kardex.producto.unidad_medida.descripcion
                stock = kardex.cantidad_total
                precio = kardex.precio_total
            except:
                codigo = producto.codigo
                descripcion = producto.descripcion
                unidad_medida = producto.unidad_medida.codigo
                stock = 0
                precio = 0
            ws.cell(row=cont, column=2).value = codigo
            ws.cell(row=cont, column=3).value = descripcion
            ws.cell(row=cont, column=4).value = unidad_medida
            ws.cell(row=cont, column=5).value = stock
            ws.cell(row=cont, column=6).value = precio
            cont = cont + 1
        nombre_archivo = "ReporteStock.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ListadoStockProducto(TemplateView):

    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            descripcion = request.GET['descripcion']
            desde = request.GET['desde']
            almacen = request.GET['almacen']
            lista_productos = []
            productos = Producto.objects.filter(descripcion__icontains=descripcion).order_by('descripcion')
            for producto in productos:
                try:
                    kardex = Kardex.objects.filter(producto=producto,
                                                   almacen__pk=almacen).latest('fecha_operacion')
                    kardex_json = {}
                    kardex_json['codigo'] = kardex.producto.codigo
                    kardex_json['label'] = kardex.producto.descripcion
                    kardex_json['unidad'] = kardex.producto.unidad_medida.codigo
                    kardex_json['stock'] = kardex.cantidad_total
                except:
                    kardex_json = {}
                    kardex_json['codigo'] = producto.codigo
                    kardex_json['label'] = producto.descripcion
                    kardex_json['unidad'] = producto.unidad_medida.codigo
                    kardex_json['stock'] = 0
                lista_productos.append(kardex_json)
            data = simplejson.dumps(lista_productos)
            return HttpResponse(data, 'application/json')

class ReporteExcelMovimientos(FormView):
    form_class = FormularioReporteMovimientos
    template_name = "almacen/reporte_movimientos.html"    

    def form_valid(self, form):
        data = form.cleaned_data
        tipo_busqueda = data['tipo_busqueda']
        p_almacen = data['almacenes']
        p_tipo_movimiento = data['tipos_movimiento']
        almacen = Almacen.objects.get(codigo=p_almacen)
        tipo_movimiento = TipoMovimiento.objects.get(codigo=p_tipo_movimiento)
        wb = Workbook()
        ws = wb.active
        if tipo_busqueda=='F':
            fecha_inicio = data['desde']
            fecha_final = data['hasta']
            ws['B1'] = 'REPORTE DE MOVIMIENTOS POR FECHA'
            ws.merge_cells('B1:H1')
            ws['B2'] = 'ALMACEN: '+ almacen.descripcion
            ws.merge_cells('B2:D2')
            ws['E2'] = 'TIPO DE MOVIMIENTO: '+ tipo_movimiento.descripcion
            ws.merge_cells('E2:H2')
            ws['B3'] = 'DESDE'
            ws['C3'] = fecha_inicio
            ws['C3'].number_format = 'dd/mm/yyyy'
            ws['D3'] = 'HASTA'
            ws['E3'] = fecha_final
            ws['F3'].number_format = 'dd/mm/yyyy'        
            movimientos = Movimiento.objects.filter(fecha_operacion__range=[fecha_inicio, fecha_final],
                                                    tipo_movimiento=tipo_movimiento,almacen=almacen)
        elif tipo_busqueda=='M':
            mes = data['mes'].strip()
            annio = data['annio'].strip()            
            ws['B1'] = 'REPORTE DE MOVIMIENTOS POR MES'
            ws.merge_cells('B1:H1')
            ws['B2'] = 'ALMACEN: '+ almacen.descripcion
            ws.merge_cells('B2:D2')
            ws['E2'] = 'TIPO DE MOVIMIENTO: '+ tipo_movimiento.descripcion
            ws.merge_cells('E2:H2')
            ws['B3'] = 'MES'
            ws['C3'] = mes
            ws['D3'] = 'AÑO'
            ws['E3'] = annio                    
            movimientos = Movimiento.objects.filter(fecha_operacion__month=mes,
                                                    fecha_operacion__year=annio,
                                                    tipo_movimiento=tipo_movimiento,
                                                    almacen=almacen)
        elif tipo_busqueda=='A':
            annio = data['annio'].strip()
            ws['B1'] = 'REPORTE DE MOVIMIENTOS POR AÑO'
            ws.merge_cells('B1:H1')
            ws['B2'] = 'ALMACEN: '+ almacen.descripcion
            ws.merge_cells('B2:D2')
            ws['E2'] = 'TIPO DE MOVIMIENTO: '+ tipo_movimiento.descripcion
            ws.merge_cells('E2:H2')
            ws['B3'] = 'AÑO'
            ws['C3'] = annio
            movimientos = Movimiento.objects.filter(fecha_operacion__year=annio,
                                                    tipo_movimiento=tipo_movimiento,
                                                    almacen=almacen)
        ws['B5'] = 'ID_MOVIMIENTO'
        ws['C5'] = 'TIPO_DOCUMENTO'
        ws['D5'] = 'SERIE'
        ws['E5'] = 'NUMERO'
        ws['F5'] = 'FECHA_OPERACION'
        ws['G5'] = 'OBSERVACION'
        ws['H5'] = 'FECHA_CREACION'
        ws['I5'] = 'ESTADO'
        cont=6
        movimientos = movimientos.order_by('fecha_operacion')
        for movimiento in movimientos:
            ws.cell(row=cont,column=2).value = movimiento.id_movimiento
            try:
                ws.cell(row=cont,column=3).value = movimiento.tipo_documento.descripcion
            except:
                ws.cell(row=cont, column=3).value = '--'
            ws.cell(row=cont,column=4).value = movimiento.serie
            ws.cell(row=cont,column=5).value = movimiento.numero
            ws.cell(row=cont,column=6).value = movimiento.fecha_operacion
            ws.cell(row=cont,column=6).number_format = 'dd/mm/yyyy hh:mm:ss'
            ws.cell(row=cont,column=7).value = movimiento.observaciones
            ws.cell(row=cont,column=8).value = movimiento.created    
            ws.cell(row=cont,column=8).number_format = 'dd/mm/yyyy hh:mm:ss'
            ws.cell(row=cont,column=9).value = movimiento.estado            
            cont = cont + 1
        nombre_archivo ="ReporteMovimientosPorFecha.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
        
class ReporteExcelMovimientosPorFecha(View):
    
    def get(self, request, *args, **kwargs):
        p_fecha_inicio = kwargs['fecha_inicio']
        p_fecha_final = kwargs['fecha_fin']
        p_almacen = kwargs['almacen']
        p_tipo_movimiento = kwargs['tipo_movimiento']
        anio = int(p_fecha_inicio[6:])        
        mes = int(p_fecha_inicio[3:5])
        dia = int(p_fecha_inicio[0:2])
        fecha_inicio = datetime.datetime(anio,mes,dia,23,59,59)
        anio = int(p_fecha_final[6:])        
        mes = int(p_fecha_final[3:5])
        dia = int(p_fecha_final[0:2])
        fecha_final = datetime.datetime(anio,mes,dia,23,59,59)
        almacen = Almacen.objects.get(codigo=p_almacen)
        tipo_movimiento = TipoMovimiento.objects.get(codigo=p_tipo_movimiento)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE MOVIMIENTOS POR FECHA'
        ws.merge_cells('B1:H1')
        ws['B2'] = 'ALMACEN: '+ almacen.descripcion
        ws.merge_cells('B2:D2')
        ws['E2'] = 'TIPO DE MOVIMIENTO: '+ tipo_movimiento.descripcion
        ws.merge_cells('E2:H2')
        ws['B3'] = 'DESDE'
        ws['C3'] = p_fecha_inicio
        ws['C3'].number_format = 'dd/mm/yyyy'
        ws['D3'] = 'HASTA'
        ws['E3'] = p_fecha_final
        ws['F3'].number_format = 'dd/mm/yyyy'        
        movimientos = Movimiento.objects.filter(fecha_operacion__range=[fecha_inicio, fecha_final],tipo_movimiento=tipo_movimiento,almacen=almacen)
        ws['B5'] = 'ID_MOVIMIENTO'
        ws['C5'] = 'TIPO_DOCUMENTO'
        ws['D5'] = 'SERIE'
        ws['E5'] = 'NUMERO'
        ws['F5'] = 'FECHA_OPERACION'
        ws['G5'] = 'OBSERVACION'
        ws['H5'] = 'FECHA_CREACION'
        cont=6
        for movimiento in movimientos:
            ws.cell(row=cont,column=2).value = movimiento.id_movimiento
            ws.cell(row=cont,column=3).value = movimiento.tipo_documento
            ws.cell(row=cont,column=4).value = movimiento.serie
            ws.cell(row=cont,column=5).value = movimiento.numero
            ws.cell(row=cont,column=6).value = movimiento.fecha_operacion
            ws.cell(row=cont,column=6).number_format = 'dd/mm/yyyy hh:mm:ss'
            ws.cell(row=cont,column=7).value = movimiento.observacion
            ws.cell(row=cont,column=8).value = movimiento.created    
            ws.cell(row=cont,column=8).number_format = 'dd/mm/yyyy hh:mm:ss'        
            cont = cont + 1
        nombre_archivo ="ReporteMovimientosPorFecha.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
class ReportePDFMovimiento(View):
    
    def get(self, request, *args, **kwargs): 
        id_movimiento = kwargs['id_movimiento']
        movimiento = Movimiento.objects.get(pk=id_movimiento)
        response = HttpResponse(content_type='application/pdf')                
        reporte = ReporteMovimiento('A4', movimiento)
        pdf = reporte.imprimir()        
        response.write(pdf)
        return response

class ReportePDFProductos(View):    
    
    def get(self, request, *args, **kwargs): 
        response = HttpResponse(content_type='application/pdf')
        pdf_name = "clientes.pdf"  # llamado clientes
        # la linea 26 es por si deseas descargar el pdf a tu computadora
        # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
        buff = BytesIO()
        doc = SimpleDocTemplate(buff,
                                pagesize=letter,
                                rightMargin=40,
                                leftMargin=40,
                                topMargin=60,
                                bottomMargin=18,
                                )
        clientes = []
        styles = getSampleStyleSheet()
        header = Paragraph("Listado de Clientes", styles['Heading1'])
        clientes.append(header)
        headings = ('Nombre', 'Email', 'Edad', 'Direccion')
        allclientes = [(p.codigo, p.descripcion, p.precio_mercado, p.grupo_suministros) for p in Producto.objects.all()]
            
        t = Table([headings] + allclientes)
        t.setStyle(TableStyle(
            [
                ('GRID', (0, 0), (3, -1), 1, colors.dodgerblue),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.darkblue),
                ('BACKGROUND', (0, 0), (-1, 0), colors.dodgerblue)
            ]
        ))
        clientes.append(t)
        doc.build(clientes)
        response.write(buff.getvalue())
        buff.close()
        return response
    
class PopupCrearTipoStock(FormView):
    template_name = 'almacen/popup_crear_tipo_stock.html'
    form_class = TipoStockForm
    success_url = reverse_lazy('almacen:crear_suministro')

    def form_valid(self, form):
        form.save()
        return super(PopupCrearTipoStock, self).form_valid(form)
    
class VerificarSolicitaDocumento(TemplateView):
    
    def get(self, request, *args, **kwargs):
        tipo = request.GET['tipo']
        tipo_movimiento = TipoMovimiento.objects.get(pk=tipo)
        json_object = {'solicita_documento': tipo_movimiento.solicita_documento}
        return JsonResponse(json_object)
    
class VerificarPideReferencia(TemplateView):
    
    def get(self, request, *args, **kwargs):
        tipo = request.GET['tipo']
        tipo_movimiento = TipoMovimiento.objects.get(pk=tipo)
        json_object = {'pide_referencia': tipo_movimiento.pide_referencia}
        return JsonResponse(json_object)
    
class VerificarStockParaPedido(TemplateView):
    
    def get(self, request, *args, **kwargs):
        almacen = request.GET['almacen']
        pedido = request.GET['pedido']
        detalles = DetallePedido.objects.filter(pedido__codigo = pedido, 
                                                estado=DetallePedido.STATUS.PEND).order_by('nro_detalle')
        lista_detalles = []
        for detalle in detalles:
            try:
                control_producto = Kardex.objects.filter(producto = detalle.producto,
                                                almacen__codigo=almacen).latest('fecha_operacion')
                
                stock = control_producto.cantidad_total
                precio = control_producto.valor_total / stock
            except:
                stock = 0
                precio = 0            
            if stock <> 0:                
                det = {}
                det['pedido'] = detalle.id
                det['codigo'] = detalle.producto.codigo               
                det['nombre'] = detalle.producto.descripcion                     
                det['unidad'] = detalle.producto.unidad_medida.descripcion                
                cantidad = detalle.cantidad-detalle.cantidad_atendida
                if cantidad > stock:
                    cantidad = stock
                valor = round(cantidad * precio, 5)
                det['cantidad'] = cantidad
                det['precio'] = round(precio,5)                
                det['valor'] = valor
                lista_detalles.append(det)
        formset = DetalleSalidaFormSet(initial=lista_detalles)
        lista_json = []
        for form in formset:
            detalle_json = {}    
            detalle_json['pedido'] = str(form['pedido'])
            detalle_json['codigo'] = str(form['codigo'])
            detalle_json['nombre'] = str(form['nombre'])
            detalle_json['cantidad'] = str(form['cantidad'])
            detalle_json['precio'] = str(form['precio'])
            detalle_json['unidad'] = str(form['unidad'])                
            detalle_json['valor'] = str(form['valor']) 
            lista_json.append(detalle_json)
        data = json.dumps(lista_json)
        return HttpResponse(data, 'application/json')