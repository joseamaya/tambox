# -*- coding: utf-8 -*- 
from django.views.generic.base import View, TemplateView
from django.views.generic.list import ListView
from django.views.generic.edit import FormView, UpdateView, CreateView
from django.core.urlresolvers import reverse_lazy, reverse
from django.http.response import HttpResponseRedirect
import json
from django.http import HttpResponse
import simplejson
from openpyxl import Workbook
from django.views.generic.detail import DetailView
from django.conf import settings
import csv
from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
import os
from contabilidad.forms import UploadForm
from django.shortcuts import render
from productos.models import Producto, UnidadMedida, GrupoProductos
from productos.forms import GrupoProductosForm, ProductoForm, ServicioForm,\
    UnidadMedidaForm
from contabilidad.models import CuentaContable, TipoExistencia
#from productos.reports import ProductosReport

# Create your views here.
class Tablero(View):
    
    def get(self, request, *args, **kwargs):
        lista_notificaciones = []
        cant_productos = Producto.objects.filter(es_servicio=False).count()
        cant_tipos_unidad_medida = UnidadMedida.objects.count()
        cant_grupos_suministros = GrupoProductos.objects.count()
        cant_servicios = Producto.objects.filter(es_servicio=True).count()
        unidad_medida, creado = UnidadMedida.objects.get_or_create(codigo = 'SERV',
                                                                   defaults = {'descripcion':'SERVICIO'})
        if creado:
            lista_notificaciones.append("Se ha creado la unidad de medida SERVICIO")
        if cant_productos == 0:
            lista_notificaciones.append("No se ha creado ningún producto")
        if cant_tipos_unidad_medida == 0:
            lista_notificaciones.append("No se ha creado ningún tipo de unidad de medida")
        if cant_grupos_suministros == 0:
            lista_notificaciones.append("No se ha creado ningún grupo de productos")
        if cant_servicios == 0:
            lista_notificaciones.append("No se ha creado ningún servicio")
        context = {'notificaciones':lista_notificaciones}
        return render(request, 'productos/tablero_productos.html', context)
    
class BusquedaProductosDescripcion(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            descripcion = request.GET['descripcion']
            tipo_busqueda =  request.GET['tipo_busqueda']
            if tipo_busqueda == 'TODOS':
                productos = Producto.objects.filter(descripcion__icontains = descripcion).order_by('descripcion')[:20]
            elif tipo_busqueda == 'PRODUCTOS':
                productos = Producto.objects.filter(descripcion__icontains = descripcion,es_servicio=False).order_by('descripcion')[:20]
            elif tipo_busqueda == 'SERVICIOS':
                productos = Producto.objects.filter(descripcion__icontains = descripcion,es_servicio=True).order_by('descripcion')[:20]
                
            lista_productos = []
            for producto in productos:
                producto_json = {}                
                producto_json['label'] = producto.descripcion                
                producto_json['codigo'] = producto.codigo
                producto_json['descripcion'] = producto.descripcion
                producto_json['unidad'] = producto.unidad_medida.descripcion
                producto_json['precio'] = str(producto.precio)            
                lista_productos.append(producto_json)                            
            data = json.dumps(lista_productos)
            return HttpResponse(data, 'application/json')
        
class BusquedaProductosCodigo(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            productos = Producto.objects.filter(codigo__icontains = codigo)[:20]    
            lista_productos = []
            for producto in productos:
                producto_json = {}                
                producto_json['label'] = producto.codigo
                producto_json['codigo'] = producto.codigo
                producto_json['descripcion'] = producto.descripcion
                producto_json['unidad'] = producto.unidad_medida.descripcion
                producto_json['precio'] = str(producto.precio)            
                lista_productos.append(producto_json)                            
            data = json.dumps(lista_productos)
            return HttpResponse(data, 'application/json')
        
class CargarGrupoProductos(FormView):
    template_name = 'productos/cargar_grupo_productos.html'
    form_class = UploadForm
    
    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']            
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT,'archivos',str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        for fila in dataReader:
            try:
                cuenta = CuentaContable.objects.get(cuenta=fila[0])
                descripcion = fila[1] 
                grupo_productos, creado = GrupoProductos.objects.get_or_create(descripcion=unicode(descripcion, errors='ignore'),
                                                                               defaults={'ctacontable' : cuenta
                                                                                        })
            except CuentaContable.DoesNotExist:
                pass                       
        return HttpResponseRedirect(reverse('productos:grupos_productos'))
        
class CargarServicios(FormView):
    template_name = 'productos/cargar_servicios.html'
    form_class = UploadForm
    
    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']            
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT,'archivos',str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        try:            
            for fila in dataReader:
                grupo = GrupoProductos.objects.get(codigo = fila[0].strip()) 
                producto, creado = Producto.objects.get_or_create(descripcion=unicode(fila[1], errors='ignore'),
                                                                  defaults={'grupo_productos' : grupo,
                                                                            'es_servicio' : True})                                    
            return HttpResponseRedirect(reverse('productos:servicios'))
        except GrupoProductos.DoesNotExist:
            return HttpResponseRedirect(reverse('productos:crear_grupo_productos'))
        
class CargarProductos(FormView):
    template_name = 'productos/cargar_productos.html'
    form_class = UploadForm
    
    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']            
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT,'archivos',str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        for fila in dataReader:
            grupo = GrupoProductos.objects.get(codigo = fila[0].strip())
            cod_und = fila[2][0:5]
            und, creado = UnidadMedida.objects.get_or_create(codigo=cod_und.strip(),
                                                             defaults={'codigo': cod_und,
                                                                       'descripcion' : fila[2].strip()})            
            if fila[3]<>'':
                precio = fila[3]
            else:
                precio = 0
            try:                
                tipo_existencia = TipoExistencia.objects.get(codigo_sunat = fila[5].strip())
            except:
                return HttpResponseRedirect(reverse('contabilidad:tablero'))
            try:
                producto, creado = Producto.objects.get_or_create(descripcion=unicode(fila[1].strip(), errors='ignore'),
                                                                  defaults={'unidad_medida' : und,
                                                                            'grupo_productos' : grupo,
                                                                            'precio' : precio,
                                                                            'tipo_existencia': tipo_existencia,
                                                                            'desc_abreviada' : unicode(fila[4].strip(), errors='ignore')})
            except:
                pass
        return HttpResponseRedirect(reverse('productos:productos'))  
    
class ConsultaStockProducto(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            producto = Producto.objects.get(codigo = codigo) 
            producto_json = {}                
            producto_json['stock'] = producto.stock
            data = simplejson.dumps(producto_json)
            return HttpResponse(data, 'application/json')      
        
class CrearGrupoProductos(CreateView):
    model = GrupoProductos
    template_name = 'productos/grupo_productos.html'
    form_class = GrupoProductosForm
    success_url = reverse_lazy('productos:grupos_productos')
    
    @method_decorator(permission_required('productos.add_grupoproductos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearGrupoProductos, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('productos:detalle_grupo_productos', args=[self.object.pk])

class CrearProducto(CreateView):
    model = Producto
    template_name = 'productos/producto.html'
    form_class = ProductoForm
    
    @method_decorator(permission_required('productos.add_producto',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearProducto, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('productos:detalle_producto', args=[self.object.pk])
    
class CrearUnidadMedida(CreateView):
    template_name = 'productos/unidad_medida.html'
    form_class = UnidadMedidaForm
        
    @method_decorator(permission_required('productos.add_unidadmedida',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearUnidadMedida, self).dispatch(*args, **kwargs)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.changed_by = self.request.user
        self.object.save()
        return super(CrearUnidadMedida, self).form_valid(form)
    
    def get_success_url(self):
        return reverse('productos:detalle_unidad_medida', args=[self.object.pk])
    
class CrearServicio(CreateView):
    template_name = 'productos/servicio.html'
    form_class = ServicioForm
        
    @method_decorator(permission_required('productos.add_producto',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearServicio, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('productos:detalle_servicio', args=[self.object.codigo])
    
"""class DownloadProductosReport(View):

    def get(self, request, *args, **kwargs):
        # debemos obtener nuestro objeto classroom haciendo la consulta a la base de datos
        report = ProductosReport()
        return report.render_to_response()"""
    
class DetalleProducto(DetailView):
    model = Producto
    template_name = 'productos/detalle_producto.html'
    
class DetalleGrupoProductos(DetailView):
    model = GrupoProductos
    template_name = 'productos/detalle_grupo_productos.html'
    
class DetalleUnidadMedida(DetailView):
    model = UnidadMedida
    template_name = 'productos/detalle_unidad_medida.html'

class DetalleServicio(DetailView):
    model = Producto
    template_name = 'productos/detalle_servicio.html'
    
class EliminarUnidadMedida(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            id = request.GET['id']
            unidad_medida = UnidadMedida.objects.get(pk=id)
            unidad_medida_json = {}
            unidad_medida_json['unidad'] = unidad_medida.unidad
            if len(unidad_medida.producto_set.all())>0:
                unidad_medida_json['productos'] = 'SI'
            else:
                unidad_medida_json['productos'] = 'NO'
                UnidadMedida.objects.filter(pk=id).update(estado = False)
            data = simplejson.dumps(unidad_medida_json)
            return HttpResponse(data, 'application/json')
        
class EliminarGrupoProductos(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            grupo_productos = GrupoProductos.objects.get(pk=codigo)
            grupo_productos_json = {}
            grupo_productos_json['codigo'] = grupo_productos.codigo
            grupo_productos_json['descripcion'] = grupo_productos.descripcion
            if len(grupo_productos.producto_set.all())>0:
                grupo_productos_json['productos'] = 'SI'
            else:
                grupo_productos_json['productos'] = 'NO'
                GrupoProductos.objects.filter(pk=codigo).update(estado = False)
            data = simplejson.dumps(grupo_productos_json)
            return HttpResponse(data, 'application/json')
        
class EliminarProducto(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            producto = Producto.objects.get(pk=codigo)
            producto_json = {}
            producto_json['codigo'] = producto.codigo
            producto_json['descripcion'] = producto.descripcion
            if len(producto.detallerequerimiento_set.all())>0:
                producto_json['relaciones'] = 'SI'
            elif len(producto.detallemovimiento_set.all())>0:
                producto_json['relaciones'] = 'SI'
            else:
                producto_json['relaciones'] = 'NO'
                Producto.objects.filter(pk=codigo).update(estado = False)
            data = simplejson.dumps(producto_json)
            return HttpResponse(data, 'application/json')
        
class EliminarServicio(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            servicio = Producto.objects.get(codigo=codigo)
            servicio_json = {}
            servicio_json['codigo'] = codigo
            if len(servicio.detalleordenservicio_set.all())>0:
                servicio_json['ordenes'] = 'SI'
            else:
                servicio_json['ordenes'] = 'NO'
                Producto.objects.filter(codigo=codigo).update(estado = False)                
            data = simplejson.dumps(servicio_json)
            return HttpResponse(data, 'application/json')
        
class ListadoUnidadesMedida(ListView):
    model = UnidadMedida
    template_name = 'productos/unidades_medida.html'
    context_object_name = 'unidades'
    queryset = UnidadMedida.objects.filter(estado=True).order_by('descripcion')
    
    @method_decorator(permission_required('productos.ver_tabla_unidades_medida',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoUnidadesMedida, self).dispatch(*args, **kwargs)
    
class ListadoServicios(ListView):
    model = Producto
    template_name = 'productos/servicios.html'
    context_object_name = 'servicios'
    queryset = Producto.objects.filter(estado=True,es_servicio=True).order_by('descripcion')
    
    @method_decorator(permission_required('productos.ver_tabla_productos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoServicios, self).dispatch(*args, **kwargs)
    
class ListadoGruposProductos(ListView):
    model = GrupoProductos
    template_name = 'productos/grupos_productos.html'
    context_object_name = 'grupos_productos'
    queryset = GrupoProductos.objects.filter(estado=True).order_by('codigo')
    
    @method_decorator(permission_required('productos.ver_tabla_grupos_productos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoGruposProductos, self).dispatch(*args, **kwargs)
    
class ListadoProductos(ListView):
    model = Producto
    template_name = 'productos/productos.html'
    context_object_name = 'productos'
    queryset = Producto.objects.filter(es_servicio=False,estado=True).order_by('codigo')
    
    @method_decorator(permission_required('productos.ver_tabla_productos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoProductos, self).dispatch(*args, **kwargs)
    
class ListadoProductosPorGrupo(ListView):
    model = Producto
    template_name = 'productos/productos.html'
    context_object_name = 'productos'    
    
    @method_decorator(permission_required('productos.ver_tabla_productos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoProductosPorGrupo, self).dispatch(*args, **kwargs)
    
    def get_queryset(self):
        grupo = GrupoProductos.objects.get(pk=self.kwargs['grupo'])
        queryset = grupo.producto_set.all()
        return queryset
    
class ModificarProducto(UpdateView):
    model = Producto
    template_name = 'productos/producto.html'
    form_class = ProductoForm    
        
    @method_decorator(permission_required('productos.change_producto',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarProducto, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('productos:detalle_producto', args=[self.object.pk])

class ModificarUnidadMedida(UpdateView):
    model = UnidadMedida
    template_name = 'productos/unidad_medida.html'
    form_class = UnidadMedidaForm
    
    @method_decorator(permission_required('productos.change_unidadmedida',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarUnidadMedida, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('productos:detalle_unidad_medida', args=[self.object.pk])
    
class ModificarGrupoProductos(UpdateView):
    model = GrupoProductos
    template_name = 'productos/grupo_productos.html'
    form_class = GrupoProductosForm
    success_url = reverse_lazy('productos:grupos_productos')
    
    @method_decorator(permission_required('productos.change_grupoproductos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarGrupoProductos, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('productos:detalle_grupo_productos', args=[self.object.pk])
    
class ModificarServicio(UpdateView):
    model = Producto
    template_name = 'productos/servicio.html'
    form_class = ServicioForm
    
    @method_decorator(permission_required('productos.change_servicio',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarServicio, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('productos:detalle_servicio', args=[self.object.pk])
    
class ReporteExcelProductos(TemplateView):
    
    def get(self, request, *args, **kwargs):
        productos = Producto.objects.filter(estado=True).order_by('codigo')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE PRODUCTOS'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'DESCRIPCION'
        ws['D3'] = 'DESCR_ABREV'
        ws['E3'] = 'GRUPO'        
        ws['F3'] = 'UNIDAD'
        ws['G3'] = 'MARCA'
        ws['H3'] = 'MODELO'
        ws['I3'] = 'PRECIO'
        ws['J3'] = 'CREADO'
        cont=4
        for producto in productos:
            ws.cell(row=cont,column=2).value = producto.codigo
            ws.cell(row=cont,column=3).value = producto.descripcion
            ws.cell(row=cont,column=4).value = producto.desc_abreviada
            ws.cell(row=cont,column=5).value = producto.grupo_productos.descripcion
            ws.cell(row=cont,column=6).value = producto.unidad_medida.descripcion
            ws.cell(row=cont,column=7).value = producto.marca
            ws.cell(row=cont,column=8).value = producto.modelo    
            ws.cell(row=cont,column=9).value = producto.precio
            ws.cell(row=cont,column=9).number_format = '#.00000'
            ws.cell(row=cont,column=10).value = producto.created  
            ws.cell(row=cont,column=10).number_format = 'dd/mm/yyyy hh:mm:ss'                  
            cont = cont + 1
        nombre_archivo ="ListadoProductos.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
class ReporteExcelGruposProductos(TemplateView):
    
    def get(self, request, *args, **kwargs):
        grupos_productos = GrupoProductos.objects.filter(estado=True).order_by('codigo')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE GRUPOS DE PRODUCTOS'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'DESCRIPCION'
        ws['D3'] = 'CTA_CONTABLE'
        ws['E3'] = 'CREADO'
        cont=4
        for grupo_productos in grupos_productos:
            ws.cell(row=cont,column=2).value = grupo_productos.codigo
            ws.cell(row=cont,column=3).value = grupo_productos.descripcion
            ws.cell(row=cont,column=4).value = grupo_productos.ctacontable.cuenta
            ws.cell(row=cont,column=5).value = grupo_productos.created
            ws.cell(row=cont,column=5).number_format = 'dd/mm/yyyy hh:mm:ss' 
            cont = cont + 1
        nombre_archivo ="ListadoGruposProductos.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
class ReporteExcelUnidadesMedida(TemplateView):
    
    def get(self, request, *args, **kwargs):
        unidades = UnidadMedida.objects.filter(estado=True).order_by('codigo')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE UNIDADES DE MEDIDA'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'UNIDAD'
        ws['C3'] = 'DESCRIPCIÓN'
        ws['D3'] = 'ESTADO'
        cont=4
        for unidad in unidades:
            ws.cell(row=cont,column=2).value = unidad.codigo
            ws.cell(row=cont,column=3).value = unidad.descripcion
            ws.cell(row=cont,column=4).value = unidad.estado
            cont = cont + 1
        nombre_archivo ="UnidadesMedida.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
class ReporteExcelServicios(TemplateView):
    
    def get(self, request, *args, **kwargs):
        servicios = Producto.objects.filter(es_servicio=True,estado=True).order_by('codigo')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE SERVICIOS'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'DESCRIPCION'
        ws['D3'] = 'ESTADO'
        cont=4
        for servicio in servicios:
            ws.cell(row=cont,column=2).value = servicio.codigo
            ws.cell(row=cont,column=3).value = servicio.descripcion
            ws.cell(row=cont,column=4).value = servicio.estado  
            cont = cont + 1
        nombre_archivo ="ListadoServicios.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response