# -*- coding: utf-8 -*-
from django.shortcuts import render, get_object_or_404
from django.views.generic.base import View, TemplateView
from django.views.generic.list import ListView
from django.views.generic.edit import FormView, UpdateView, CreateView
from django.core.urlresolvers import reverse_lazy, reverse
from django.http.response import HttpResponseRedirect
import json
from django.http import HttpResponse
import simplejson
from django.views.generic.detail import DetailView
from administracion.models import Oficina, NivelAprobacion
import locale
from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
import os
from django.db import transaction, IntegrityError
from django.db.models import Q
from django.contrib import messages
from requerimientos.models import AprobacionRequerimiento, Requerimiento, \
    DetalleRequerimiento
from requerimientos.forms import AprobacionRequerimientoForm, RequerimientoForm, DetalleRequerimientoFormSet
from compras.forms import DetalleCotizacionFormSet
from compras.models import Cotizacion
from productos.models import Producto
from requerimientos.mail import correo_creacion_requerimiento
from openpyxl import Workbook
from requerimientos.reports import ReporteRequerimiento
from datetime import date
from requerimientos.settings import CONFIGURACION, OFICINA_ADMINISTRACION, \
    LOGISTICA, PRESUPUESTO

locale.setlocale(locale.LC_ALL, "")


class Tablero(View):
    def get(self, request, *args, **kwargs):
        lista_notificaciones = []
        context = {'notificaciones': lista_notificaciones}
        return render(request, 'requerimientos/tablero_requerimientos.html', context)


class AprobarRequerimiento(UpdateView):
    model = AprobacionRequerimiento
    template_name = 'requerimientos/aprobar_requerimiento.html'
    form_class = AprobacionRequerimientoForm
    success_url = reverse_lazy('requerimientos:listado_aprobacion_requerimientos')

    @method_decorator(permission_required('requerimientos.change_aprobacionrequerimiento',
                                          reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        aprobacion_requerimiento = get_object_or_404(self.model, pk=kwargs['pk'])
        usuario = self.request.user
        if aprobacion_requerimiento.verificar_acceso_aprobacion(usuario):
            return super(AprobarRequerimiento, self).dispatch(*args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))

    def get_form_kwargs(self):
        kwargs = super(AprobarRequerimiento, self).get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.save()
        return HttpResponseRedirect(self.get_success_url())

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))


class CrearDetalleRequerimiento(FormView):
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            lista_detalles = []
            det = {}
            det['codigo'] = ''
            det['producto'] = ''
            det['unidad'] = ''
            det['cantidad'] = '0'
            det['uso'] = ''
            lista_detalles.append(det)
            formset = DetalleRequerimientoFormSet(initial=lista_detalles)
            lista_json = []
            for form in formset:
                detalle_json = {}
                detalle_json['codigo'] = str(form['codigo'])
                detalle_json['producto'] = str(form['producto'])
                detalle_json['unidad'] = str(form['unidad'])
                detalle_json['cantidad'] = str(form['cantidad'])
                detalle_json['uso'] = str(form['uso'])
                lista_json.append(detalle_json)
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')


class CrearRequerimiento(CreateView):
    template_name = 'requerimientos/requerimiento.html'
    form_class = RequerimientoForm
    model = Requerimiento

    def get_form_kwargs(self):
        kwargs = super(CrearRequerimiento, self).get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_initial(self):
        initial = super(CrearRequerimiento, self).get_initial()
        initial['annio'] = date.today().year
        initial['fecha'] = date.today().strftime('%d/%m/%Y')
        initial['mes'] = date.today().month
        return initial

    def get(self, request, *args, **kwargs):
        self.object = None
        oficinas = Oficina.objects.all()
        if not oficinas:
            return HttpResponseRedirect(reverse('administracion:crear_oficina'))
        try:
            trabajador = self.request.user.trabajador
        except:
            return HttpResponseRedirect(reverse('administracion:crear_trabajador'))
        if trabajador.firma == '':
            return HttpResponseRedirect(reverse('administracion:modificar_trabajador', args=[trabajador.pk]))
        puesto = trabajador.puesto
        if puesto is None:
            return HttpResponseRedirect(reverse('administracion:crear_puesto'))
        puesto_jefe = puesto.puesto_superior
        if puesto_jefe is None:
            return HttpResponseRedirect(reverse('administracion:crear_puesto'))
        niveles_aprobacion = NivelAprobacion.objects.all()
        if not niveles_aprobacion:
            return HttpResponseRedirect(reverse('administracion:crear_nivel_aprobacion'))
        if CONFIGURACION is not None:
            form_class = self.get_form_class()
            form = self.get_form(form_class)
            detalle_requerimiento_formset = DetalleRequerimientoFormSet()
            return self.render_to_response(self.get_context_data(form=form,
                                                                 detalle_requerimiento_formset=detalle_requerimiento_formset))
        else:
            return HttpResponseRedirect(reverse('contabilidad:configuracion'))

    def post(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_requerimiento_formset = DetalleRequerimientoFormSet(request.POST)
        if form.is_valid() and detalle_requerimiento_formset.is_valid():
            return self.form_valid(form, detalle_requerimiento_formset)
        else:
            return self.form_invalid(form, detalle_requerimiento_formset)

    def form_valid(self, form, detalle_requerimiento_formset):
        try:
            with transaction.atomic():
                self.object = form.save()
                detalles = []
                cont = 1
                for detalle_requerimiento_form in detalle_requerimiento_formset:
                    codigo = detalle_requerimiento_form.cleaned_data.get('codigo')
                    cantidad = detalle_requerimiento_form.cleaned_data.get('cantidad')
                    uso = detalle_requerimiento_form.cleaned_data.get('uso')
                    if codigo and cantidad:
                        producto = Producto.objects.get(codigo=codigo)
                        detalles.append(DetalleRequerimiento(requerimiento=self.object,
                                                             nro_detalle=cont,
                                                             producto=producto,
                                                             cantidad=cantidad,
                                                             uso=uso))
                        cont = cont + 1
                DetalleRequerimiento.objects.bulk_create(detalles)
                puesto_jefe = self.object.solicitante.puesto.puesto_superior  # Puesto.objects.get(oficina=self.object.oficina, es_jefatura=True, estado=True)
                jefe = puesto_jefe.trabajador
                destinatario = jefe.usuario.email
                if jefe.pk != self.object.solicitante.pk:
                    correo_creacion_requerimiento(destinatario, self.object)
                return HttpResponseRedirect(reverse('requerimientos:detalle_requerimiento', args=[self.object.codigo]))
        except IntegrityError:
            messages.error(self.request, 'Error guardando el requerimiento.')

    def form_invalid(self, form, detalle_requerimiento_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_requerimiento_formset=detalle_requerimiento_formset))


class DetalleOperacionRequerimiento(DetailView):
    model = Requerimiento
    slug_field = 'codigo'
    slug_url_kwarg = 'codigo'
    template_name = 'requerimientos/detalle_requerimiento.html'

    @method_decorator(
        permission_required('requerimientos.ver_detalle_requerimiento', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        requerimiento = self.get_object()
        if requerimiento.verificar_acceso(self.request.user, OFICINA_ADMINISTRACION, LOGISTICA, PRESUPUESTO):
            return super(DetalleOperacionRequerimiento, self).dispatch(*args, **kwargs)
        else:
            return HttpResponseRedirect(
                reverse('requerimientos:detalle_requerimiento', args=[requerimiento.siguiente()]))


class EliminarRequerimiento(TemplateView):
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            requerimiento = Requerimiento.objects.get(codigo=codigo)
            requerimiento_json = {}
            requerimiento_json['codigo'] = codigo
            cotizaciones = requerimiento.cotizacion_set.all()
            if len(cotizaciones) > 0:
                requerimiento_json['cotizaciones'] = 'SI'
            else:
                requerimiento_json['cotizaciones'] = 'NO'
                with transaction.atomic():
                    requerimiento.eliminar_requerimiento()
                    DetalleRequerimiento.objects.filter(requerimiento=requerimiento).delete()
            data = simplejson.dumps(requerimiento_json)
            return HttpResponse(data, 'application/json')


class ListadoAprobacionRequerimientos(ListView):
    model = AprobacionRequerimiento
    template_name = 'requerimientos/listado_aprobacion_requerimientos.html'
    context_object_name = 'aprobacion_requerimientos'

    @method_decorator(
        permission_required('requerimientos.ver_tabla_requerimientos', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoAprobacionRequerimientos, self).dispatch(*args, **kwargs)

    def get(self, request, *args, **kwargs):
        try:
            trabajador = self.request.user.trabajador
        except:
            return HttpResponseRedirect(reverse('administracion:crear_trabajador'))
        if trabajador.firma == '':
            return HttpResponseRedirect(reverse('administracion:modificar_trabajador', args=[trabajador.pk]))
        puesto = trabajador.puesto
        if puesto is None:
            return HttpResponseRedirect(reverse('administracion:crear_puesto'))
        if not puesto.es_jefatura:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))
        return super(ListadoAprobacionRequerimientos, self).get(request, *args, **kwargs)

    def get_queryset(self):
        return AprobacionRequerimiento.obtener_aprobaciones_pendientes(self.request.user)


class ListadoCotizacionesPorRequerimiento(ListView):
    model = Cotizacion
    template_name = 'compras/cotizaciones.html'
    context_object_name = 'cotizaciones'

    @method_decorator(permission_required('compras.ver_tabla_cotizaciones', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoCotizacionesPorRequerimiento, self).dispatch(*args, **kwargs)

    def get_queryset(self):
        requerimiento = Requerimiento.objects.get(pk=self.kwargs['requerimiento'])
        queryset = requerimiento.cotizacion_set.all()
        return queryset


class ListadoRequerimientos(ListView):
    model = Requerimiento
    template_name = 'requerimientos/listado_requerimientos.html'
    context_object_name = 'requerimientos'

    def get_queryset(self):
        usuario = self.request.user
        requerimientos_visibles = Requerimiento.obtener_requerimientos_visibles(self, usuario)
        return requerimientos_visibles

    @method_decorator(
        permission_required('requerimientos.ver_tabla_requerimientos', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoRequerimientos, self).dispatch(*args, **kwargs)


class ModificarRequerimiento(UpdateView):
    template_name = 'requerimientos/requerimiento.html'
    model = Requerimiento
    form_class = RequerimientoForm

    @method_decorator(
        permission_required('requerimientos.change_requerimiento', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        requerimiento = self.get_object()
        if (requerimiento.aprobacionrequerimiento.estado == AprobacionRequerimiento.NIVEL.USU or
                requerimiento.aprobacionrequerimiento.estado == AprobacionRequerimiento.NIVEL.JEF or
                self.request.user.is_superuser):
            return super(ModificarRequerimiento, self).dispatch(*args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('seguridad:permiso_denegado'))

    def get_form_kwargs(self):
        kwargs = super(ModificarRequerimiento, self).get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_initial(self):
        initial = super(ModificarRequerimiento, self).get_initial()
        initial['fecha'] = self.object.fecha.strftime('%d/%m/%Y')
        return initial

    def get_context_data(self, **kwargs):
        context = super(ModificarRequerimiento, self).get_context_data(**kwargs)
        try:
            context['archivo_informe'] = os.path.join('/siad', 'media', self.object.informe.url)
        except:
            pass
        return context

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalles = DetalleRequerimiento.objects.filter(requerimiento=self.object).order_by('nro_detalle')
        detalles_data = []
        for detalle in detalles:
            try:
                d = {'codigo': detalle.producto.codigo,
                     'producto': detalle.producto.descripcion,
                     'cantidad': detalle.cantidad,
                     'unidad': detalle.producto.unidad_medida.codigo,
                     'uso': detalle.uso}
            except:
                d = {'codigo': '',
                     'producto': detalle.otro,
                     'cantidad': detalle.cantidad,
                     'unidad': '',
                     'uso': detalle.uso}
            detalles_data.append(d)
        detalle_requerimiento_formset = DetalleRequerimientoFormSet(initial=detalles_data)
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_requerimiento_formset=detalle_requerimiento_formset))

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        detalle_requerimiento_formset = DetalleRequerimientoFormSet(request.POST)
        if form.is_valid() and detalle_requerimiento_formset.is_valid():
            return self.form_valid(form, detalle_requerimiento_formset)
        else:
            return self.form_invalid(form, detalle_requerimiento_formset)

    def form_valid(self, form, detalle_requerimiento_formset):
        try:
            with transaction.atomic():
                DetalleRequerimiento.objects.filter(requerimiento=self.object).delete()
                form.save()
                detalles = []
                cont = 1
                for detalle_requerimiento_form in detalle_requerimiento_formset:
                    codigo = detalle_requerimiento_form.cleaned_data.get('codigo')
                    cantidad = detalle_requerimiento_form.cleaned_data.get('cantidad')
                    uso = detalle_requerimiento_form.cleaned_data.get('uso')
                    if codigo and cantidad:
                        producto = Producto.objects.get(codigo=codigo)
                        detalles.append(
                            DetalleRequerimiento(requerimiento=self.object, nro_detalle=cont, producto=producto,
                                                 cantidad=cantidad, uso=uso))
                        cont = cont + 1
                    elif cantidad:
                        producto = detalle_requerimiento_form.cleaned_data.get('producto')
                        detalles.append(DetalleRequerimiento(requerimiento=self.object, nro_detalle=cont, otro=producto,
                                                             cantidad=cantidad, uso=uso))
                        cont = cont + 1
                DetalleRequerimiento.objects.bulk_create(detalles)
                return HttpResponseRedirect(reverse('requerimientos:detalle_requerimiento', args=[self.object.codigo]))
        except IntegrityError:
            messages.error(self.request, 'Error guardando el requerimiento.')

    def form_invalid(self, form, detalle_requerimiento_formset):
        return self.render_to_response(self.get_context_data(form=form,
                                                             detalle_requerimiento_formset=detalle_requerimiento_formset))


class ObtenerDetalleRequerimiento(TemplateView):
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            requerimiento = request.GET['requerimiento']
            tipo_busqueda = request.GET['tipo_busqueda']
            if tipo_busqueda == 'TODOS':
                detalles = DetalleRequerimiento.objects.filter(
                    Q(estado=DetalleRequerimiento.STATUS.PEND) | Q(estado=DetalleRequerimiento.STATUS.COTIZ),
                    requerimiento__codigo=requerimiento).order_by('nro_detalle')
            elif tipo_busqueda == 'PRODUCTOS':
                detalles = DetalleRequerimiento.objects.filter(Q(estado=DetalleRequerimiento.STATUS.PEND) |
                                                               Q(estado=DetalleRequerimiento.STATUS.COTIZ),
                                                               requerimiento__codigo=requerimiento,
                                                               producto__isnull=False).order_by('nro_detalle')
            lista_detalles = []
            for detalle in detalles:
                det = {}
                det['requerimiento'] = detalle.id
                try:
                    det['codigo'] = detalle.producto.codigo
                    det['nombre'] = detalle.producto.descripcion
                    det['unidad'] = detalle.producto.unidad_medida.codigo
                    # det['uso'] = detalle.uso
                    det['cantidad'] = str(detalle.cantidad - detalle.cantidad_atendida)
                    # det['precio'] = str(detalle.producto.precio)
                    # det['valor'] = str(detalle.producto.precio*(detalle.cantidad-detalle.cantidad_atendida))
                    lista_detalles.append(det)
                except:
                    pass
            formset = DetalleCotizacionFormSet(initial=lista_detalles)
            lista_json = []
            for form in formset:
                detalle_json = {}
                detalle_json['requerimiento'] = str(form['requerimiento'])
                detalle_json['codigo'] = str(form['codigo'])
                detalle_json['nombre'] = str(form['nombre'])
                detalle_json['unidad'] = str(form['unidad'])
                detalle_json['cantidad'] = str(form['cantidad'])
                lista_json.append(detalle_json)
            data = json.dumps(lista_json)
            return HttpResponse(data, 'application/json')


class TransferenciaRequerimiento(TemplateView):
    template_name = 'requerimientos/transferencia_requerimiento.html'

    def get_context_data(self, **kwargs):
        context = super(TransferenciaRequerimiento, self).get_context_data(**kwargs)
        # requerimientos = Requerimiento.objects.all()
        requerimientos = Requerimiento.obtener_requerimientos_listos_transferencia()
        context['requerimientos'] = requerimientos
        return context


class ReporteExcelRequerimientos(TemplateView):
    def get(self, request, *args, **kwargs):
        requerimientos = Requerimiento.objects.requerimientos_activos_por_usuario(request.user,
                                                                                  Requerimiento.STATUS.CANC)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE REQUERIMIENTOS'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'OFICINA'
        ws['D3'] = 'ESTADO APROBACION'
        ws['E3'] = 'ESTADO'
        ws['F3'] = 'FECHA'
        cont = 4
        for requerimiento in requerimientos:
            ws.cell(row=cont, column=2).value = requerimiento.codigo
            ws.cell(row=cont, column=3).value = requerimiento.oficina.nombre
            ws.cell(row=cont, column=4).value = requerimiento.get_estado_display()
            ws.cell(row=cont, column=5).value = requerimiento.aprobacionrequerimiento.get_estado_display()
            ws.cell(row=cont, column=6).value = requerimiento.created
            ws.cell(row=cont, column=6).number_format = 'dd/mm/yyyy hh:mm:ss'
            cont = cont + 1
        nombre_archivo = "ListadoRequerimientos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


class ReportePDFRequerimiento(View):
    def get(self, request, *args, **kwargs):
        codigo = kwargs['codigo']
        requerimiento = Requerimiento.objects.get(codigo=codigo)
        response = HttpResponse(content_type='application/pdf')
        reporte = ReporteRequerimiento('A4', requerimiento)
        pdf = reporte.imprimir()
        response.write(pdf)
        return response
