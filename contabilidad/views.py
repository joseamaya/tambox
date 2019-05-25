# -*- coding: utf-8 -*- 
from django.shortcuts import render
from django.views.generic.list import ListView
from contabilidad.models import CuentaContable, TipoDocumento, Impuesto,\
    Configuracion, FormaPago, Empresa, TipoExistencia, TipoCambio
from django.views.generic.base import View, TemplateView
from contabilidad.forms import TipoDocumentoForm,CuentaContableForm,\
    ImpuestoForm, ConfiguracionForm, FormaPagoForm, TipoCambioForm
from django.conf import settings
import csv
from django.http.response import HttpResponseRedirect
from django.core.urlresolvers import reverse, reverse_lazy
from django.views.generic.edit import FormView, UpdateView, CreateView,\
    BaseCreateView
import simplejson
from django.http import HttpResponse
from django.views.generic.detail import DetailView
from openpyxl import Workbook
from contabilidad.forms import UploadForm
import os
from django.contrib.auth.decorators import permission_required
from django.utils.decorators import method_decorator
import datetime

class Tablero(View):
    
    def get(self, request, *args, **kwargs):
        lista_notificaciones = []
        cant_cuentas_contables = CuentaContable.objects.count()
        tipo_documento, creado = TipoDocumento.objects.get_or_create(codigo_sunat = 'PEC',
                                                                     defaults = {'descripcion':'PECOSA',
                                                                                 'nombre': 'PECOSA'})
        if creado:
            lista_notificaciones.append("Se ha creado el tipo de documento PECOSA")
        if cant_cuentas_contables == 0:
            lista_notificaciones.append("No se ha creado ninguna cuenta contable")
        context = {'notificaciones':lista_notificaciones}
        return render(request, 'contabilidad/tablero_contabilidad.html', context)
    
class CargarCuentasContables(FormView):
    template_name = 'contabilidad/cargar_cuentas_contables.html'
    form_class = UploadForm
    
    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT,'archivos',str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        for fila in dataReader:
            CuentaContable.objects.get_or_create(cuenta=fila[0].strip(),
                                                 defaults={'descripcion': fila[1].strip()})
        return HttpResponseRedirect(reverse('contabilidad:cuentas_contables'))
    
class CargarTiposDocumentos(FormView):
    template_name = 'contabilidad/cargar_tipos_documentos.html'
    form_class = UploadForm
    
    def form_valid(self, form):
        data = form.cleaned_data
        docfile = data['archivo']            
        form.save()
        csv_filepathname = os.path.join(settings.MEDIA_ROOT,'archivos',str(docfile))
        dataReader = csv.reader(open(csv_filepathname), delimiter=',', quotechar='"')
        for fila in dataReader:
            TipoDocumento.objects.create(codigo_sunat = fila[0],
                                         nombre=fila[1],
                                         descripcion =fila[1])
        return HttpResponseRedirect(reverse('contabilidad:tipos_documentos'))   
    
class CrearFormaPago(CreateView):
    model = FormaPago
    template_name = 'contabilidad/forma_pago.html'
    form_class = FormaPagoForm
    
    @method_decorator(permission_required('compras.add_formapago',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearFormaPago, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('contabilidad:detalle_forma_pago', args=[self.object.pk])

class CrearTipoDocumento(CreateView):
    model = TipoDocumento
    template_name = 'contabilidad/tipo_documento.html'
    form_class = TipoDocumentoForm
    
    @method_decorator(permission_required('contabilidad.add_tipodocumento',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearTipoDocumento, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('contabilidad:detalle_tipo_documento', args=[self.object.pk])


class CrearTipoCambio(CreateView):
    model = TipoCambio
    template_name = 'contabilidad/tipo_cambio.html'
    form_class = TipoCambioForm

    @method_decorator(permission_required('contabilidad.add_tipocambio', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearTipoCambio, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('contabilidad:detalle_tipo_cambio', args=[self.object.pk])

class CrearCuentaContable(CreateView):
    model = CuentaContable
    template_name = 'contabilidad/cuenta_contable.html'
    form_class = CuentaContableForm
        
    @method_decorator(permission_required('contabilidad.add_cuenta_contable',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearCuentaContable, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('contabilidad:detalle_cuenta_contable', args=[self.object.pk])
    
class CrearImpuesto(CreateView):
    model = Impuesto
    template_name = 'contabilidad/impuesto.html'
    form_class = ImpuestoForm
        
    @method_decorator(permission_required('contabilidad.add_impuesto',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearImpuesto, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('contabilidad:detalle_impuesto', args=[self.object.pk])
    
class CrearConfiguracion(CreateView):
    model = Configuracion
    template_name = 'contabilidad/configuracion.html'
    form_class = ConfiguracionForm
        
    @method_decorator(permission_required('contabilidad.add_configuracion',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(CrearConfiguracion, self).dispatch(*args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        self.object = None
        configuracion = Configuracion.objects.first()
        if configuracion is None:
            return super(BaseCreateView, self).get(request, *args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('contabilidad:modificar_configuracion',args=[configuracion.pk]))
    
    def get_success_url(self):
        return reverse('contabilidad:modificar_configuracion', args=[self.object.pk])

class DetalleTipoCambio(DetailView):
    model = TipoCambio
    template_name = 'contabilidad/detalle_tipo_cambio.html'

class DetalleTipoDocumento(DetailView):
    model = TipoDocumento
    template_name = 'contabilidad/detalle_tipo_documento.html'
    
class DetalleCuentaContable(DetailView):
    model = CuentaContable
    template_name = 'contabilidad/detalle_cuenta_contable.html'
    
class DetalleImpuesto(DetailView):
    model = Impuesto
    template_name = 'contabilidad/detalle_impuesto.html'
    
class DetalleEmpresa(DetailView):
    model = Empresa
    template_name = 'contabilidad/detalle_empresa.html'
    
class DetalleFormaPago(DetailView):
    model = FormaPago
    template_name = 'contabilidad/detalle_forma_pago.html'
    
class EliminarFormaPago(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            codigo = request.GET['codigo']
            forma_pago = FormaPago.objects.get(pk=codigo)
            forma_pago_json = {}
            forma_pago_json['codigo'] = forma_pago.codigo
            forma_pago_json['descripcion'] = forma_pago.descripcion
            if len(forma_pago.ordencompra_set.all())>0:
                forma_pago_json['relaciones'] = 'SI'
            elif len(forma_pago.detalleordencompra_set.all())>0:
                forma_pago_json['relaciones'] = 'SI'
            elif len(forma_pago.detallemovimiento_set.all())>0:
                forma_pago_json['relaciones'] = 'SI'
            else:
                forma_pago_json['relaciones'] = 'NO'
                FormaPago.objects.filter(pk=codigo).update(estado = False)
            data = simplejson.dumps(forma_pago_json)
            return HttpResponse(data, 'application/json')
    
class EliminarTipoDocumento(TemplateView):
    
    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            id = request.GET['id']
            tipo_documento = TipoDocumento.objects.get(pk=id)
            tipo_documento_json = {}
            tipo_documento_json['codigo_sunat'] = tipo_documento.codigo_sunat
            tipo_documento_json['nombre'] = tipo_documento.nombre
            if len(tipo_documento.movimiento_set.all())>0:
                tipo_documento_json['relaciones'] = 'SI'
            else:
                tipo_documento_json['relaciones'] = 'NO'
                TipoDocumento.objects.filter(pk=id).update(estado = False)
            data = simplejson.dumps(tipo_documento_json)
            return HttpResponse(data, 'application/json') 
        
class ListadoTiposDocumentos(ListView):
    model = TipoDocumento
    template_name = 'contabilidad/tipos_documento.html'
    context_object_name = 'tipos'
    queryset = TipoDocumento.objects.filter(estado=True).order_by('nombre')
    
    @method_decorator(permission_required('contabilidad.ver_tabla_tipos_documentos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoTiposDocumentos, self).dispatch(*args, **kwargs)


class ListadoTiposCambio(ListView):
    model = TipoCambio
    template_name = 'contabilidad/tipos_cambio.html'
    context_object_name = 'tipos'

    @method_decorator(permission_required('contabilidad.ver_tabla_tipos_cambio', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoTiposCambio, self).dispatch(*args, **kwargs)
    
class ListadoCuentasContables(ListView):
    model = CuentaContable
    template_name = 'contabilidad/cuentas_contables.html'
    context_object_name = 'cuentas_contables'
    queryset = CuentaContable.objects.all().order_by('cuenta')
    
    @method_decorator(permission_required('contabilidad.ver_tabla_cuentas_contables',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoCuentasContables, self).dispatch(*args, **kwargs)
    
class ListadoFormasPago(ListView):
    model = FormaPago
    template_name = 'contabilidad/formas_pago.html'
    context_object_name = 'formas_pago'
    paginate_by = 10
    queryset = FormaPago.objects.order_by('codigo')
    
    @method_decorator(permission_required('compras.ver_tabla_formas_pago',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoFormasPago, self).dispatch(*args, **kwargs)
    
class ListadoImpuestos(ListView):
    model = Impuesto
    template_name = 'contabilidad/impuestos.html'
    context_object_name = 'impuestos'    
    
    @method_decorator(permission_required('contabilidad.ver_tabla_impuestos',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ListadoImpuestos, self).dispatch(*args, **kwargs)
    
class ModificarFormaPago(UpdateView):
    model = FormaPago
    template_name = 'contabilidad/forma_pago.html'
    form_class = FormaPagoForm
    
    @method_decorator(permission_required('compras.change_formapago',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarFormaPago, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('compras:detalle_forma_pago', args=[self.object.pk])


class ModificarTipoCambio(UpdateView):
    model = TipoCambio
    template_name = 'contabilidad/tipo_cambio.html'
    form_class = TipoCambioForm

    @method_decorator(permission_required('contabilidad.change_tipocambio', reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarTipoCambio, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('contabilidad:detalle_tipo_cambio', args=[self.object.pk])

class ModificarTipoDocumento(UpdateView):
    model = TipoDocumento
    template_name = 'contabilidad/tipo_documento.html'
    form_class = TipoDocumentoForm    

    @method_decorator(permission_required('contabilidad.change_tipodocumento',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarTipoDocumento, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('contabilidad:detalle_tipo_documento', args=[self.object.pk])
    
class ModificarCuentaContable(UpdateView):
    model = CuentaContable
    template_name = 'contabilidad/cuenta_contable.html'
    form_class = CuentaContableForm
    
    @method_decorator(permission_required('contabilidad.change_cuenta_contable',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarCuentaContable, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('contabilidad:detalle_cuenta_contable', args=[self.object.pk])

class ModificarConfiguracion(UpdateView):
    model = Configuracion
    template_name = 'contabilidad/configuracion.html'
    form_class = ConfiguracionForm
    
    @method_decorator(permission_required('contabilidad.change_configuracion',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarConfiguracion, self).dispatch(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('contabilidad:modificar_configuracion', args=[self.object.pk])
    
class ModificarImpuesto(UpdateView):
    model = Impuesto
    template_name = 'contabilidad/impuesto.html'
    form_class = ImpuestoForm
    
    @method_decorator(permission_required('contabilidad.change_impuesto',reverse_lazy('seguridad:permiso_denegado')))
    def dispatch(self, *args, **kwargs):
        return super(ModificarImpuesto, self).dispatch(*args, **kwargs)
    
    def get_initial(self):
        initial = super(ModificarImpuesto, self).get_initial()
        initial['fecha_inicio'] = self.object.fecha_inicio.strftime('%d/%m/%Y')
        if self.object.fecha_fin is not None:
            initial['fecha_fin'] = self.object.fecha_fin.strftime('%d/%m/%Y')        
        return initial
    
    def get_success_url(self):
        return reverse('contabilidad:detalle_impuesto', args=[self.object.pk])


class ObtenerTipoCambio(TemplateView):

    def get(self, request, *args, **kwargs):
        if request.is_ajax():
            fecha_get = request.GET['fecha']
            anio = int(fecha_get[6:])
            mes = int(fecha_get[3:5])
            dia = int(fecha_get[0:2])
            fecha = datetime.date(anio, mes, dia)
            try:
                tipo_cambio = TipoCambio.objects.get(fecha=fecha)
            except TipoCambio.DoesNotExist:
                tipo_cambio = {'fecha':fecha_get,'monto':0}
            data = simplejson.dumps(tipo_cambio)
            return HttpResponse(data, 'application/json')

class ReporteExcelCuentasContables(TemplateView):
    
    def get(self, request, *args, **kwargs):
        cuentas = CuentaContable.objects.all().order_by('cuenta')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE UNIDADES DE MEDIDA'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CUENTA'
        ws['C3'] = 'DESCRIPCIÓN'
        ws['D3'] = 'DEPRECIACION'
        cont=4
        for cuenta in cuentas:
            ws.cell(row=cont,column=2).value = cuenta.cuenta
            ws.cell(row=cont,column=3).value = cuenta.descripcion
            ws.cell(row=cont,column=4).value = cuenta.depreciacion
            cont = cont + 1
        nombre_archivo ="ListadoCuentasContables.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
class ReporteExcelFormasPago(TemplateView):
    
    def get(self, request, *args, **kwargs):
        formas_pago = FormaPago.objects.all().order_by('codigo')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE FORMAS DE PAGO'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'DESCRIPCIÓN'
        ws['D3'] = 'DIAS_CREDITO'
        cont=4
        for forma_pago in formas_pago:
            ws.cell(row=cont,column=2).value = forma_pago.codigo
            ws.cell(row=cont,column=3).value = forma_pago.descripcion
            ws.cell(row=cont,column=4).value = forma_pago.dias_credito
            cont = cont + 1
        nombre_archivo ="ListadoFormasPago.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
    
class ReporteExcelTiposDocumentos(TemplateView):
    
    def get(self, request, *args, **kwargs):
        tipos = TipoDocumento.objects.all().order_by('codigo_sunat')
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE TIPOS DE DOCUMENTOS'
        ws.merge_cells('B1:J1')
        ws['B3'] = 'CODIGO SUNAT'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'DESCRIPCIÓN'
        cont=4
        for tipo in tipos:
            ws.cell(row=cont,column=2).value = tipo.codigo_sunat
            ws.cell(row=cont,column=3).value = tipo.nombre
            ws.cell(row=cont,column=4).value = tipo.descripcion
            cont = cont + 1
        nombre_archivo ="ListadoTiposDocumentos.xlsx" 
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response